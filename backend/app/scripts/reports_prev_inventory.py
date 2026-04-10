# reports_prev_inventory.py
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, Tuple, Optional, Any, List
from openpyxl import load_workbook, Workbook

try:
    import xlrd  # .xls
except Exception:
    xlrd = None


def norm_code(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().replace("\u00A0", " ")
    s = re.sub(r"\s+", "", s).upper()
    if re.fullmatch(r"-?\d+\.0", s):  # 50205.0 -> 50205
        s = s[:-2]
    return s


def to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace(",", "")  # 1,234 -> 1234
    try:
        return float(s)
    except Exception:
        return None


def _detect_data_start_xlsx(ws, code_col_1based: int = 1, scan_rows: int = 30) -> int:
    # Хэрэв эхний мөр нь "Код"/"code" байвал дараагийн мөрөөс өгөгдөл эхэлнэ.
    for r in range(1, scan_rows + 1):
        v = ws.cell(r, code_col_1based).value
        if v is None:
            continue
        t = str(v).strip().lower()
        if t in ("код", "code"):
            return r + 1
        # Эхний мөр нь бодит код байж болох тул 1 гэж үзэж буцаана
        return 1
    return 1


def read_cols_any(path: str, col_map_0based: Dict[str, int]) -> Dict[str, Dict[str, Any]]:
    """
    col_map_0based: {'code':0, 'name':1, 'qty':8} гэх мэт
    Return: {CODE: {'name':..., 'qty':...}}
    """
    ext = Path(path).suffix.lower()

    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        start_row = _detect_data_start_xlsx(ws, code_col_1based=col_map_0based["code"] + 1)
        out: Dict[str, Dict[str, Any]] = {}
        for r in range(start_row, ws.max_row + 1):
            code = norm_code(ws.cell(r, col_map_0based["code"] + 1).value)
            if not code:
                continue
            name = ws.cell(r, col_map_0based["name"] + 1).value if "name" in col_map_0based else ""
            qtyv = ws.cell(r, col_map_0based["qty"] + 1).value if "qty" in col_map_0based else None
            out[code] = {"name": name or "", "qty": to_float(qtyv)}
        wb.close()
        return out

    if ext == ".xls":
        if xlrd is None:
            raise RuntimeError("xls унших xlrd суусангүй. requirements.txt дээр xlrd==2.0.1 нэмнэ үү.")
        book = xlrd.open_workbook(path)
        sh = book.sheet_by_index(0)

        # header detection
        start = 0
        if sh.nrows > 0 and sh.ncols > 0:
            t = str(sh.cell_value(0, col_map_0based["code"]) or "").strip().lower()
            if t in ("код", "code"):
                start = 1

        out: Dict[str, Dict[str, Any]] = {}
        for r in range(start, sh.nrows):
            code = norm_code(sh.cell_value(r, col_map_0based["code"]) if col_map_0based["code"] < sh.ncols else "")
            if not code:
                continue
            name = sh.cell_value(r, col_map_0based["name"]) if "name" in col_map_0based and col_map_0based["name"] < sh.ncols else ""
            qtyv = sh.cell_value(r, col_map_0based["qty"]) if "qty" in col_map_0based and col_map_0based["qty"] < sh.ncols else None
            out[code] = {"name": name or "", "qty": to_float(qtyv)}
        return out

    raise ValueError(f"Дэмжихгүй файл: {ext}")


def build_prev_inventory_check_report(after_path: str, counted_path: str, out_xlsx_path: str) -> None:
    # 1) Унших
    # After-adjustment: A=code(0), B=name(1), I=qty(8)
    after = read_cols_any(after_path, {"code": 0, "name": 1, "qty": 8})
    # Counted: A=code(0), B=name(1), D=qty(3)
    counted = read_cols_any(counted_path, {"code": 0, "name": 1, "qty": 3})

    # Fail-fast: хоосон уншилт бол хоосон файл экспортлохгүй
    if len(after) == 0 and len(counted) == 0:
        raise ValueError("Хоёр файл хоёулаа хоосон уншигдлаа. (Sheet/багана буруу эсэхийг шалгана уу.)")

    # 2) Зөрүү 1: нэг талд байгаа мөрүүд
    all_codes = sorted(set(after.keys()) | set(counted.keys()))
    missing_rows: List[List[Any]] = []
    mismatch_rows: List[List[Any]] = []

    for code in all_codes:
        a = after.get(code)
        c = counted.get(code)

        a_qty = a.get("qty") if a else None
        c_qty = c.get("qty") if c else None
        name = (c.get("name") if c and c.get("name") else (a.get("name") if a else "")) if True else ""

        in_after = "Тийм" if a else "Үгүй"
        in_counted = "Тийм" if c else "Үгүй"

        if (a is None) or (c is None):
            note = (
                "Тооллогын тайланд байна, тохируулгын дараах тайланд алга"
                if (c is not None and a is None)
                else "Тохируулгын дараах тайланд байна, тооллогын тайланд алга"
            )
            missing_rows.append([code, name, in_after, a_qty, in_counted, c_qty, note])
            continue

        # 3) Зөрүү 2: тоо зөрүүтэй
        # None-г 0 гэж үзэхгүй (жишиг: qty огт байхгүй бол алдаа гэж харуулах)
        if a_qty is None or c_qty is None:
            note = "Тоо хэмжээ уншигдаагүй (хоосон/формат буруу байж магадгүй)"
            mismatch_rows.append([code, name, c_qty, a_qty, None, note])
            continue

        if float(a_qty) != float(c_qty):
            diff = float(a_qty) - float(c_qty)
            note = f"Тооллого {c_qty}ш, тохируулгын дараах {a_qty}ш (зөрүү {diff:+g})"
            mismatch_rows.append([code, name, c_qty, a_qty, diff, note])

    # 4) Excel бичих
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Бүртгэл дутуу-илүү"
    ws1.append(["Код", "Нэр", "Тохируулгын дараах тайланд", "Үлдэгдэл (I)", "Тооллогын тайланд", "Тоолсон (D)", "Тайлбар"])
    for r in missing_rows:
        ws1.append(r)

    ws2 = wb.create_sheet("Тоо зөрүүтэй")
    ws2.append(["Код", "Нэр", "Тоолсон (D)", "Тохируулгын дараах (I)", "Зөрүү (I-D)", "Тайлбар"])
    for r in mismatch_rows:
        ws2.append(r)

    # simple formatting
    for ws in (ws1, ws2):
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 18 if col != 7 and col != 6 else 22
        if ws.max_column >= 7:
            ws.column_dimensions["G"].width = 60
        if ws.max_column >= 6:
            ws.column_dimensions["F"].width = 18

    wb.save(out_xlsx_path)
