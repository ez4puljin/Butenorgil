# warehouse_report_picker.py
# Windows: ажиллуулахад файл сонгох цонх нээгдэнэ.
# pip install pandas openpyxl
# (.xls унших бол) pip install xlrd==2.0.1

import os
import re
import math
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None
    filedialog = None
    messagebox = None



def to_float(x):
    try:
        if x is None:
            return float("nan")
        if isinstance(x, str):
            s = x.replace(",", "").strip()
            if s == "" or s.lower() in ("null", "none"):
                return float("nan")
            return float(s)
        if isinstance(x, (int, float)):
            return float(x)
        return float("nan")
    except Exception:
        return float("nan")


def parse_int_code(x):
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        if s == "":
            return None
        # "50205.0" гэх мэтийг int болгох
        if re.fullmatch(r"\d+(\.0+)?", s):
            return int(float(s))
        if s.isdigit():
            return int(s)
        return None
    if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x)):
        return int(float(x))
    return None


def is_numeric(x):
    try:
        if x is None:
            return False
        if isinstance(x, str):
            s = x.strip()
            if s == "":
                return False
            float(s.replace(",", ""))
            return True
        if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x)):
            return True
        return False
    except Exception:
        return False


def detect_warehouse_header_row(row):
    # A: 1..99, B: агуулахын нэр (string), C: тоон утгатай байх
    a = parse_int_code(row[0]) if len(row) > 0 else None
    b = row[1] if len(row) > 1 else None
    c = row[2] if len(row) > 2 else None
    return (
        a is not None
        and 0 < a < 100
        and isinstance(b, str) and b.strip() != ""
        and is_numeric(c)
    )


def detect_item_row(row):
    # A: барааны код 100000+, B: нэр, мөн C..K дотор дор хаяж 1 утга байна
    a = parse_int_code(row[0]) if len(row) > 0 else None
    b = row[1] if len(row) > 1 else None
    if a is None or a < 100000:
        return False
    if not isinstance(b, str) or b.strip() == "":
        return False

    has_any = False
    for j in range(2, min(len(row), 11)):  # C..K
        v = row[j]
        if v is None:
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        has_any = True
        break
    return has_any


def read_excel_any(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        # Excel 97-2003: xlrd шаардлагатай
        return pd.read_excel(path, header=None, engine="xlrd")
    return pd.read_excel(path, header=None)


def add_sheet(out_wb, name, df):
    ws = out_wb.create_sheet(title=name[:31])
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(r_idx, c_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[:2000]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def _parse_records(input_path: str) -> "pd.DataFrame":
    """Read the ERP stock file and return a raw records DataFrame."""
    raw = read_excel_any(input_path)
    records = []
    current_wh = None

    for idx in range(len(raw)):
        row = raw.iloc[idx].tolist()
        if detect_warehouse_header_row(row):
            current_wh = str(row[1]).strip()
            continue
        if current_wh is None:
            continue
        if not detect_item_row(row):
            continue
        code = parse_int_code(row[0])
        name = str(row[1]).strip()
        qty = to_float(row[8]) if len(row) > 8 else float("nan")
        is_red = (not math.isnan(qty)) and qty < 0
        nonzero = (not math.isnan(qty)) and abs(qty) > 0
        records.append({
            "Warehouse": current_wh,
            "Code": code,
            "Name": name,
            "FinalQty_I": None if math.isnan(qty) else qty,
            "IsRed": is_red,
            "NonZero": nonzero,
        })

    df = pd.DataFrame(records)
    if df.empty:
        raise RuntimeError("Барааны мөрүүд олдсонгүй.")
    return df


def get_stats(input_path: str) -> dict:
    """Return dashboard-ready summary stats without writing any file."""
    df = _parse_records(input_path)

    # Per-warehouse summary
    wh_summary = (
        df.groupby("Warehouse")
        .agg(
            items=("Code", "nunique"),
            red_items=("IsRed", "sum"),
            total_qty=(
                "FinalQty_I",
                lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum(),
            ),
        )
        .reset_index()
        .sort_values("Warehouse")
    )

    # Multi-location count
    nonzero_df = df[df["NonZero"]].copy()
    wh_counts = nonzero_df.groupby("Code")["Warehouse"].nunique()
    multi_location_count = int((wh_counts > 1).sum())

    warehouses = [
        {
            "name": row["Warehouse"],
            "items": int(row["items"]),
            "red_items": int(row["red_items"]),
            "total_qty": round(float(row["total_qty"]), 2),
        }
        for _, row in wh_summary.iterrows()
    ]

    return {
        "warehouse_count": len(warehouses),
        "total_items": int(df["Code"].nunique()),
        "total_red_items": int(df["IsRed"].sum()),
        "multi_location_count": multi_location_count,
        "warehouses": warehouses,
    }


def build_report(input_path, output_path):
    df = _parse_records(input_path)
    df["Row"] = 0  # Row info not tracked in _parse_records; keep column for compat

    # Warehouse_Summary: нийт бараа, улайсан тоо, эцсийн тоо
    wh_summary = df.groupby("Warehouse").agg(
        Items=("Code", "nunique"),
        RedItems=("IsRed", "sum"),
        TotalFinalQty=("FinalQty_I", lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum())
    ).reset_index().sort_values("Warehouse")

    # RedItems: зөвхөн I<0
    red_df = df[df["IsRed"]].copy()
    red_df["Reason"] = "FinalQty_I<0"
    red_df = red_df[["Warehouse", "Code", "Name", "FinalQty_I", "Reason", "Row"]].sort_values(["Warehouse", "Code"])

    # MultiLocation: давхар байршил дээр үлдэгдэлтэй бараа, агуулах бүрийн үлдэгдлийг баганаар
    nonzero_df = df[df["NonZero"]].copy()
    wh_counts = nonzero_df.groupby("Code")["Warehouse"].nunique().reset_index(name="WarehouseCount")
    multi_codes = set(wh_counts[wh_counts["WarehouseCount"] > 1]["Code"].tolist())

    multi_base = nonzero_df[nonzero_df["Code"].isin(multi_codes)].copy()
    pivot = multi_base.pivot_table(
        index=["Code", "Name"],
        columns="Warehouse",
        values="FinalQty_I",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    counts_map = wh_counts.set_index("Code")["WarehouseCount"].to_dict()
    pivot["WarehouseCount"] = pivot["Code"].map(counts_map).fillna(0).astype(int)

    warehouse_cols = [c for c in pivot.columns if c not in ("Code", "Name", "WarehouseCount")]
    pivot["TotalQty"] = pivot[warehouse_cols].sum(axis=1)

    multi_report = pivot[["Code", "Name", "WarehouseCount", "TotalQty"] + warehouse_cols].sort_values(
        ["WarehouseCount", "Code"], ascending=[False, True]
    )

    # Export
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    add_sheet(out_wb, "Warehouse_Summary", wh_summary)
    add_sheet(out_wb, "RedItems", red_df)
    add_sheet(out_wb, "MultiLocation", multi_report)

    out_wb.save(output_path)


def main():
    root = tk.Tk()
    root.withdraw()

    input_path = filedialog.askopenfilename(
        title="ERP татсан үлдэгдлийн файлаа сонго",
        filetypes=[
            ("Excel files", "*.xls *.xlsx"),
            ("Excel 97-2003", "*.xls"),
            ("Excel Workbook", "*.xlsx"),
            ("All files", "*.*"),
        ],
    )

    if not input_path:
        return

    out_dir = os.path.dirname(os.path.abspath(input_path))
    date_str = datetime.now().strftime("%Y_%m_%d")
    output_name = f"{date_str} улайлт тайлан.xlsx"
    output_path = os.path.join(out_dir, output_name)

    try:
        build_report(input_path, output_path)
        messagebox.showinfo("Амжилттай", f"Тайлан export хийгдлээ:\n{output_path}")
    except Exception as e:
        messagebox.showerror("Алдаа", str(e))


if __name__ == "__main__":
    main()
