from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def make_consolidated_excel(rows: list[dict], brand_final_weight: dict[str, float], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    headers = ["WarehouseTag", "Brand", "ItemCode", "Name", "Qty(Box)", "Qty(Pcs)", "UnitWeight", "AutoWeight"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["warehouse_tag_id"], r["brand"], r["item_code"], r["name"],
            r["order_qty_box"], r["order_qty_pcs"], r["unit_weight"], r["computed_weight"]
        ])

    # summary sheet
    ws2 = wb.create_sheet("Brand_Summary")
    ws2.append(["Brand", "AutoSumWeight", "FinalWeight(Override)"])
    for b, auto_sum in sorted({k: v for k, v in r_agg(rows).items()}.items()):
        ws2.append([b, auto_sum, brand_final_weight.get(b, auto_sum)])

    # autosize
    for sheet in [ws, ws2]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path)

def r_agg(rows: list[dict]) -> dict[str, float]:
    agg: dict[str, float] = {}
    for r in rows:
        agg[r["brand"]] = agg.get(r["brand"], 0.0) + float(r["computed_weight"] or 0)
    return agg