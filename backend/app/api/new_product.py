"""
Шинэ бараа таниулах — AI ашиглан бараа бүртгэх
- POST /new-product/analyze   : зураг → AI дүн шинжилгээ
- POST /new-product/generate-excel : баталгаажсан бараанууд → Excel татах
"""
from __future__ import annotations

import base64
import difflib
import io
import json
import re
from pathlib import Path
from typing import List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel

from app.core.config import settings
from app.api.deps import get_current_user
from app.models.user import User

router = APIRouter(prefix="/new-product", tags=["new-product"])

MASTER_PATH = Path("app/data/outputs/master_latest.xlsx")

# ──────────────────────────────────────────────────────────────
# Category lookup — master_latest.xlsx-аас ангилал болон max код
# ──────────────────────────────────────────────────────────────
def _load_category_map() -> dict[str, int]:
    """Ангилал нэр → тухайн ангиллын max код (int)"""
    if not MASTER_PATH.exists():
        return {}
    try:
        df = pd.read_excel(MASTER_PATH, sheet_name="Нэгтгэл")
        # Баганы нэрийг жижиг үсэг рүү хөрвүүлнэ
        df.columns = [str(c).strip() for c in df.columns]
        cat_col = next((c for c in df.columns if "ангилал" in c.lower()), None)
        code_col = next((c for c in df.columns if c.lower() in ["код", "item_code", "code"]), None)
        if not cat_col or not code_col:
            return {}
        df = df[[cat_col, code_col]].dropna()
        # Кодыг тооруу хувиргана (хэрэв боломжтой бол)
        def _to_int(v):
            try:
                return int(str(v).split(".")[0])
            except Exception:
                return 0
        df["_code_int"] = df[code_col].apply(_to_int)
        result = df.groupby(cat_col)["_code_int"].max().to_dict()
        return {str(k): int(v) for k, v in result.items()}
    except Exception:
        return {}


def _next_code(suggested_category: str, cat_map: dict[str, int]) -> tuple[str, str]:
    """Fuzzy match → matched_category, next_code (string)"""
    if not cat_map:
        return suggested_category, ""
    keys = list(cat_map.keys())
    matches = difflib.get_close_matches(suggested_category, keys, n=1, cutoff=0.3)
    if matches:
        matched = matches[0]
        return matched, str(cat_map[matched] + 1)
    return suggested_category, ""


# ──────────────────────────────────────────────────────────────
# Background removal — rembg (optional)
# ──────────────────────────────────────────────────────────────
def _remove_bg(image_bytes: bytes) -> str:
    """Зургийн фонг устгаж base64 PNG буцаана. rembg байхгүй бол оригинал зурагтай ажиллана."""
    try:
        from rembg import remove  # type: ignore
        from PIL import Image  # type: ignore

        img = Image.open(io.BytesIO(image_bytes)).convert("RGBA")
        out = remove(img)
        buf = io.BytesIO()
        out.save(buf, format="PNG")
        buf.seek(0)
        return "data:image/png;base64," + base64.b64encode(buf.read()).decode()
    except ImportError:
        # rembg суугаагүй тохиолдолд оригинал зургийг буцаана
        encoded = base64.b64encode(image_bytes).decode()
        return "data:image/jpeg;base64," + encoded
    except Exception:
        encoded = base64.b64encode(image_bytes).decode()
        return "data:image/jpeg;base64," + encoded


# ──────────────────────────────────────────────────────────────
# Google Gemini Vision дүн шинжилгээ (үнэгүй)
# ──────────────────────────────────────────────────────────────
def _analyze_with_gemini(photos_b64: list[str], brand: str, categories: list[str]) -> dict:
    """Gemini 1.5 Flash Vision ашиглан барааны мэдээлэл гаргана. Үнэгүй."""
    api_key = settings.gemini_api_key
    if not api_key:
        raise HTTPException(500, "Gemini API key тохируулагдаагүй. backend/.env-д GEMINI_API_KEY=AIza... нэмнэ үү. (aistudio.google.com-д үнэгүй авна)")

    try:
        from google import genai as ggenai          # type: ignore
        from google.genai import types as gtypes    # type: ignore
        from PIL import Image as PILImage           # type: ignore
    except ImportError:
        raise HTTPException(500, "google-genai пакет суугаагүй. pip install google-genai")

    client = ggenai.Client(api_key=api_key)

    cat_list = ", ".join(categories[:60]) if categories else "мэдэгдэхгүй"
    prompt = f"""Та барааны зургуудыг шинжлэн, дараах мэдээллийг ЗӨВХӨН JSON форматаар буцаана:
{{
  "name": "барааны бүтэн нэр (монгол, орос эсвэл латин хэлээр байсан хамаагүй, хамгийн тодорхой нэрийг)",
  "weight_kg": 0.288,
  "barcode": "4006996010457",
  "suggested_category": "КОФЕ"
}}

Баримтлах дүрмүүд:
- name: Зурган дээрх үгсийг бичнэ. Хайрцагны тоо (24ш, 12рх гэх мэт) байвал нэрэнд нэмнэ.
- weight_kg: Зурган дээрх жинг кг-аар (гр байвал 1000-д хуваана). Олдохгүй бол 0.
- barcode: EAN-13 эсвэл EAN-8 дугаар. Олдохгүй бол хоосон мөр "".
- suggested_category: Доорх жагсаалтаас хамгийн тохирохыг сонгоно: {cat_list}
- Тайлбар нэмэхгүй, зөвхөн JSON буцаана.

Бренд: {brand}"""

    # base64 → PIL Image болгоно
    parts: list = [prompt]
    for b64 in photos_b64[:5]:
        data = b64.split(",", 1)[1] if "," in b64 else b64
        img_bytes = base64.b64decode(data)
        img = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        parts.append(img)

    try:
        response = client.models.generate_content(
            model="gemini-1.5-flash",
            contents=parts,
        )
        raw = response.text or "{}"
    except Exception as e:
        err = str(e)
        if "quota" in err.lower() or "429" in err or "RESOURCE_EXHAUSTED" in err:
            raise HTTPException(429, "Gemini API хязгаарт хүрлээ. Хэсэг хүлээгээд дахин оролдоно уу.")
        if "api_key" in err.lower() or "API_KEY" in err or "INVALID_ARGUMENT" in err:
            raise HTTPException(500, "Gemini API key буруу. backend/.env файлыг шалгана уу.")
        raise HTTPException(500, f"Gemini алдаа: {err[:200]}")

    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return {"name": "", "weight_kg": 0, "barcode": "", "suggested_category": ""}


# ──────────────────────────────────────────────────────────────
# Endpoint: POST /new-product/analyze
# ──────────────────────────────────────────────────────────────
@router.post("/analyze")
async def analyze_product(
    photos: List[UploadFile] = File(...),
    brand: str = Form(""),
    pack_ratio: int = Form(1),
    _u: User = Depends(get_current_user),
):
    if not photos:
        raise HTTPException(400, "Хамгийн багадаа 1 зураг оруулна уу")

    # 1. Зургуудыг унших
    photos_bytes = [await f.read() for f in photos]

    # 2. Эхний зургийн фонг устгах (rembg)
    processed_image_b64 = _remove_bg(photos_bytes[0])

    # 3. Зургуудыг base64 болгох
    photos_b64 = [
        "data:image/jpeg;base64," + base64.b64encode(b).decode()
        for b in photos_bytes
    ]

    # 4. Category map ачаалах
    cat_map = _load_category_map()
    categories = sorted(cat_map.keys())

    # 5. Gemini Vision шинжилгээ
    ai_result = _analyze_with_gemini(photos_b64, brand, categories)

    # 6. Ангилал болон дараагийн код тодорхойлох
    suggested_cat = ai_result.get("suggested_category", "")
    matched_category, next_code = _next_code(suggested_cat, cat_map)

    return {
        "processed_image_b64": processed_image_b64,
        "name": ai_result.get("name", ""),
        "barcode": ai_result.get("barcode", ""),
        "weight_kg": ai_result.get("weight_kg", 0),
        "category": matched_category,
        "suggested_code": next_code,
        "brand": brand,
        "pack_ratio": pack_ratio,
    }


# ──────────────────────────────────────────────────────────────
# Endpoint: POST /new-product/generate-excel
# ──────────────────────────────────────────────────────────────
class ProductConfirmed(BaseModel):
    item_code: str
    name: str
    category: str
    weight_kg: float
    pack_ratio: int
    brand: str
    barcode: str
    processed_image_b64: Optional[str] = None


class GenerateExcelIn(BaseModel):
    products: List[ProductConfirmed]


@router.post("/generate-excel")
def generate_excel(
    body: GenerateExcelIn,
    _u: User = Depends(get_current_user),
):
    wb = Workbook()

    # ── 1-р хуудас: Бараа мастер мэдээлэл ──
    ws1 = wb.active
    ws1.title = "Шинэ бараа"
    headers1 = ["Код", "Нэр", "Ангилал нэр", "Жин (кг)", "Хайрцагны тоо", "Брэнд нэр", "Баркод"]
    ws1.append(headers1)

    # Header style
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="0071E3")
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for p in body.products:
        ws1.append([
            p.item_code,
            p.name,
            p.category,
            p.weight_kg,
            p.pack_ratio,
            p.brand,
            p.barcode,
        ])

    # Column width
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 25
    ws1.column_dimensions["G"].width = 18

    # ── 2-р хуудас: ERP Орлого авах format ──
    ws2 = wb.create_sheet("ERP Орлого")
    erp_headers = [
        "Данс", "Бараа материал", "Ангилал", "Тоо хэмжээ",
        "Нэгж үнэ", "Нийт дүн", "Хувийн жин", "Баркод"
    ]
    ws2.append(erp_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = PatternFill(fill_type="solid", fgColor="16A34A")
        cell.alignment = Alignment(horizontal="center")

    for p in body.products:
        ws2.append([
            "150101",       # Данс — тогтмол
            p.item_code,    # Бараа материал код
            p.category,
            1,              # Тоо хэмжээ — хэрэглэгч өөрчлөнө
            0,              # Нэгж үнэ — хэрэглэгч оруулна
            0,              # Нийт дүн
            1,              # Хувийн жин — тогтмол
            p.barcode,
        ])

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date
    filename = f"new_products_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
