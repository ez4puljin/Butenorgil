"""
Тооцоо хаах — Хаанбанкны хуулга Excel файлаар импортлох,
Календар харагдац, гүйлгээ засах, тохиргоо.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
from typing import Optional
from pydantic import BaseModel
from datetime import datetime
from pathlib import Path
import pandas as pd
import io
import re
import zipfile
from openpyxl import Workbook

from app.api.deps import get_db, get_current_user
from app.models.bank_statement import BankStatement, BankTransaction, BankAccountConfig, SettlementConfig, CrossAccountPreset, FeeConfig
from app.models.user import User

router = APIRouter(prefix="/bank-statements", tags=["bank-statements"])

# ── Helpers ───────────────────────────────────────────────────────────────────

_FEE_KEYWORDS = ["хураамж", "commission", "fee", "үйлчилгээний төлбөр"]

# POS гүйлгээ — Хаанбанк-ын settlement description-д "SETTLEMENT" гэж байна
# Жишээ: "29/04/2026 SETTLEMENT - ORGIL BUUNII TUV"
_SETTLEMENT_RE = re.compile(r"\bSETTLEMENT\b", re.IGNORECASE)
# Огноо хайх: эхлээд YYYY-prefix, дараа нь DD-prefix (Монгол стандарт)
_DATE_YMD_RE = re.compile(r"\b(\d{4})[-./](\d{1,2})[-./](\d{1,2})\b")
_DATE_DMY_RE = re.compile(r"\b(\d{1,2})[-./](\d{1,2})[-./](\d{4})\b")


def _is_fee(desc: str) -> bool:
    d = (desc or "").lower()
    return any(k in d for k in _FEE_KEYWORDS)


def _is_pos_income(desc: str) -> bool:
    """bank_description дотор 'SETTLEMENT' гэсэн pattern байгаа эсэх."""
    return bool(desc) and bool(_SETTLEMENT_RE.search(desc))


def _settlement_description(cfg_text: str, bank_desc: str) -> str:
    """Settlement-ийн Гүйлгээний утга = config-ын тэкст + банкны утга.
    Config хоосон бол зөвхөн банкны утга. Хоёулаа хоосон бол хоосон."""
    bd  = (bank_desc or "").strip()
    cfg = (cfg_text or "").strip()
    if cfg and bd:
        return f"{cfg} {bd}"
    return cfg or bd


def _extract_date_from_desc(desc: str):
    """bank_description-аас огноог parse хийнэ. Олдохгүй бол None буцаана."""
    if not desc:
        return None
    from datetime import date as _dt
    # YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD
    m = _DATE_YMD_RE.search(desc)
    if m:
        try:
            return _dt(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    # DD/MM/YYYY (Монгол стандарт)
    m = _DATE_DMY_RE.search(desc)
    if m:
        try:
            return _dt(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def _parse_excel(content: bytes, filename: str) -> dict:
    """Хаанбанкны бодит хуулга Excel-ийн форматыг танина:

    Row 0: Хэвлэсэн огноо / Хэрэглэгч / Интервал гэх мэт нийтлэг
            мэдээлэл (col[6] нь интервал огноо, e.g. "2026/05/07").
    Row 1: Багана-ын гарчиг
        col[0] Гүйлгээний огноо   col[1] Салбар
        col[2] Эхний үлдэгдэл     col[3] Дебит гүйлгээ (- утгатай)
        col[4] Кредит гүйлгээ     col[5] Эцсийн үлдэгдэл
        col[6] Гүйлгээний утга    col[7] Харьцсан данс
    Row 2..n: Гүйлгээний мөрүүд.
    Last row: "Нийт дүн:" нийлбэр (хасна).
    """
    account_number, currency = "", "MNT"
    m = re.search(r"Statement_([A-Z]+)_(\d+)", filename)
    if m:
        currency = m.group(1)
        account_number = m.group(2)

    raw = pd.read_excel(io.BytesIO(content), header=None, sheet_name=0)

    # Row 0-аас Интервал огноо ялгаж авна. col[5] = "Интервал:", col[6] = "YYYY/MM/DD"
    # Зарим хуулга нэг өдрийнх (col[6] -д нэг л огноо), зарим нь "from to" range.
    date_from = date_to = None
    try:
        # Бүх row 0-н текстээс огноо хайна (Хэвлэсэн огноо + Интервал хоёулангаас)
        joined = " ".join(str(v) for v in raw.iloc[0].tolist() if pd.notna(v))
        parts = re.findall(r"\d{4}[/\-]\d{2}[/\-]\d{2}", joined)
        # Эхний нь Хэвлэсэн огноо (MM/DD/YYYY байж болно), сүүлийн 1-2 нь Интервал
        # → Интервал-ыг сонгохын тулд YYYY-prefix хоёрт нь анхаарна
        ymd_parts = [p for p in parts if re.match(r"^\d{4}", p)]
        if len(ymd_parts) >= 2:
            date_from = datetime.strptime(ymd_parts[-2], "%Y/%m/%d").date()
            date_to   = datetime.strptime(ymd_parts[-1], "%Y/%m/%d").date()
        elif len(ymd_parts) == 1:
            date_from = date_to = datetime.strptime(ymd_parts[-1], "%Y/%m/%d").date()
    except Exception:
        pass

    # Header нь row 1 — гарчиг тэмдэглэгээ. Data row 2-оос эхэлнэ.
    # `header=None`-аар уншиж шууд багана index-ээр хандана.
    df = raw

    transactions = []
    for i in range(2, len(df)):
        row = df.iloc[i]
        # Эхний баганд "Нийт дүн" эсвэл хоосон бол алгасна (нийлбэр мөр)
        first = row.iloc[0]
        if pd.isna(first):
            continue
        first_str = str(first).strip()
        if not first_str or first_str.startswith("Нийт"):
            continue

        # col[0] — Гүйлгээний огноо (datetime)
        txn_date = None
        try:
            txn_date = pd.to_datetime(first).to_pydatetime()
        except Exception:
            continue  # огноо парс хийгдэхгүй бол data row биш

        # col[3] — Дебит (Excel-д сөрөг утгатай → abs())
        raw_debit  = 0.0
        try:
            v = row.iloc[3] if df.shape[1] > 3 else 0
            if pd.notna(v): raw_debit = float(v)
        except Exception:
            pass
        debit = abs(raw_debit)

        # col[4] — Кредит
        credit = 0.0
        try:
            v = row.iloc[4] if df.shape[1] > 4 else 0
            if pd.notna(v): credit = float(v)
        except Exception:
            pass

        # col[6] — Гүйлгээний утга
        desc = ""
        if df.shape[1] > 6:
            v = row.iloc[6]
            if pd.notna(v):
                desc = str(v).strip()
                if desc.lower() == "nan":
                    desc = ""

        # col[7] — Харьцсан данс (зарим мөрд хоосон)
        cpart = ""
        if df.shape[1] > 7:
            v = row.iloc[7]
            if pd.notna(v):
                try:
                    # 5303363476.0 → "5303363476"
                    cpart = str(int(float(str(v))))
                except (ValueError, OverflowError):
                    cpart = str(v).strip()
                if cpart.lower() == "nan":
                    cpart = ""

        is_fee_row = _is_fee(desc)
        default_action = "close" if (credit > 0 and not is_fee_row) else ""
        transactions.append({
            "txn_date":         txn_date,
            "debit":            debit,
            "credit":           credit,
            "bank_description": desc,
            "bank_counterpart": cpart,
            "is_fee":           is_fee_row,
            "action":           default_action,
        })

    return {
        "account_number": account_number,
        "currency":       currency,
        "date_from":      date_from,
        "date_to":        date_to,
        "filename":       filename,
        "transactions":   transactions,
    }


def _ser_stmt(s: BankStatement, include_txns: bool = False, erp_map: Optional[dict] = None, db: Optional[Session] = None) -> dict:
    txns = s.transactions
    main_txns = [t for t in txns if not t.is_fee]

    # ERP код хайх — config-аас account_number-ээр
    if erp_map is None and db is not None:
        cfg = db.query(BankAccountConfig).filter(
            BankAccountConfig.account_number == s.account_number
        ).first()
        erp_code = (cfg.erp_account_code or "") if cfg else ""
        is_registered = bool(cfg)
    elif erp_map is not None:
        if s.account_number in erp_map:
            erp_code = erp_map[s.account_number]
            is_registered = True
        else:
            erp_code = ""
            is_registered = False
    else:
        erp_code = ""
        is_registered = False

    d = {
        "id":               s.id,
        "account_number":   s.account_number,
        "currency":         s.currency,
        "date_from":        s.date_from.isoformat()   if s.date_from   else None,
        "date_to":          s.date_to.isoformat()     if s.date_to     else None,
        "filename":         s.filename,
        "uploaded_at":      s.uploaded_at.isoformat() if s.uploaded_at else None,
        "txn_count":        len(main_txns),
        "fee_count":        len(txns) - len(main_txns),
        "total_credit":     sum(t.credit for t in main_txns),
        "total_debit":      sum(t.debit  for t in main_txns),
        "filled_count":     sum(1 for t in main_txns if t.partner_name or t.action),
        "erp_account_code": erp_code,
        "is_registered":    is_registered,
    }
    if include_txns:
        d["transactions"] = [_ser_txn(t) for t in txns]
    return d


def _build_erp_map(db: Session) -> dict:
    """account_number → erp_account_code; зөвхөн config-д бүртгэлтэй дансууд."""
    return {
        a.account_number: (getattr(a, "erp_account_code", "") or "")
        for a in db.query(BankAccountConfig).all()
        if a.account_number
    }


def _ser_txn(t: BankTransaction) -> dict:
    return {
        "id":                 t.id,
        "txn_date":           t.txn_date.isoformat() if t.txn_date else None,
        "debit":              t.debit,
        "credit":             t.credit,
        "bank_description":   t.bank_description,
        "bank_counterpart":   t.bank_counterpart,
        "is_fee":             t.is_fee,
        "partner_name":       t.partner_name,
        "partner_account":    t.partner_account,
        "custom_description": t.custom_description,
        "action":             t.action,
        "export_type":        getattr(t, "export_type", "") or "",
        "is_settlement":      _is_pos_income(t.bank_description or "") and t.credit > 0,
    }


def _ser_acct(a: BankAccountConfig) -> dict:
    return {
        "id":               a.id,
        "account_number":   a.account_number,
        "partner_name":     a.partner_name,
        "bank_name":        a.bank_name,
        "is_fee_default":   a.is_fee_default,
        "erp_account_code": getattr(a, "erp_account_code", "") or "",
        "note":             a.note,
        "sort_order":       a.sort_order,
    }


# ── Pydantic ──────────────────────────────────────────────────────────────────

class TxnUpdate(BaseModel):
    partner_name:       Optional[str] = None
    partner_account:    Optional[str] = None
    custom_description: Optional[str] = None
    action:             Optional[str] = None
    export_type:        Optional[str] = None   # "" | "kass" | "hariltsah"


class AccountIn(BaseModel):
    account_number:   str  = ""
    partner_name:     str  = ""
    bank_name:        str  = "Хаанбанк"
    is_fee_default:   bool = False
    erp_account_code: str  = ""     # Эрхэт систем дахь дансны код (жишээ: "110104")
    note:             str  = ""
    sort_order:       int  = 0


# ── Calendar ──────────────────────────────────────────────────────────────────

@router.get("/calendar")
def get_calendar(
    year:  int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Тухайн сард өдөр тус бүрт хэдэн хуулга оруулсныг буцаана.
    date_from байгаа бол тэрийг, байхгүй бол uploaded_at-ийг ашиглана."""
    day_expr = func.coalesce(
        BankStatement.date_from,
        func.date(BankStatement.uploaded_at),
    )
    rows = db.query(
        day_expr.label("day"),
        func.count(BankStatement.id).label("cnt"),
    ).filter(
        func.strftime("%Y", day_expr) == str(year),
        func.strftime("%m", day_expr) == f"{month:02d}",
    ).group_by(day_expr).all()
    return {r.day: r.cnt for r in rows}


@router.get("/by-date")
def get_by_date(
    date: str = Query(...),   # "YYYY-MM-DD"
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Тухайн өдрийн хуулгуудыг буцаана.
    date_from байгаа бол тэрийг, байхгүй бол uploaded_at-ийг ашиглана."""
    day_expr = func.coalesce(
        BankStatement.date_from,
        func.date(BankStatement.uploaded_at),
    )
    rows = db.query(BankStatement).filter(
        day_expr == date,
    ).order_by(BankStatement.uploaded_at).all()
    erp_map = _build_erp_map(db)
    return [_ser_stmt(s, erp_map=erp_map) for s in rows]


# ── One-time data fix: fix negative debits + .0 counterparts ─────────────────

@router.post("/fix-legacy-data")
def fix_legacy_data(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Хуучин өгөгдлийн дебит сөрөг байгааг abs() болгож,
    харьцсан данс '5893180078.0' хэлбэрийг '5893180078' болгоно.
    Мөн action хоосон кредит мөрүүдэд default 'close' тавина.
    SETTLEMENT (POS) мөрийн partner_account-ыг банкны ERP код руу шилжүүлнэ."""
    # account_number → bank_erp_code map
    erp_map = _build_erp_map(db)
    # statement_id → account_number map
    stmt_map = {s.id: s.account_number for s in db.query(BankStatement).all()}

    txns = db.query(BankTransaction).all()
    fixed = 0
    for t in txns:
        changed = False
        if t.debit < 0:
            t.debit = abs(t.debit)
            changed = True
        if t.bank_counterpart and t.bank_counterpart.endswith(".0"):
            try:
                t.bank_counterpart = str(int(float(t.bank_counterpart)))
                changed = True
            except Exception:
                pass
        # Кредит мөр шимтгэл биш ба action хоосон бол → "close"
        if t.credit > 0 and not t.is_fee and not (t.action or "").strip():
            t.action = "close"
            changed = True
        # SETTLEMENT мөрийн partner_account нь банкны ERP код байх ёстой
        if t.credit > 0 and not t.is_fee and _is_pos_income(t.bank_description or ""):
            acct_no = stmt_map.get(t.statement_id, "")
            bank_erp = erp_map.get(acct_no, "")
            if bank_erp and (t.partner_account or "").strip() != bank_erp:
                t.partner_account = bank_erp
                changed = True
            if not (t.partner_name or "").strip() or t.partner_name == "Хаалт хийх харилцагч":
                t.partner_name = "30000"
                changed = True
        if changed:
            fixed += 1
    db.commit()
    return {"fixed": fixed}


# ── Customer search — in-memory cache (хурдан хайлт) ─────────────────────────

_CUSTOMER_FILE  = Path("app/data/outputs/customer_info_last.xlsx")
_customers_rows: list[dict] = []   # санах ойд байх жагсаалт
_customers_mtime: float = 0.0     # сүүлийн ачааллын цаг


def _ensure_customers_loaded() -> list[dict]:
    """Файл өөрчлөгдсөн эсвэл санах ой хоосон бол дахин ачаална, үгүй бол cache буцаана."""
    global _customers_rows, _customers_mtime
    if not _CUSTOMER_FILE.exists():
        return []
    try:
        mtime = _CUSTOMER_FILE.stat().st_mtime
    except OSError:
        return []
    if _customers_rows and mtime <= _customers_mtime:
        return _customers_rows      # cache хэвээр байна

    # Excel-г нэг удаа уншаад санах ойд хадгална
    try:
        df = pd.read_excel(str(_CUSTOMER_FILE), header=0, dtype=str)
    except Exception:
        return []

    rows = []
    for _, row in df.iterrows():
        def clean(col: str) -> str:
            v = row.get(col)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            v = str(v).strip()
            return "" if v in ("nan", "NaN", "None") else v

        name = clean("Нэр")
        if not name:
            continue
        rows.append({
            "code":       clean("Код"),
            "name":       name,
            "group":      clean("Бүлэг нэр"),
            "phone":      clean("Утас"),
            "account":    clean("Банк дахь данс"),
            # хайхад ашиглах lowercase нэгтгэл
            "_search":    f"{name} {clean('Код')} {clean('Бүлэг нэр')} {clean('Утас')}".lower(),
        })

    _customers_rows  = rows
    _customers_mtime = mtime
    print(f"[customers] {len(rows)} харилцагч санах ойд ачааллаа")
    return rows


@router.get("/customers/search")
def search_customers(
    q: str = Query(""),
    _: User = Depends(get_current_user),
):
    """Санах ойноос нэр/код/бүлгээр хурдан хайна (Excel-г дахин уншихгүй)."""
    rows = _ensure_customers_loaded()
    q_low = q.strip().lower()

    results = []
    for r in rows:
        if not q_low or q_low in r["_search"]:
            results.append({k: v for k, v in r.items() if k != "_search"})
            if len(results) >= 40:
                break
    return results


# ── Global fee export ─────────────────────────────────────────────────────────

@router.get("/export/fees")
def export_fees(
    date_from:   Optional[str] = Query(None),       # "YYYY-MM-DD" хуулгын эхлэх огноо
    date_to:     Optional[str] = Query(None),       # "YYYY-MM-DD" хуулгын дуусах огноо
    fee_partner: str           = Query("30000"),    # Эрхэт дахь харилцагчийн код
    fee_account: str           = Query("703012"),   # Эрхэт дахь харьцсан дансны код
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Бүх хуулгын банкны шимтгэл гүйлгээг нэгтгэн
    Мөнгөн хөрөнгийн харилцахын гүйлгээ Excel-ийг буцаана.
    Дансны код нь BankAccountConfig.erp_account_code-оос авна.
    """
    from datetime import date as dt_date

    # ── date_from / date_to filter ─────────────────────────────────
    day_expr = func.coalesce(BankStatement.date_from, func.date(BankStatement.uploaded_at))

    q = (
        db.query(BankTransaction, BankStatement)
        .join(BankStatement, BankTransaction.statement_id == BankStatement.id)
        .filter(BankTransaction.is_fee == True)
    )
    if date_from:
        try:
            q = q.filter(day_expr >= date_from)
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(day_expr <= date_to)
        except Exception:
            pass
    q = q.order_by(day_expr, BankTransaction.txn_date)

    rows = q.all()

    # ── ERP код харагдацын толь: account_number → erp_account_code ──
    erp_map = {
        a.account_number: (getattr(a, "erp_account_code", "") or "")
        for a in db.query(BankAccountConfig).all()
    }

    # ── (огноо, ERP код) түлхүүрээр шимтгэлийг нэгтгэнэ ─────────────
    # key = (date, erp_code, account_number) → нийлбэр дүн
    aggregated: dict[tuple, float] = {}
    for txn, stmt in rows:
        txn_date = txn.txn_date.date() if txn.txn_date else (stmt.date_from or dt_date.today())
        erp_code = erp_map.get(stmt.account_number, "")
        amount   = txn.debit if txn.debit > 0 else txn.credit
        key = (txn_date, erp_code, stmt.account_number)
        aggregated[key] = aggregated.get(key, 0.0) + (amount or 0.0)

    # ── Excel үүсгэх ───────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Шимтгэл"
    ws.append(_KASS_HARILTSAH_HEADERS)

    # Огноо → ERP код-оор эрэмбэлж бичнэ
    written = 0
    for (txn_date, erp_code, _acct), total in sorted(
        aggregated.items(), key=lambda x: (x[0][0], x[0][1])
    ):
        if total <= 0:
            continue
        ws.append([
            txn_date,         # Огноо — date object
            "Банкны шимтгэл", # Гүйлгээний утга — fixed
            fee_partner,      # Харилцагч — "30000"
            fee_account,      # Харьцсан данс — "703012"
            "",               # Харьцсан ялгаатай харилцагч
            "", "", "", "",   # НӨАТ 4 col
            "", "", "", "",   # НХАТ 4 col
            2,                # Орлого бол 1 Зарлага бол 2 — 2 (зарлага)
            erp_code,         # Дансны код — "110104" гэх мэт
            "",               # Валютын дүн
            "",               # Авах / Зарах ханш
            "",               # Олз / Гарзын дансны код
            total,            # Дүн — нэгтгэсэн нийлбэр
        ])
        written += 1

    # Огноо баганыг Short Date (MM/DD/YYYY) форматтай болгоно
    if written > 0:
        _apply_date_format(ws, "A", 2, written + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    from_lbl = (date_from or "all").replace("-", "")
    to_lbl   = ("_" + date_to.replace("-", "")) if date_to else ""
    filename = f"Shimtgel_{from_lbl}{to_lbl}.xlsx"
    display  = f"Шимтгэл_{from_lbl}{to_lbl}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(display)}"},
    )


# ── Config: accounts ──────────────────────────────────────────────────────────

@router.get("/config/accounts")
def list_accounts(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.query(BankAccountConfig).order_by(
        BankAccountConfig.sort_order, BankAccountConfig.id
    ).all()
    return [_ser_acct(a) for a in rows]


@router.post("/config/accounts")
def create_account(
    body: AccountIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    # Шимтгэлийн данс нэг л байна
    if body.is_fee_default:
        db.query(BankAccountConfig).update({"is_fee_default": False})
    a = BankAccountConfig(**body.model_dump())
    db.add(a)
    db.commit()
    db.refresh(a)
    return _ser_acct(a)


@router.patch("/config/accounts/{acct_id}")
def update_account(
    acct_id: int,
    body: AccountIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    a = db.query(BankAccountConfig).filter(BankAccountConfig.id == acct_id).first()
    if not a:
        raise HTTPException(404, "Данс олдсонгүй")
    if body.is_fee_default:
        db.query(BankAccountConfig).filter(BankAccountConfig.id != acct_id).update({"is_fee_default": False})
    for k, v in body.model_dump().items():
        setattr(a, k, v)
    db.commit()
    return _ser_acct(a)


@router.delete("/config/accounts/{acct_id}")
def delete_account(
    acct_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    a = db.query(BankAccountConfig).filter(BankAccountConfig.id == acct_id).first()
    if not a:
        raise HTTPException(404, "Данс олдсонгүй")
    db.delete(a)
    db.commit()
    return {"ok": True}


# ── Cross-Account Presets (310101, 120101 гэх мэт) ──────────────────────────

class CrossAcctIn(BaseModel):
    code:       str = ""
    label:      str = ""
    sort_order: int = 0


def _ser_cross(c: CrossAccountPreset) -> dict:
    return {
        "id":         c.id,
        "code":       c.code or "",
        "label":      c.label or "",
        "sort_order": c.sort_order or 0,
    }


@router.get("/config/cross-accounts")
def list_cross_accounts(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.query(CrossAccountPreset).order_by(
        CrossAccountPreset.sort_order, CrossAccountPreset.id
    ).all()
    if not rows:
        # Анх удаа дуудахад default 2 утга seed хийнэ
        defaults = [("310101", "Авлагын данс"), ("120101", "Өглөгийн данс")]
        for i, (code, label) in enumerate(defaults):
            db.add(CrossAccountPreset(code=code, label=label, sort_order=i))
        db.commit()
        rows = db.query(CrossAccountPreset).order_by(
            CrossAccountPreset.sort_order, CrossAccountPreset.id
        ).all()
    return [_ser_cross(r) for r in rows]


@router.post("/config/cross-accounts")
def create_cross_account(
    body: CrossAcctIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rec = CrossAccountPreset(**body.model_dump())
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return _ser_cross(rec)


@router.patch("/config/cross-accounts/{cid}")
def update_cross_account(
    cid: int,
    body: CrossAcctIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rec = db.query(CrossAccountPreset).filter(CrossAccountPreset.id == cid).first()
    if not rec:
        raise HTTPException(404, "Олдсонгүй")
    for k, v in body.model_dump().items():
        setattr(rec, k, v)
    db.commit()
    return _ser_cross(rec)


@router.delete("/config/cross-accounts/{cid}")
def delete_cross_account(
    cid: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rec = db.query(CrossAccountPreset).filter(CrossAccountPreset.id == cid).first()
    if not rec:
        raise HTTPException(404, "Олдсонгүй")
    db.delete(rec)
    db.commit()
    return {"ok": True}


# ── Settlement Config (singleton) ────────────────────────────────────────────

def _get_settlement_config(db: Session) -> SettlementConfig:
    """SETTLEMENT-ийн тохиргоо (singleton). Байхгүй бол default-аар үүсгэнэ."""
    cfg = db.query(SettlementConfig).filter(SettlementConfig.id == 1).first()
    if not cfg:
        cfg = SettlementConfig(
            id=1, partner_name="30000", partner_account="",
            custom_description="", action="close", account_code="120105",
        )
        db.add(cfg)
        db.commit()
        db.refresh(cfg)
    return cfg


def _ser_settlement(cfg: SettlementConfig) -> dict:
    return {
        "partner_name":       cfg.partner_name or "",
        "partner_account":    cfg.partner_account or "",
        "custom_description": cfg.custom_description or "",
        "action":             cfg.action or "",
        "account_code":       cfg.account_code or "",
    }


class SettlementConfigIn(BaseModel):
    partner_name:       Optional[str] = None
    partner_account:    Optional[str] = None
    custom_description: Optional[str] = None
    action:             Optional[str] = None
    account_code:       Optional[str] = None


@router.get("/config/settlement")
def get_settlement_config(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return _ser_settlement(_get_settlement_config(db))


@router.patch("/config/settlement")
def update_settlement_config(
    body: SettlementConfigIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    cfg = _get_settlement_config(db)
    if body.partner_name       is not None: cfg.partner_name       = body.partner_name
    if body.partner_account    is not None: cfg.partner_account    = body.partner_account
    if body.custom_description is not None: cfg.custom_description = body.custom_description
    if body.action             is not None: cfg.action             = body.action
    if body.account_code       is not None: cfg.account_code       = body.account_code
    db.commit()
    db.refresh(cfg)
    return _ser_settlement(cfg)


# ── Fee Config (singleton) ───────────────────────────────────────────────────

def _get_fee_config(db: Session) -> FeeConfig:
    cfg = db.query(FeeConfig).filter(FeeConfig.id == 1).first()
    if not cfg:
        cfg = FeeConfig(
            id=1, partner_name="30000", partner_account="703012",
            custom_description="Банкны шимтгэл", action="close",
            export_type="hariltsah", account_code="",
        )
        db.add(cfg)
        db.commit()
        db.refresh(cfg)
    return cfg


def _ser_fee_cfg(cfg: FeeConfig) -> dict:
    return {
        "partner_name":       cfg.partner_name or "",
        "partner_account":    cfg.partner_account or "",
        "custom_description": cfg.custom_description or "",
        "action":             cfg.action or "",
        "export_type":        cfg.export_type or "hariltsah",
        "account_code":       cfg.account_code or "",
    }


class FeeConfigIn(BaseModel):
    partner_name:       Optional[str] = None
    partner_account:    Optional[str] = None
    custom_description: Optional[str] = None
    action:             Optional[str] = None
    export_type:        Optional[str] = None
    account_code:       Optional[str] = None


@router.get("/config/fee")
def get_fee_config(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return _ser_fee_cfg(_get_fee_config(db))


@router.patch("/config/fee")
def update_fee_config(
    body: FeeConfigIn,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    cfg = _get_fee_config(db)
    if body.partner_name       is not None: cfg.partner_name       = body.partner_name
    if body.partner_account    is not None: cfg.partner_account    = body.partner_account
    if body.custom_description is not None: cfg.custom_description = body.custom_description
    if body.action             is not None: cfg.action             = body.action
    if body.export_type        is not None: cfg.export_type        = body.export_type
    if body.account_code       is not None: cfg.account_code       = body.account_code
    db.commit()
    db.refresh(cfg)
    return _ser_fee_cfg(cfg)


@router.post("/config/fee/reapply")
def reapply_fee_config(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Одоо байгаа шимтгэл (is_fee=True) мөрүүдэд тохиргоог дахин хэрэглэнэ."""
    fc = _get_fee_config(db)
    fixed = 0
    for t in db.query(BankTransaction).filter(BankTransaction.is_fee == True).all():
        t.partner_name       = fc.partner_name or "30000"
        t.partner_account    = fc.partner_account or "703012"
        t.custom_description = fc.custom_description or "Банкны шимтгэл"
        t.action             = fc.action or "close"
        t.export_type        = fc.export_type or "hariltsah"
        fixed += 1
    db.commit()
    return {"ok": True, "fixed": fixed}


@router.post("/config/settlement/reapply")
def reapply_settlement_config(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Одоо байгаа SETTLEMENT мөрүүдэд тохиргоог дахин хэрэглэнэ."""
    sc = _get_settlement_config(db)
    erp_map = _build_erp_map(db)
    stmt_map = {s.id: s.account_number for s in db.query(BankStatement).all()}
    fixed = 0
    for t in db.query(BankTransaction).all():
        if t.credit > 0 and not t.is_fee and _is_pos_income(t.bank_description or ""):
            acct_no  = stmt_map.get(t.statement_id, "")
            bank_erp = erp_map.get(acct_no, "")
            t.partner_name    = sc.partner_name or "30000"
            t.partner_account = (sc.partner_account or "").strip() or bank_erp
            t.custom_description = _settlement_description(sc.custom_description, t.bank_description or "")
            t.action          = sc.action or "close"
            fixed += 1
    db.commit()
    return {"ok": True, "fixed": fixed}


# ── Statements CRUD ───────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_statement(
    file: UploadFile = File(...),
    selected_date: Optional[str] = Form(None),   # frontend-ийн сонгосон огноо "YYYY-MM-DD"
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content  = await file.read()
    filename = file.filename or "statement.xlsx"
    try:
        parsed = _parse_excel(content, filename)
    except Exception as e:
        raise HTTPException(400, f"Excel уншихад алдаа: {e}")

    # Сонгосон огноог date_from-д ашиглана (Excel-ийн header parse амжилтгүй болсон тохиолдолд ч зөв ажиллана)
    effective_date_from = parsed["date_from"]
    effective_date_to   = parsed["date_to"]
    if selected_date:
        from datetime import date as date_type
        try:
            parsed_sel = date_type.fromisoformat(selected_date)
            # Excel-ийн date_from байхгүй эсвэл огноо зөрүүтэй бол override хийнэ
            if effective_date_from is None:
                effective_date_from = parsed_sel
                effective_date_to   = parsed_sel

        except ValueError:
            pass

    stmt = BankStatement(
        account_number=parsed["account_number"],
        currency=parsed["currency"],
        date_from=effective_date_from,
        date_to=effective_date_to,
        filename=parsed["filename"],
        uploaded_by_id=current_user.id,
    )
    db.add(stmt)
    db.flush()

    # Тухайн дансны ERP код хайх (SETTLEMENT мөрд автоматаар бөглөх)
    cfg = db.query(BankAccountConfig).filter(
        BankAccountConfig.account_number == parsed["account_number"]
    ).first()
    bank_erp = (cfg.erp_account_code or "") if cfg else ""
    sc = _get_settlement_config(db)
    fc = _get_fee_config(db)

    for t in parsed["transactions"]:
        # 1) Шимтгэл мөр илрүүлбэл Fee Config-ийн утгуудаар бөглөнө
        if t.get("is_fee"):
            t["partner_name"]       = fc.partner_name or "30000"
            t["partner_account"]    = fc.partner_account or "703012"
            t["custom_description"] = fc.custom_description or "Банкны шимтгэл"
            t["action"]             = fc.action or "close"
            t["export_type"]        = fc.export_type or "hariltsah"
        # 2) SETTLEMENT (POS) мөр илрүүлбэл Settlement Config-оор автоматаар бөглөнө
        elif t.get("credit", 0) > 0 and _is_pos_income(t.get("bank_description", "")):
            t["partner_name"]    = sc.partner_name or "30000"
            t["partner_account"] = (sc.partner_account or "").strip() or bank_erp
            t["custom_description"] = _settlement_description(sc.custom_description, t.get("bank_description", ""))
            t["action"]          = sc.action or "close"
        db.add(BankTransaction(statement_id=stmt.id, **t))

    db.commit()
    db.refresh(stmt)
    return _ser_stmt(stmt, db=db)


@router.get("/")
def list_statements(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.query(BankStatement).order_by(BankStatement.uploaded_at.desc()).all()
    erp_map = _build_erp_map(db)
    return [_ser_stmt(s, erp_map=erp_map) for s in rows]


@router.get("/{stmt_id}")
def get_statement(
    stmt_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = db.query(BankStatement).filter(BankStatement.id == stmt_id).first()
    if not s:
        raise HTTPException(404, "Хуулга олдсонгүй")
    return _ser_stmt(s, include_txns=True, db=db)


@router.patch("/{stmt_id}/transactions/{txn_id}")
def update_transaction(
    stmt_id: int,
    txn_id: int,
    body: TxnUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    t = db.query(BankTransaction).filter(
        BankTransaction.id == txn_id,
        BankTransaction.statement_id == stmt_id,
    ).first()
    if not t:
        raise HTTPException(404, "Гүйлгээ олдсонгүй")

    if body.partner_name       is not None: t.partner_name       = body.partner_name
    if body.partner_account    is not None: t.partner_account    = body.partner_account
    if body.custom_description is not None: t.custom_description = body.custom_description
    if body.action             is not None: t.action             = body.action
    if body.export_type        is not None: t.export_type        = body.export_type

    db.commit()
    return _ser_txn(t)


# ── Эрхэт Excel export helpers ───────────────────────────────────────────────

_AVLAGA_HEADERS = [
    "Огноо", "Гүйлгээний утга", "Дуусах огноо", "Харилцагч",
    "Харьцсан данс", "Харьцсан ялгаатай харилцагч",
    "НӨАТ тэй эсэх", "НӨАТ автоматаар бодох эсэх",
    "НӨАТ ийн үзүүлэлтийн код", "НӨАТ ийн дүн",
    "НХАТ тэй эсэх", "НХАТ автоматаар бодох эсэх",
    "НХАТ ийн дүн", "НХАТ ийн үзүүлэлтийн код",
    "Нээлт(1) / Хаалт(2)", "Дансны код", "Валют",
    "Авах / Зарах ханш", "Олз / Гарзын дансны код", "Дүн",
]

_KASS_HARILTSAH_HEADERS = [
    "Огноо", "Гүйлгээний утга", "Харилцагч",
    "Харьцсан данс", "Харьцсан ялгаатай харилцагч",
    "НӨАТ тэй эсэх", "НӨАТ автоматаар бодох эсэх",
    "НӨАТ ийн үзүүлэлтийн код", "НӨАТ ийн дүн",
    "НХАТ тэй эсэх", "НХАТ автоматаар бодох эсэх",
    "НХАТ ийн дүн", "НХАТ ийн үзүүлэлтийн код",
    "Орлого бол 1 Зарлага бол 2", "Дансны код",
    "Валютын дүн", "Авах / Зарах ханш",
    "Олз / Гарзын дансны код", "Дүн",
]


# Excel-ийн "Short Date" формат — 04/30/2026
_DATE_FMT = "MM/DD/YYYY"


def _apply_date_format(ws, col_letter: str, start_row: int, end_row: int) -> None:
    """Тухайн баганы заасан мөрүүдэд Short Date формат тавина."""
    for r in range(start_row, end_row + 1):
        ws[f"{col_letter}{r}"].number_format = _DATE_FMT


def _build_avlaga_excel(txns: list, eff_date, bank_erp_code: str = "") -> bytes:
    """Авлага өглөгийн гүйлгээ Excel (зөвхөн кредит гүйлгээ).
    SETTLEMENT pattern илрүүлбэл автомат бөглөнө:
      • Огноо       — bank_description-аас parse
      • Утга        — "Пос орлого " + bank_description
      • Харилцагч   — "30000"
      • Харьцсан    — банкны ERP код (bank_erp_code)
      • Дансны код  — "120105"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Гүйлгээ"
    ws.append(_AVLAGA_HEADERS)
    for t in txns:
        bd = t.bank_description or ""
        is_pos = _is_pos_income(bd)

        # Кредит мөрийн анхдагч action = "Хаах" (хэрэв хэрэглэгч өөрөөр сонгоогүй бол)
        # 2026-05: Хэрэглэгчийн хүсэлтээр O баганы утга шинэчлэгдэв
        #   close        → 1   (Хаах — нэг үйлдэл)
        #   create       → 1   (Үүсгэх — нэг үйлдэл)
        #   close_create → 2   (Хаах Үүсгэх — хосолсон үйлдэл)
        #   бусад/хоосон → ""
        eff_action = (t.action or "").strip() or "close"
        if eff_action == "close_create":
            action_val: object = 2
        elif eff_action in ("close", "create"):
            action_val = 1
        else:
            action_val = ""

        if is_pos:
            pos_date  = _extract_date_from_desc(bd)
            row_date  = pos_date or eff_date
            partner   = (t.partner_name or "").strip() or "30000"
            # SETTLEMENT POS — "Харьцсан данс" нь үргэлж банкны ERP код байна
            # (хэрэглэгчийн partner_account-ыг override хийнэ)
            cross     = bank_erp_code or (t.partner_account or "").strip()
            # custom_description нь parse үед концат хийгдсэн байна, эс бөгөөс bd
            desc_text = (t.custom_description or "").strip() or bd
        else:
            row_date  = eff_date
            partner   = t.partner_name or ""
            cross     = t.partner_account or ""
            desc_text = t.custom_description or bd
        # Авлага бүх кредит мөрийн "Дансны код" = 120105 (анхдагч авлагын данс)
        main_acct = "120105"

        ws.append([
            row_date,        # Огноо — date object (POS бол bank_description-ийн огноо)
            desc_text,       # Гүйлгээний утга
            "",              # Дуусах огноо
            partner,         # Харилцагч (POS → "30000")
            cross,           # Харьцсан данс (POS → bank ERP код)
            "",              # Харьцсан ялгаатай харилцагч (хоосон)
            "", "", "", "",  # НӨАТ 4 col
            "", "", "", "",  # НХАТ 4 col
            action_val,      # Нээлт(1) / Хаалт(2)
            main_acct,       # Дансны код (POS → "120105")
            "",              # Валют
            "",              # Авах / Зарах ханш
            "",              # Олз / Гарзын дансны код
            t.credit,        # Дүн
        ])
    if txns:
        _apply_date_format(ws, "A", 2, len(txns) + 1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_kass_hariltsah_excel(txns: list, eff_date, bank_erp_code: str = "") -> bytes:
    """Кассын / Харилцахын гүйлгээ Excel (дебит гүйлгээ)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Гүйлгээ"
    ws.append(_KASS_HARILTSAH_HEADERS)
    for t in txns:
        desc = t.custom_description or t.bank_description or ""
        ws.append([
            eff_date,                # Огноо — date object
            desc,                    # Гүйлгээний утга
            t.partner_name or "",    # Харилцагч
            t.partner_account or "", # Харьцсан данс
            "",                      # Харьцсан ялгаатай харилцагч (хоосон)
            "", "", "", "",          # НӨАТ 4 col
            "", "", "", "",          # НХАТ 4 col
            2,                       # Орлого бол 1 Зарлага бол 2 — always 2 (зарлага)
            bank_erp_code,           # Дансны код — банкны ERP код
            "",                      # Валютын дүн (хоосон — Валют холбоотой)
            "",                      # Авах / Зарах ханш
            "",                      # Олз / Гарзын дансны код
            t.debit,                 # Дүн
        ])
    if txns and eff_date:
        _apply_date_format(ws, "A", 2, len(txns) + 1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{stmt_id}/export")
def export_erkhet(
    stmt_id: int,
    export_date: Optional[str] = Query(None),   # "YYYY-MM-DD" — сонгосон огноо
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Хуулгын гүйлгээг 3 Эрхэт импорт Excel файл болгон ZIP архивт буцаана.
    • Кредит гүйлгээ → Авлага өглөгийн гүйлгээ.xlsx
    • Дебит (export_type=kass) → Мөнгөн хөрөнгийн кассын гүйлгээ.xlsx
    • Дебит (export_type=hariltsah) → Мөнгөн хөрөнгийн харилцахын гүйлгээ.xlsx
    """
    stmt = db.query(BankStatement).filter(BankStatement.id == stmt_id).first()
    if not stmt:
        raise HTTPException(404, "Хуулга олдсонгүй")

    from datetime import date as dt_date
    eff_date = None
    if export_date:
        try:
            eff_date = dt_date.fromisoformat(export_date)
        except ValueError:
            pass
    if not eff_date:
        eff_date = stmt.date_from

    # Тухайн дансны ERP код (POS auto-fill-д хэрэгтэй)
    cfg = db.query(BankAccountConfig).filter(
        BankAccountConfig.account_number == stmt.account_number
    ).first()
    bank_erp_code = (cfg.erp_account_code or "") if cfg else ""

    txns = [t for t in stmt.transactions if not t.is_fee]
    credit_txns    = [t for t in txns if t.credit > 0]
    kass_txns      = [t for t in txns if t.debit > 0 and (getattr(t, "export_type", "") or "") == "kass"]
    hariltsah_txns = [t for t in txns if t.debit > 0 and (getattr(t, "export_type", "") or "") == "hariltsah"]

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if credit_txns:
            zf.writestr("Авлага өглөгийн гүйлгээ.xlsx",
                        _build_avlaga_excel(credit_txns, eff_date, bank_erp_code))
        if kass_txns:
            zf.writestr("Мөнгөн хөрөнгийн кассын гүйлгээ.xlsx",
                        _build_kass_hariltsah_excel(kass_txns, eff_date, bank_erp_code))
        if hariltsah_txns:
            zf.writestr("Мөнгөн хөрөнгийн харилцахын гүйлгээ.xlsx",
                        _build_kass_hariltsah_excel(hariltsah_txns, eff_date))

    from urllib.parse import quote
    zip_buf.seek(0)
    date_label = eff_date.strftime("%Y%m%d") if eff_date else "export"
    acct = stmt.account_number or "statement"
    ascii_name = f"Erkhet_{acct}_{date_label}.zip"
    display    = f"Эрхэт_{acct}_{date_label}.zip"

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(display)}"},
    )


@router.delete("/{stmt_id}")
def delete_statement(
    stmt_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    s = db.query(BankStatement).filter(BankStatement.id == stmt_id).first()
    if not s:
        raise HTTPException(404, "Хуулга олдсонгүй")
    db.delete(s)
    db.commit()
    return {"ok": True}
