"""
Бараа тулгаж авах (receiving) endpoint-ууд.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as _func
from datetime import date as date_type, datetime
from pathlib import Path
from pydantic import BaseModel
from typing import Optional
import io
import re

from app.api.deps import get_db, get_current_user, require_role
from app.models.receiving import ReceivingSession, ReceivingLine, ReceivingBrandStatus
from app.models.product import Product
from app.models.user import User
from app.core.event_bus import publish as _publish_event


def _notify(session_id: int, action: str = "update") -> None:
    """Бусад device-ийн UI-д live refresh trigger өгнө."""
    try:
        _publish_event("receivings", session_id=session_id, action=action)
    except Exception:
        pass


router = APIRouter(prefix="/receivings", tags=["receivings"])

UPLOAD_DIR = Path("app/data/uploads/receiving_receipts")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

STATUS_SEQUENCE = ["matching", "price_review", "received"]
STATUS_LABEL = {
    "matching": "Тулгаж байна",
    "price_review": "Үнэ хянагдаж байна",
    "received": "Орлого авсан",
}


# ── Schemas ────────────────────────────────────────────────────────────────────

class SessionCreateIn(BaseModel):
    date: date_type
    notes: str = ""


class SessionUpdateIn(BaseModel):
    notes: Optional[str] = None
    date: Optional[date_type] = None


class LineIn(BaseModel):
    product_id: int
    qty_pcs: float = 0.0
    unit_price: float = 0.0
    note: str = ""
    override_brand: Optional[str] = None  # хоосон/None бол Product.brand-аар явна


class LineUpdateIn(BaseModel):
    qty_pcs: Optional[float] = None
    unit_price: Optional[float] = None
    note: Optional[str] = None
    override_brand: Optional[str] = None  # "" → override арилгах; None → хөдөлгөхгүй


class BrandMatchIn(BaseModel):
    supplier_total_pcs: float
    supplier_total_amount: float


class StatusIn(BaseModel):
    status: str


# ── Helpers ────────────────────────────────────────────────────────────────────

def _effective_brand(line: ReceivingLine, product: Optional[Product]) -> str:
    """Override-той бол override-ыг, үгүй бол Product.brand-ыг буцаана."""
    ob = (getattr(line, "override_brand", "") or "").strip()
    if ob:
        return ob
    return (product.brand if product else "") or ""


def _serialize_line(ln: ReceivingLine, product: Optional[Product]) -> dict:
    pack = float(product.pack_ratio or 1) if product else 1.0
    if pack <= 0:
        pack = 1.0
    box = int(ln.qty_pcs // pack)
    extra = int(round(ln.qty_pcs - box * pack))
    total_amount = round(ln.qty_pcs * ln.unit_price, 2)
    return {
        "id": ln.id,
        "product_id": ln.product_id,
        "item_code": product.item_code if product else "",
        "name": product.name if product else "(устгагдсан)",
        "brand": _effective_brand(ln, product),
        "original_brand": product.brand if product else "",
        "override_brand": (getattr(ln, "override_brand", "") or ""),
        "warehouse_name": product.warehouse_name if product else "",
        "pack_ratio": pack,
        "unit_weight": float(product.unit_weight or 0) if product else 0,
        "last_purchase_price": float(product.last_purchase_price or 0) if product else 0,
        "qty_pcs": ln.qty_pcs,
        "stock_box": box,
        "stock_extra_pcs": extra,
        "unit_price": ln.unit_price,
        "total_amount": total_amount,
        "price_reviewed": bool(getattr(ln, "price_reviewed", False)),
        "note": ln.note or "",
    }


def _brand_aggregate(session_id: int, db: Session) -> dict:
    """Бренд тус бүрээр нийт ширхэг + нийт дүн."""
    lines = db.query(ReceivingLine).filter(ReceivingLine.session_id == session_id).all()
    prod_ids = [l.product_id for l in lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()} if prod_ids else {}
    brands: dict[str, dict] = {}
    for l in lines:
        p = products.get(l.product_id)
        brand = _effective_brand(l, p) or "Брэнд байхгүй"
        if brand not in brands:
            brands[brand] = {"brand": brand, "line_count": 0, "total_pcs": 0.0, "total_amount": 0.0, "has_price_diff": False}
        brands[brand]["line_count"] += 1
        brands[brand]["total_pcs"] += l.qty_pcs
        brands[brand]["total_amount"] += l.qty_pcs * l.unit_price
        if p and p.last_purchase_price and abs((l.unit_price or 0) - float(p.last_purchase_price)) > 0.01 and l.unit_price > 0:
            brands[brand]["has_price_diff"] = True
    # Attach brand statuses
    statuses = {
        bs.brand: bs
        for bs in db.query(ReceivingBrandStatus).filter(ReceivingBrandStatus.session_id == session_id).all()
    }
    out = []
    for b, stats in sorted(brands.items()):
        bs = statuses.get(b)
        out.append({
            **stats,
            "total_pcs": round(stats["total_pcs"], 2),
            "total_amount": round(stats["total_amount"], 2),
            "is_matched": bool(bs.is_matched) if bs else False,
            "supplier_total_pcs": float(bs.supplier_total_pcs) if bs else 0.0,
            "supplier_total_amount": float(bs.supplier_total_amount) if bs else 0.0,
            "receipt_image_path": (bs.receipt_image_path if bs else "") or "",
            "matched_at": bs.matched_at.isoformat() if bs and bs.matched_at else None,
        })
    return out


def _serialize_session(s: ReceivingSession, db: Session, include_lines: bool = True) -> dict:
    creator = db.query(User).filter(User.id == s.created_by_user_id).first()
    result = {
        "id": s.id,
        "date": s.date.isoformat(),
        "notes": s.notes or "",
        "status": s.status,
        "status_label": STATUS_LABEL.get(s.status, s.status),
        "is_archived": bool(s.is_archived),
        "created_by_username": creator.username if creator else "",
        "created_at": s.created_at.isoformat() if s.created_at else None,
    }
    # Line + brand stats
    lines = db.query(ReceivingLine).filter(ReceivingLine.session_id == s.id).all()
    prod_ids = [l.product_id for l in lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()} if prod_ids else {}
    result["line_count"] = len(lines)
    result["total_pcs"] = round(sum(l.qty_pcs for l in lines), 2)
    result["total_amount"] = round(sum(l.qty_pcs * l.unit_price for l in lines), 2)
    # Brand info
    result["brands"] = _brand_aggregate(s.id, db)
    all_brands_matched = result["brands"] and all(b["is_matched"] for b in result["brands"])
    result["all_brands_matched"] = bool(all_brands_matched)
    if include_lines:
        result["lines"] = [_serialize_line(l, products.get(l.product_id)) for l in lines]
    return result


def _auto_advance_if_all_matched(session: ReceivingSession, db: Session):
    """Бүх effective brand matched болсон бөгөөд одоогоор 'matching' төлөвт байвал 'price_review' рүү шилжүүлнэ."""
    if session.status != "matching":
        return
    # Override-тай line-ыг тооцох тул Python-side-аар effective brand тооцно.
    # Хоосон brand нь UI-д "Брэнд байхгүй" гэж confirm хийгддэг тул тэр хэлбэрт нь хөрвүүлэв.
    lines = db.query(ReceivingLine).filter(ReceivingLine.session_id == session.id).all()
    prod_ids = [l.product_id for l in lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()} if prod_ids else {}
    all_brands = {(_effective_brand(l, products.get(l.product_id)) or "Брэнд байхгүй") for l in lines}
    if not all_brands:
        return
    matched_brands = set(
        bs.brand for bs in db.query(ReceivingBrandStatus).filter(
            ReceivingBrandStatus.session_id == session.id,
            ReceivingBrandStatus.is_matched == True,
        ).all()
    )
    if all_brands.issubset(matched_brands):
        session.status = "price_review"
        db.commit()


# ── Brand list (override dropdown-д ашиглагдана) ──────────────────────────────

@router.get("/brands/all")
def list_all_brands(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Системд бүртгэлтэй бүх distinct brand-ийг буцаана. Override dropdown-д хэрэглэнэ."""
    rows = (
        db.query(Product.brand)
        .filter(Product.brand != "")
        .filter(Product.brand.isnot(None))
        .distinct()
        .order_by(Product.brand.asc())
        .all()
    )
    return [r[0] for r in rows if r[0]]


# ── Session endpoints ─────────────────────────────────────────────────────────

@router.get("")
def list_sessions(
    status: Optional[str] = Query(None),
    archived: str = Query("false"),  # false | true | only
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    q = db.query(ReceivingSession)
    arch = (archived or "false").lower()
    if arch == "only":
        if u.role not in ("admin", "manager", "supervisor"):
            raise HTTPException(403, "Архив харах эрхгүй")
        q = q.filter(ReceivingSession.is_archived == True)
    elif arch != "true":
        q = q.filter(ReceivingSession.is_archived == False)
    if status:
        q = q.filter(ReceivingSession.status == status)
    rows = q.order_by(ReceivingSession.date.desc(), ReceivingSession.id.desc()).all()
    return [_serialize_session(s, db, include_lines=False) for s in rows]


@router.post("")
def create_session(
    body: SessionCreateIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    s = ReceivingSession(
        date=body.date,
        notes=(body.notes or "").strip(),
        status="matching",
        is_archived=False,
        created_by_user_id=u.id,
    )
    db.add(s); db.commit(); db.refresh(s)
    _notify(s.id, "session_created")
    return _serialize_session(s, db, include_lines=True)


@router.get("/{session_id}")
def get_session(
    session_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")
    if s.is_archived and u.role not in ("admin", "manager", "supervisor"):
        raise HTTPException(403, "Архивлагдсан тулгалтыг харах эрхгүй")
    return _serialize_session(s, db, include_lines=True)


@router.patch("/{session_id}")
def update_session(
    session_id: int,
    body: SessionUpdateIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")
    if body.notes is not None:
        s.notes = body.notes.strip()
    if body.date is not None:
        s.date = body.date
    db.commit()
    return _serialize_session(s, db, include_lines=False)


@router.patch("/{session_id}/status")
def set_status(
    session_id: int,
    body: StatusIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "accountant")),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")
    if body.status not in STATUS_SEQUENCE:
        raise HTTPException(400, "Буруу статус")
    s.status = body.status
    db.commit()
    _notify(session_id, "status_changed")
    return _serialize_session(s, db, include_lines=False)


@router.patch("/{session_id}/archive")
def archive_session(
    session_id: int,
    archived: bool = Query(True),
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor")),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")
    s.is_archived = bool(archived)
    db.commit()
    return {"ok": True, "is_archived": s.is_archived}


@router.delete("/{session_id}")
def delete_session(
    session_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor")),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")
    db.delete(s); db.commit()
    return {"ok": True}


# ── Line endpoints ────────────────────────────────────────────────────────────

@router.post("/{session_id}/lines")
def add_line(
    session_id: int,
    body: LineIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")
    p = db.query(Product).filter(Product.id == body.product_id).first()
    if not p:
        raise HTTPException(404, "Бараа олдсонгүй")
    # Override-ыг normalise: trim, мөн product.brand-тай тэнцэх бол хадгалахгүй (хоосон гэж үзнэ)
    incoming_override = (body.override_brand or "").strip()
    if incoming_override and incoming_override == (p.brand or ""):
        incoming_override = ""
    # Давхардсан бараа — нэг session-д нэг (product, override_brand) хослолыг merge
    existing = db.query(ReceivingLine).filter(
        ReceivingLine.session_id == session_id,
        ReceivingLine.product_id == body.product_id,
        ReceivingLine.override_brand == incoming_override,
    ).first()
    if existing:
        existing.qty_pcs += float(body.qty_pcs or 0)
        if body.unit_price and body.unit_price > 0:
            existing.unit_price = float(body.unit_price)
        if body.note:
            existing.note = body.note
        db.commit(); db.refresh(existing)
        _notify(session_id, "line_merged")
        return _serialize_line(existing, p)
    ln = ReceivingLine(
        session_id=session_id,
        product_id=body.product_id,
        qty_pcs=float(body.qty_pcs or 0),
        unit_price=float(body.unit_price or 0),
        override_brand=incoming_override,
        note=body.note or "",
    )
    db.add(ln); db.commit(); db.refresh(ln)
    _notify(session_id, "line_added")
    return _serialize_line(ln, p)


@router.patch("/{session_id}/lines/{line_id}")
def update_line(
    session_id: int,
    line_id: int,
    body: LineUpdateIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    ln = db.query(ReceivingLine).filter(
        ReceivingLine.id == line_id,
        ReceivingLine.session_id == session_id,
    ).first()
    if not ln:
        raise HTTPException(404, "Мөр олдсонгүй")
    p = db.query(Product).filter(Product.id == ln.product_id).first()
    # Edit guard: override_brand өөрчилбөл, өмнөх эсвэл шинэ effective brand нь
    # аль хэдий нь is_matched=True байвал бүү зөвшөөр.
    if body.override_brand is not None:
        new_override = (body.override_brand or "").strip()
        if new_override and new_override == (p.brand if p else ""):
            new_override = ""
        old_eff = _effective_brand(ln, p)
        new_eff = new_override if new_override else (p.brand if p else "") or ""
        if old_eff != new_eff:
            locked = db.query(ReceivingBrandStatus).filter(
                ReceivingBrandStatus.session_id == ln.session_id,
                ReceivingBrandStatus.brand.in_([old_eff, new_eff]),
                ReceivingBrandStatus.is_matched == True,
            ).first()
            if locked:
                raise HTTPException(
                    400,
                    f"'{locked.brand}' бренд аль хэдий нь тулгагдсан тул бренд override-ыг өөрчлөх боломжгүй. "
                    f"Эхлээд тулгалтыг буцаана уу.",
                )
        ln.override_brand = new_override
    if body.qty_pcs is not None:
        ln.qty_pcs = float(body.qty_pcs)
    if body.unit_price is not None:
        ln.unit_price = float(body.unit_price)
    if body.note is not None:
        ln.note = body.note
    db.commit(); db.refresh(ln)
    _notify(session_id, "line_updated")
    return _serialize_line(ln, p)


@router.patch("/{session_id}/lines/{line_id}/price-review")
def toggle_price_review(
    session_id: int,
    line_id: int,
    reviewed: bool = Query(...),
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    """price_review статусын үед хэрэглэгч мөрийг 'Хянасан/Хянагдаагүй' гэж toggle хийнэ."""
    ln = db.query(ReceivingLine).filter(
        ReceivingLine.id == line_id,
        ReceivingLine.session_id == session_id,
    ).first()
    if not ln:
        raise HTTPException(404, "Мөр олдсонгүй")
    ln.price_reviewed = bool(reviewed)
    db.commit(); db.refresh(ln)
    _notify(session_id, "price_reviewed")
    p = db.query(Product).filter(Product.id == ln.product_id).first()
    return _serialize_line(ln, p)


@router.delete("/{session_id}/lines/{line_id}")
def delete_line(
    session_id: int,
    line_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    ln = db.query(ReceivingLine).filter(
        ReceivingLine.id == line_id,
        ReceivingLine.session_id == session_id,
    ).first()
    if not ln:
        raise HTTPException(404, "Мөр олдсонгүй")
    db.delete(ln); db.commit()
    _notify(session_id, "line_deleted")
    return {"ok": True}


# ── Brand match + receipt upload ──────────────────────────────────────────────

@router.post("/{session_id}/brands/confirm")
async def confirm_brand(
    session_id: int,
    brand: str = Query(...),
    supplier_total_pcs: float = Form(...),
    supplier_total_amount: float = Form(...),
    receipt: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "warehouse_clerk", "accountant")),
):
    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")

    # "Брэнд байхгүй" гэдэг нь UI-ийн placeholder бөгөөд aggregator-аас гардаг.
    # DB дээр энэ нь үнэн хэрэгтээ хоосон string ("") юм.
    target_brand = "" if brand == "Брэнд байхгүй" else brand

    # Нийт тоог тулгах
    lines = db.query(ReceivingLine).filter(ReceivingLine.session_id == session_id).all()
    prod_ids = [l.product_id for l in lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()} if prod_ids else {}
    my_pcs = 0.0
    my_amount = 0.0
    matched_lines = 0
    for l in lines:
        p = products.get(l.product_id)
        if not p:
            continue
        if _effective_brand(l, p) != target_brand:
            continue
        my_pcs += l.qty_pcs
        my_amount += l.qty_pcs * l.unit_price
        matched_lines += 1

    if matched_lines == 0:
        raise HTTPException(400, f"'{brand}' брэндэд тохирох мөр олдсонгүй.")
    if abs(my_pcs - float(supplier_total_pcs)) > 0.01:
        raise HTTPException(400, f"Ширхэгийн тоо таарсангүй. Таны оруулсан: {my_pcs:.0f}ш, баримт дээр: {supplier_total_pcs:.0f}ш")
    if abs(my_amount - float(supplier_total_amount)) > 1.0:
        raise HTTPException(400, f"Нийт дүн таарсангүй. Таны оруулсан: {my_amount:.2f}₮, баримт дээр: {supplier_total_amount:.2f}₮")

    # Баримтны зураг хадгалах (заавал биш — зураггүй ч тулгаж болно)
    saved_path = ""
    if receipt and receipt.filename:
        try:
            sess_dir = UPLOAD_DIR / str(session_id)
            sess_dir.mkdir(parents=True, exist_ok=True)
            safe_brand = re.sub(r"[\\/:*?\"<>|]", "_", brand) or "no_brand"
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            suffix = Path(receipt.filename or "").suffix or ".jpg"
            saved_file = sess_dir / f"{safe_brand}_{ts}{suffix}"
            content = await receipt.read()
            if content:
                saved_file.write_bytes(content)
                saved_path = str(saved_file).replace("\\", "/")
        except Exception as e:
            raise HTTPException(500, f"Баримт хадгалахад алдаа гарлаа: {type(e).__name__}: {e}")

    bs = db.query(ReceivingBrandStatus).filter(
        ReceivingBrandStatus.session_id == session_id,
        ReceivingBrandStatus.brand == brand,
    ).first()
    if not bs:
        bs = ReceivingBrandStatus(session_id=session_id, brand=brand)
        db.add(bs)
    bs.is_matched = True
    if saved_path:
        bs.receipt_image_path = saved_path
    bs.supplier_total_pcs = float(supplier_total_pcs)
    bs.supplier_total_amount = float(supplier_total_amount)
    bs.matched_at = datetime.utcnow()
    db.commit()

    _auto_advance_if_all_matched(s, db)
    db.refresh(s)
    _notify(session_id, "brand_confirmed")
    return _serialize_session(s, db, include_lines=False)


@router.post("/{session_id}/brands/unmatch")
def unmatch_brand(
    session_id: int,
    brand: str = Query(...),
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor")),
):
    bs = db.query(ReceivingBrandStatus).filter(
        ReceivingBrandStatus.session_id == session_id,
        ReceivingBrandStatus.brand == brand,
    ).first()
    if not bs:
        raise HTTPException(404, "Тулгалт олдсонгүй")
    bs.is_matched = False
    bs.matched_at = None
    db.commit()
    _notify(session_id, "brand_unmatched")
    return {"ok": True}


@router.get("/{session_id}/brands/receipt")
def get_receipt(
    session_id: int,
    brand: str = Query(...),
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    bs = db.query(ReceivingBrandStatus).filter(
        ReceivingBrandStatus.session_id == session_id,
        ReceivingBrandStatus.brand == brand,
    ).first()
    if not bs or not bs.receipt_image_path:
        raise HTTPException(404, "Баримт олдсонгүй")
    p = Path(bs.receipt_image_path)
    if not p.exists():
        raise HTTPException(404, "Баримт файл байхгүй")
    return FileResponse(str(p))


# ── ERP Excel export ──────────────────────────────────────────────────────────

class ERPCfg(BaseModel):
    company: str  # "buten_orgil" | "orgil_khorum"
    date: Optional[str] = None
    document_note: str = ""
    related_account: str = ""
    account: str = ""
    warehouse_map: dict = {}
    single_location: str = ""
    brand_filter: str = ""


@router.post("/{session_id}/export-erp-excel")
def export_erp_excel(
    session_id: int,
    body: ERPCfg,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor", "accountant")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    s = db.query(ReceivingSession).filter(ReceivingSession.id == session_id).first()
    if not s:
        raise HTTPException(404, "Receiving session олдсонгүй")

    try:
        date_val = datetime.strptime(body.date, "%Y-%m-%d").date() if body.date else s.date
    except Exception:
        date_val = s.date

    lines = db.query(ReceivingLine).filter(
        ReceivingLine.session_id == session_id,
        ReceivingLine.qty_pcs > 0,
    ).all()
    prod_ids = [l.product_id for l in lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()} if prod_ids else {}
    # Effective brand -> brand_code lookup (prefer non-empty codes).
    # Needed so override_brand lines are grouped under the selected brand's supplier code.
    brand_code_rows = db.query(Product.brand, Product.brand_code).filter(
        Product.brand.isnot(None),
        Product.brand != "",
    ).all()
    brand_code_by_brand: dict[str, str] = {}
    for b, bc in brand_code_rows:
        key = (b or "").strip()
        code = (bc or "").strip()
        if not key:
            continue
        if code and not brand_code_by_brand.get(key):
            brand_code_by_brand[key] = code
        elif key not in brand_code_by_brand:
            brand_code_by_brand[key] = ""
    matched_brands = {
        (bs.brand or "").strip()
        for bs in db.query(ReceivingBrandStatus).filter(
            ReceivingBrandStatus.session_id == session_id,
            ReceivingBrandStatus.is_matched == True,
        ).all()
    }

    brand_filter = (body.brand_filter or "").strip()
    valid = []
    for ln in lines:
        p = products.get(ln.product_id)
        if not p:
            continue
        brand = (_effective_brand(ln, p) or "").strip()
        # Consolidated ERP export must follow only confirmed/matched brand sets.
        # This avoids exporting stale lines from unmatched (old/error) reconciliation.
        if brand not in matched_brands:
            continue
        if brand_filter and brand != brand_filter:
            continue
        if body.company == "orgil_khorum":
            location = body.single_location
        else:
            location = body.warehouse_map.get(p.warehouse_name, "")
        total = round(ln.qty_pcs * ln.unit_price, 2)
        # Use effective brand code (override-aware), fallback to current product brand_code.
        eff_brand_code = (brand_code_by_brand.get(brand, "") or (p.brand_code or "")).strip()
        valid.append((brand, p.item_code, p, ln, location, ln.qty_pcs, ln.unit_price, total, eff_brand_code))

    valid.sort(key=lambda x: (x[8] or "", x[0], x[1]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    col_headers = [
        "Огноо", "Баримтын дугаар", "Гүйлгээний утга", "Харилцагч",
        "Харьцсан данс", "Харьцсан ялгаатай харилцагч",
        "НӨАТ тай эсэх", "НӨАТ-н үзүүлэлт", "НӨАТ автоматаар бодох эсэх", "НӨАТ-н дүн",
        "НХАТ тай эсэх", "НХАТ автоматаар бодох эсэх", "НХАТ-н дүн",
        "Данс", "Бараа материал", "Барааны байршил",
        "Тоо хэмжээ", "Нэгж үнэ", "Хувийн жин", "Нийт дүн",
        "НӨАТ тооцох эсэх", "НХАТ тооцох эсэх", "НХАТ мөр",
    ]
    hdr_fill = PatternFill("solid", fgColor="3258A0")
    hdr_font = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    from collections import defaultdict
    groups: dict = defaultdict(list)
    for item in valid:
        groups[item[8] or ""].append(item)

    current_row = 2
    for supplier_code, items in groups.items():
        for i, (brand, item_code, p, ln, location, qty, price, total, _bc) in enumerate(items):
            is_first = (i == 0)
            row = [
                date_val if is_first else None,
                None,
                body.document_note if is_first else None,
                supplier_code if is_first else None,
                body.related_account if is_first else None, None,
                0 if is_first else None,
                None, None,
                0 if is_first else None,
                0 if is_first else None,
                None,
                0 if is_first else None,
                body.account,
                p.item_code,
                location,
                qty,
                price,
                1.0,
                total,
                0,
                0,
                None,
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(row=current_row, column=ci, value=val)
            current_row += 1

    widths = [14, 16, 28, 14, 14, 20, 14, 18, 22, 12, 14, 22, 12,
              16, 18, 20, 12, 12, 12, 14, 18, 18, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = s.date.strftime("%Y%m%d")
    brand_part = re.sub(r"[\\/:*?\"<>|]", "_", brand_filter) if brand_filter else "all"
    filename = f"{date_str}_RECV{s.id}_{brand_part}.xlsx"
    ascii_fallback = re.sub(r"[^\w\-.]", "_", filename.encode("ascii", "ignore").decode("ascii")) or f"RECV{s.id}.xlsx"
    utf8_quoted = quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={ascii_fallback}; filename*=UTF-8''{utf8_quoted}"},
    )
