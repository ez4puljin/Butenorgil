from pathlib import Path
from datetime import date as date_type
from typing import Optional, List
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.api.deps import get_db, get_current_user, require_role, parse_tag_ids
from app.models.purchase_order import (
    PurchaseOrder, PurchaseOrderLine, PurchaseOrderBrandVehicle,
    PurchaseOrderBrandStatus, OrderExtraLine, POShipment, POShipmentLine,
)
from app.models.product import Product
from app.models.user import User
from app.models.logistics import Vehicle

router = APIRouter(prefix="/purchase-orders", tags=["purchase-orders"])

# master_latest.xlsx location (same OUTPUT_DIR as imports.py)
MASTER_FILE = Path("app/data/outputs/master_latest.xlsx")

# Status transition map
STATUS_SEQUENCE = [
    "preparing",
    "reviewing",
    "sending",
    "loading",
    "transit",
    "arrived",
    "accounting",
    "confirmed",
    "received",
]

STATUS_LABEL = {
    "preparing":  "Захиалга бэлдэж байна",
    "reviewing":  "Хянаж байна",
    "sending":    "Захиалга илгээж байна",
    "loading":    "Ачигдаж байна",
    "transit":    "Замд явж байна",
    "arrived":    "Ачаа ирсэн",
    "accounting": "Нягтлан шалгаж байна",
    "confirmed":  "Нягтлан Баталгаажсан",
    "received":   "Орлого авагдсан",
}


def _next_status(current: str) -> Optional[str]:
    try:
        idx = STATUS_SEQUENCE.index(current)
        if idx + 1 < len(STATUS_SEQUENCE):
            return STATUS_SEQUENCE[idx + 1]
    except ValueError:
        pass
    return None


def _serialize_order(o: PurchaseOrder, db: Session) -> dict:
    """Compact summary for list view."""
    creator = db.query(User).filter(User.id == o.created_by_user_id).first()
    total_boxes = sum(l.order_qty_box for l in o.lines)
    total_weight = sum(l.computed_weight for l in o.lines)
    vehicle = db.query(Vehicle).filter(Vehicle.id == o.vehicle_id).first() if o.vehicle_id else None
    return {
        "id": o.id,
        "order_date": o.order_date.isoformat(),
        "status": o.status,
        "status_label": STATUS_LABEL.get(o.status, o.status),
        "created_by_username": creator.username if creator else "",
        "line_count": len(o.lines),
        "total_boxes": round(total_boxes, 2),
        "total_weight": round(total_weight, 2),
        "created_at": o.created_at.isoformat() if o.created_at else None,
        "vehicle_id": o.vehicle_id,
        "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
        "notes": o.notes or "",
    }


def _serialize_order_detail(
    o: PurchaseOrder,
    db: Session,
    filter_tag_ids: Optional[List[int]] = None,
) -> dict:
    """Full detail including product info for each line.
    If filter_tag_ids is provided, only lines whose product belongs to those warehouses are returned.
    Uses SQL queries instead of lazy-loading 10K+ lines into memory.
    """
    # Use SQL aggregate for base stats (don't lazy-load o.lines for summary)
    agg = db.query(
        func.count(PurchaseOrderLine.id),
        func.sum(PurchaseOrderLine.order_qty_box),
        func.sum(PurchaseOrderLine.computed_weight),
    ).filter(PurchaseOrderLine.purchase_order_id == o.id).first()

    creator = db.query(User).filter(User.id == o.created_by_user_id).first()
    vehicle = db.query(Vehicle).filter(Vehicle.id == o.vehicle_id).first() if o.vehicle_id else None

    base = {
        "id": o.id,
        "order_date": o.order_date.isoformat(),
        "status": o.status,
        "status_label": STATUS_LABEL.get(o.status, o.status),
        "created_by_username": creator.username if creator else "",
        "line_count": agg[0] or 0,
        "total_boxes": round(float(agg[1] or 0), 2),
        "total_weight": round(float(agg[2] or 0), 2),
        "created_at": o.created_at.isoformat() if o.created_at else None,
        "vehicle_id": o.vehicle_id,
        "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
        "notes": o.notes or "",
        "is_archived": bool(o.is_archived),
    }

    # Load lines via explicit query (NOT lazy o.lines which loads ALL 10K+ rows)
    # For preparing stage: load all lines (user needs to enter quantities for any product)
    # For other stages: only load lines with order_qty_box > 0 (or supplier_qty > 0 for cancelled tracking)
    lines_q = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == o.id)
    if o.status not in ("preparing",):
        lines_q = lines_q.filter(
            (PurchaseOrderLine.order_qty_box > 0) | (PurchaseOrderLine.supplier_qty_box > 0)
        )
    all_lines = lines_q.all()

    # Bulk load products
    product_ids = [l.product_id for l in all_lines]
    products = db.query(Product).filter(Product.id.in_(product_ids)).all() if product_ids else []
    product_map = {p.id: p for p in products}

    # Min-stock rules (нэг удаа ачаалаад бараа бүрт match хийнэ)
    from app.models.min_stock_rule import MinStockRule
    from app.services.min_stock_check import find_rule_for_product, compute_needs_reorder, stock_breakdown
    _ms_rules = db.query(MinStockRule).filter(MinStockRule.is_active == True).all()

    # Bulk load shipment line totals: po_line_id → total loaded (sum across shipments)
    line_ids = [l.id for l in all_lines]
    shipped_loaded: dict[int, float] = {}
    shipped_received: dict[int, float] = {}
    if line_ids:
        from sqlalchemy import func as _func
        rows = (
            db.query(
                POShipmentLine.po_line_id,
                _func.sum(POShipmentLine.loaded_qty_box),
                _func.sum(POShipmentLine.received_qty_box),
            )
            .filter(POShipmentLine.po_line_id.in_(line_ids))
            .group_by(POShipmentLine.po_line_id)
            .all()
        )
        for row in rows:
            shipped_loaded[row[0]] = float(row[1] or 0)
            shipped_received[row[0]] = float(row[2] or 0)

    lines_out = []
    for l in all_lines:
        p = product_map.get(l.product_id)
        if not p:
            continue
        # Warehouse clerk sees only their assigned warehouses.
        # Products with warehouse_tag_id=0 are shared — visible to everyone.
        if filter_tag_ids and p.warehouse_tag_id != 0 and p.warehouse_tag_id not in filter_tag_ids:
            continue
        lpp = float(p.last_purchase_price or 0)
        up = float(l.unit_price or 0)
        qty = float(l.order_qty_box or 0)
        estimated_cost = round(lpp * qty, 2)
        price_diff = round(up - lpp, 2) if up > 0 and lpp > 0 else None

        # Prefer shipment-aggregated loaded/received values if shipments exist
        loaded_effective = shipped_loaded.get(l.id, float(l.loaded_qty_box or 0))
        if loaded_effective == 0 and (l.loaded_qty_box or 0) > 0:
            loaded_effective = float(l.loaded_qty_box)
        received_effective = shipped_received.get(l.id, float(l.received_qty_box or 0))
        if received_effective == 0 and (l.received_qty_box or 0) > 0:
            received_effective = float(l.received_qty_box)

        # Min-stock rule match → needs_reorder + min_stock_box
        matched_rule = find_rule_for_product(p, _ms_rules)
        needs_reorder, min_stock_box = compute_needs_reorder(p, matched_rule)
        bd = stock_breakdown(p)

        # Үр дүнгийн бренд: override_brand тохиргоотой бол түүнийг, эс бол product.brand
        eff_brand = (l.override_brand or "").strip() or p.brand
        lines_out.append({
            "line_id": l.id,
            "product_id": l.product_id,
            "item_code": p.item_code,
            "name": p.name,
            "brand": eff_brand,
            "original_brand": p.brand,
            "override_brand": l.override_brand or "",
            "warehouse_tag_id": p.warehouse_tag_id,
            "warehouse_name": p.warehouse_name or "",
            "price_tag": p.price_tag or "",
            "unit_weight": p.unit_weight,
            "pack_ratio": p.pack_ratio,
            "stock_qty": p.stock_qty,
            "stock_box": bd["stock_box"],
            "stock_extra_pcs": bd["stock_extra_pcs"],
            "sales_qty": p.sales_qty,
            "needs_reorder": needs_reorder,
            "min_stock_box": min_stock_box,
            "order_qty_box": l.order_qty_box,
            "order_qty_pcs": l.order_qty_pcs,
            "computed_weight": l.computed_weight,
            "supplier_qty_box": l.supplier_qty_box,
            "loaded_qty_box": loaded_effective,
            "received_qty_box": received_effective,
            "received_qty_extra_pcs": float(l.received_qty_extra_pcs or 0),
            "difference": round(loaded_effective - received_effective, 2),
            "unit_price": up,
            "last_purchase_price": lpp,
            "estimated_cost": estimated_cost,
            "price_diff": price_diff,
            "remark": l.line_remark or "",
        })
    # Sort by brand, then item_code
    lines_out.sort(key=lambda x: (x["brand"], x["item_code"]))
    base["notes"] = o.notes
    base["lines"] = lines_out
    base["total_estimated_cost"] = round(sum(l["estimated_cost"] for l in lines_out), 2)
    base["price_diff_count"] = sum(1 for l in lines_out if l["price_diff"] is not None and abs(l["price_diff"]) > 0.01)
    base["next_status"] = _next_status(o.status)
    base["next_status_label"] = STATUS_LABEL.get(_next_status(o.status) or "", "")

    # Brand-vehicle assignments
    bvs = db.query(PurchaseOrderBrandVehicle).filter(
        PurchaseOrderBrandVehicle.purchase_order_id == o.id
    ).all()
    # Bulk load vehicles for brand-vehicle assignments
    bv_vehicle_ids = {bv.vehicle_id for bv in bvs if bv.vehicle_id}
    bv_vehicles = {v.id: v for v in db.query(Vehicle).filter(Vehicle.id.in_(bv_vehicle_ids)).all()} if bv_vehicle_ids else {}
    bv_out = []
    for bv in bvs:
        vehicle = bv_vehicles.get(bv.vehicle_id)
        bv_out.append({
            "brand": bv.brand,
            "vehicle_id": bv.vehicle_id,
            "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
        })
    base["brand_vehicles"] = bv_out

    # Extra lines (supplier-added items not in product catalog)
    base["extra_lines"] = [{
        "id": el.id,
        "brand": el.brand,
        "name": el.name,
        "item_code": el.item_code,
        "warehouse_name": el.warehouse_name,
        "unit_weight": el.unit_weight,
        "pack_ratio": el.pack_ratio,
        "qty_box": el.qty_box,
        "computed_weight": el.computed_weight,
    } for el in o.extra_lines]

    # Per-brand statuses
    bs_rows = db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == o.id
    ).all()
    base["brand_statuses"] = {bs.brand: bs.status for bs in bs_rows}

    return base


# ── Endpoints ──────────────────────────────────────────────────────────────────

class POCreateIn(BaseModel):
    order_date: date_type
    notes: str = ""
    # None эсвэл хоосон бол бүх бренд, эс бол зөвхөн сонгосон бренд-ийн барааг л оруулна.
    brands: Optional[List[str]] = None


class POLineIn(BaseModel):
    product_id: int
    order_qty_box: float
    supplier_qty_box: Optional[float] = None
    loaded_qty_box: Optional[float] = None
    received_qty_box: Optional[float] = None
    received_qty_extra_pcs: Optional[float] = None
    unit_price: Optional[float] = None
    remark: Optional[str] = None


class AddLineIn(BaseModel):
    product_id: int
    order_qty_box: float = 1.0
    # Онцгой тохиолдолд: тухайн бараа Q бренд-ээс байвал override_brand="W" бичвэл
    # W брендийн хэсэгт харагдана. Зөвхөн admin эрхтэй үед зөвшөөрнө.
    override_brand: Optional[str] = None


class POVehicleIn(BaseModel):
    vehicle_id: Optional[int] = None


# ── Dashboard endpoint ────────────────────────────────────────────────────────

@router.get("/{order_id}/dashboard")
def get_order_dashboard(
    order_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    """Захиалгын бренд-түвшний нэгтгэсэн dashboard мэдээлэл."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    # Load only active lines (NOT lazy po.lines which loads ALL 10K+ rows)
    active_lines = db.query(PurchaseOrderLine).filter(
        PurchaseOrderLine.purchase_order_id == order_id,
        (PurchaseOrderLine.order_qty_box > 0) | (PurchaseOrderLine.supplier_qty_box > 0)
    ).all()

    # Bulk load products for active lines only
    product_ids = [l.product_id for l in active_lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}

    # All shipments + shipment lines for this PO
    shipments = db.query(POShipment).filter(POShipment.purchase_order_id == order_id).order_by(POShipment.id).all()
    all_ship_lines = (
        db.query(POShipmentLine)
        .join(POShipment, POShipment.id == POShipmentLine.shipment_id)
        .filter(POShipment.purchase_order_id == order_id)
        .all()
    )

    # assigned_map: po_line_id → total loaded across all shipments
    assigned_map: dict[int, float] = {}
    # shipment_map: po_line_id → set of shipment_ids
    shipment_line_map: dict[int, set[int]] = {}
    for sl in all_ship_lines:
        assigned_map[sl.po_line_id] = assigned_map.get(sl.po_line_id, 0) + sl.loaded_qty_box
        shipment_line_map.setdefault(sl.po_line_id, set()).add(sl.shipment_id)

    # Shipment status map
    shipment_status_map = {s.id: s.status for s in shipments}

    # ── Group lines by brand ──
    brand_data: dict[str, dict] = {}
    cancelled_lines = 0
    cancelled_brands_set: set[str] = set()

    for l in active_lines:
        p = products.get(l.product_id)
        if not p:
            continue

        is_cancelled = l.order_qty_box == 0 and (l.supplier_qty_box or 0) > 0

        brand = (l.override_brand or "").strip() or p.brand or "Брэнд байхгүй"
        if brand not in brand_data:
            brand_data[brand] = {
                "brand": brand,
                "line_count": 0,
                "total_order_boxes": 0.0,
                "total_loaded_boxes": 0.0,
                "total_received_boxes": 0.0,
                "total_weight": 0.0,
                "estimated_cost": 0.0,
                "shipment_ids": set(),
                "vehicle_ids": set(),
                "items": [],
                "has_active_lines": False,
                "all_cancelled": True,
            }
        bd = brand_data[brand]

        loaded = assigned_map.get(l.id, 0)
        remaining = max(0.0, l.order_qty_box - loaded)
        weight = l.order_qty_box * float(p.pack_ratio or 1) * float(p.unit_weight or 0)
        lpp = float(p.last_purchase_price or 0)

        if is_cancelled:
            cancelled_lines += 1
        else:
            bd["all_cancelled"] = False

        if l.order_qty_box > 0:
            bd["has_active_lines"] = True
            bd["all_cancelled"] = False

        bd["line_count"] += 1
        bd["total_order_boxes"] += l.order_qty_box
        bd["total_loaded_boxes"] += loaded
        bd["total_received_boxes"] += (l.received_qty_box or 0)
        bd["total_weight"] += weight
        bd["estimated_cost"] += lpp * l.order_qty_box

        # Track shipments/vehicles
        sids = shipment_line_map.get(l.id, set())
        bd["shipment_ids"].update(sids)
        for sid in sids:
            sh = next((s for s in shipments if s.id == sid), None)
            if sh and sh.vehicle_id:
                bd["vehicle_ids"].add(sh.vehicle_id)

        bd["items"].append({
            "item_code": p.item_code,
            "name": p.name,
            "order_qty_box": l.order_qty_box,
            "loaded_qty_box": loaded,
            "unloaded_qty": remaining,
            "received_qty_box": l.received_qty_box or 0,
            "weight": round(weight, 2),
            "is_cancelled": is_cancelled,
        })

    # Derive brand_status
    def _brand_status(bd: dict) -> str:
        if bd["all_cancelled"]:
            return "cancelled"
        if not bd["has_active_lines"]:
            return "cancelled"
        loaded = bd["total_loaded_boxes"]
        ordered = bd["total_order_boxes"]
        if ordered <= 0:
            return "cancelled"
        if loaded <= 0:
            return "unloaded"
        # Check shipment statuses for loaded lines
        statuses = {shipment_status_map.get(sid, "loading") for sid in bd["shipment_ids"]}
        if loaded < ordered:
            if statuses & {"transit", "arrived", "accounting", "confirmed", "received"}:
                return "partial"
            return "partial"
        # Fully loaded — return highest shipment status
        if "received" in statuses and len(statuses) == 1:
            return "received"
        if statuses <= {"confirmed", "received"}:
            return "confirmed"
        if statuses <= {"arrived", "accounting", "confirmed", "received"}:
            return "arrived"
        if "transit" in statuses:
            return "transit"
        return "loaded"

    # Load persisted brand statuses from DB
    persisted_bs = {
        bs.brand: bs.status
        for bs in db.query(PurchaseOrderBrandStatus).filter(
            PurchaseOrderBrandStatus.purchase_order_id == order_id
        ).all()
    }

    # Bulk load all vehicles referenced by brands
    all_vehicle_ids: set[int] = set()
    for bd in brand_data.values():
        all_vehicle_ids.update(bd["vehicle_ids"])
    for sh in shipments:
        if sh.vehicle_id:
            all_vehicle_ids.add(sh.vehicle_id)
    vehicle_map = {v.id: v for v in db.query(Vehicle).filter(Vehicle.id.in_(all_vehicle_ids)).all()} if all_vehicle_ids else {}

    # Build brands list
    brands_list = []
    for bd in sorted(brand_data.values(), key=lambda x: x["brand"]):
        vehicle_names = []
        for vid in bd["vehicle_ids"]:
            v = vehicle_map.get(vid)
            if v:
                vehicle_names.append(f"{v.name} ({v.plate})")

        brands_list.append({
            "brand": bd["brand"],
            "line_count": bd["line_count"],
            "total_order_boxes": round(bd["total_order_boxes"], 1),
            "total_loaded_boxes": round(bd["total_loaded_boxes"], 1),
            "total_unloaded_boxes": round(max(0, bd["total_order_boxes"] - bd["total_loaded_boxes"]), 1),
            "total_received_boxes": round(bd["total_received_boxes"], 1),
            "total_weight": round(bd["total_weight"], 1),
            "estimated_cost": round(bd["estimated_cost"], 2),
            "brand_status": persisted_bs.get(bd["brand"], _brand_status(bd)),
            "brand_status_label": STATUS_LABEL.get(persisted_bs.get(bd["brand"], _brand_status(bd)), ""),
            "vehicle_names": vehicle_names,
            "items": bd["items"],
        })

    cancelled_brands_count = sum(1 for b in brands_list if b["brand_status"] == "cancelled")

    # ── Extra lines ──
    extra_brand_map: dict[str, list] = {}
    for el in po.extra_lines:
        b = el.brand or "Нэмэлт бараа"
        extra_brand_map.setdefault(b, []).append({
            "name": el.name, "item_code": el.item_code,
            "qty_box": el.qty_box, "computed_weight": el.computed_weight,
        })
    extra_brands = [
        {
            "brand": b,
            "items": items,
            "total_boxes": sum(i["qty_box"] for i in items),
            "total_weight": round(sum(i["computed_weight"] for i in items), 2),
        }
        for b, items in sorted(extra_brand_map.items())
    ]

    # ── Shipments with per-brand breakdown ──
    active_line_by_id = {l.id: l for l in active_lines}
    shipments_out = []
    for sh in shipments:
        v = vehicle_map.get(sh.vehicle_id) if sh.vehicle_id else None
        sh_lines = [sl for sl in all_ship_lines if sl.shipment_id == sh.id]
        sh_brand_map: dict[str, dict] = {}
        total_weight = 0.0
        for sl in sh_lines:
            pl = active_line_by_id.get(sl.po_line_id)
            if not pl:
                continue
            p = products.get(pl.product_id)
            if not p:
                continue
            b = (pl.override_brand or "").strip() or p.brand or "Брэнд байхгүй"
            w = sl.loaded_qty_box * float(p.pack_ratio or 1) * float(p.unit_weight or 0)
            total_weight += w
            if b not in sh_brand_map:
                sh_brand_map[b] = {"brand": b, "loaded_boxes": 0, "received_boxes": 0, "weight": 0, "line_count": 0}
            sh_brand_map[b]["loaded_boxes"] += sl.loaded_qty_box
            sh_brand_map[b]["received_boxes"] += sl.received_qty_box
            sh_brand_map[b]["weight"] += w
            sh_brand_map[b]["line_count"] += 1

        cap_kg = float(v.capacity_kg) if v else 0
        shipments_out.append({
            "id": sh.id,
            "vehicle_id": sh.vehicle_id,
            "vehicle_name": f"{v.name} ({v.plate})" if v else None,
            "driver_name": v.driver_name if v else None,
            "capacity_kg": cap_kg,
            "status": sh.status,
            "status_label": SHIPMENT_STATUS_LABEL.get(sh.status, sh.status),
            "brands": sorted(sh_brand_map.values(), key=lambda x: x["brand"]),
            "total_loaded_boxes": round(sum(sl.loaded_qty_box for sl in sh_lines), 1),
            "total_weight": round(total_weight, 1),
            "capacity_pct": round(total_weight / cap_kg * 100, 1) if cap_kg > 0 else 0,
        })

    # ── Unloaded pool ──
    unloaded_brands: dict[str, dict] = {}
    for l in active_lines:
        if l.order_qty_box <= 0:
            continue
        loaded = assigned_map.get(l.id, 0)
        remaining = l.order_qty_box - loaded
        if remaining <= 0.001:
            continue
        p = products.get(l.product_id)
        if not p:
            continue
        b = (l.override_brand or "").strip() or p.brand or "Брэнд байхгүй"
        if b not in unloaded_brands:
            unloaded_brands[b] = {"brand": b, "total_remaining_boxes": 0, "total_weight": 0, "items": []}
        w = remaining * float(p.pack_ratio or 1) * float(p.unit_weight or 0)
        unloaded_brands[b]["total_remaining_boxes"] += remaining
        unloaded_brands[b]["total_weight"] += w
        unloaded_brands[b]["items"].append({
            "item_code": p.item_code, "name": p.name,
            "remaining_boxes": round(remaining, 1), "weight": round(w, 2),
        })

    unloaded_list = sorted(unloaded_brands.values(), key=lambda x: x["brand"])
    for ub in unloaded_list:
        ub["total_remaining_boxes"] = round(ub["total_remaining_boxes"], 1)
        ub["total_weight"] = round(ub["total_weight"], 1)

    active_brands = [b for b in brands_list if b["brand_status"] != "cancelled" and b["total_order_boxes"] > 0]

    return {
        "order": {
            "id": po.id,
            "order_date": po.order_date.isoformat(),
            "status": po.status,
            "status_label": STATUS_LABEL.get(po.status, po.status),
            "notes": po.notes or "",
        },
        "summary": {
            "total_brands": len(active_brands),
            "total_boxes": round(sum(b["total_order_boxes"] for b in active_brands), 1),
            "total_weight": round(sum(b["total_weight"] for b in active_brands), 1),
            "total_estimated_cost": round(sum(b["estimated_cost"] for b in active_brands), 2),
            "cancelled_lines": cancelled_lines,
            "cancelled_brands": cancelled_brands_count,
        },
        "brands": brands_list,
        "extra_brands": extra_brands,
        "shipments": shipments_out,
        "unloaded_pool": {
            "brands": unloaded_list,
            "total_remaining_boxes": round(sum(ub["total_remaining_boxes"] for ub in unloaded_list), 1),
            "total_weight": round(sum(ub["total_weight"] for ub in unloaded_list), 1),
        },
        "available_vehicles": [
            {"id": v.id, "name": v.name, "plate": v.plate, "is_active": v.is_active}
            for v in db.query(Vehicle).filter(Vehicle.is_active == True).order_by(Vehicle.name).all()
        ],
    }


# ── Per-brand status endpoints ────────────────────────────────────────────────

@router.patch("/{order_id}/brand-advance")
def advance_brand_status(
    order_id: int,
    brand: str = Query(...),
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    """Тухайн брендийн статусыг дараагийн stage руу шилжүүлнэ."""

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    bs = db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id,
        PurchaseOrderBrandStatus.brand == brand,
    ).first()
    if not bs:
        raise HTTPException(404, f"'{brand}' брендийн статус олдсонгүй")

    next_st = _next_status(bs.status)
    if not next_st:
        raise HTTPException(400, "Эцсийн статуст хүрсэн")

    # Role check (same logic as advance_status)
    if u.role == "warehouse_clerk":
        raise HTTPException(403, "Статус шилжүүлэх эрхгүй")
    if u.role == "accountant" and bs.status not in ("accounting", "confirmed"):
        raise HTTPException(403, "Нягтлан зөвхөн accounting/confirmed статусыг шилжүүлнэ")

    # Side effect: arrived → accounting → pre-fill unit_price for this brand's lines
    if bs.status == "arrived" and next_st == "accounting":
        product_ids = [l.product_id for l in po.lines]
        products = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()}
        for line in po.lines:
            p = products.get(line.product_id)
            if not p:
                continue
            eff_brand = (line.override_brand or "").strip() or p.brand
            if eff_brand == brand:
                if (line.unit_price or 0) == 0 and (p.last_purchase_price or 0) > 0:
                    line.unit_price = p.last_purchase_price

    bs.status = next_st
    db.commit()

    _sync_po_status_from_brands(order_id, db)

    return {
        "brand": brand,
        "new_status": next_st,
        "new_status_label": STATUS_LABEL.get(next_st, next_st),
        "po_status": po.status,
    }


@router.get("/{order_id}/brand-detail")
def get_brand_detail(
    order_id: int,
    brand: str = Query(...),
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    """Тухайн брендийн дэлгэрэнгүй мэдээлэл (filtered detail)."""

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    bs = db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id,
        PurchaseOrderBrandStatus.brand == brand,
    ).first()

    brand_status = bs.status if bs else po.status

    # Filter lines to this brand
    detail = _serialize_order_detail(po, db)
    detail["lines"] = [l for l in detail["lines"] if l["brand"] == brand]
    detail["brand_filter"] = brand
    detail["brand_status"] = brand_status
    detail["brand_next_status"] = _next_status(brand_status)
    detail["brand_next_status_label"] = STATUS_LABEL.get(_next_status(brand_status) or "", "")

    return detail


@router.patch("/{order_id}/vehicle")
def set_order_vehicle(
    order_id: int,
    body: POVehicleIn,
    db: Session = Depends(get_db),
    u=Depends(require_role("admin", "manager")),
):
    o = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not o:
        raise HTTPException(404, "Order not found")
    if body.vehicle_id is not None:
        v = db.query(Vehicle).filter(Vehicle.id == body.vehicle_id, Vehicle.is_active == True).first()
        if not v:
            raise HTTPException(400, "Vehicle not found")
    o.vehicle_id = body.vehicle_id
    db.commit()
    vehicle = db.query(Vehicle).filter(Vehicle.id == o.vehicle_id).first() if o.vehicle_id else None
    return {
        "vehicle_id": o.vehicle_id,
        "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
    }


@router.get("/dashboard-stats")
def purchase_order_dashboard_stats(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Хяналтын самбарт харуулах захиалгын хураангуй статистик."""
    orders = db.query(PurchaseOrder).order_by(PurchaseOrder.order_date.desc()).all()

    by_status: dict[str, int] = {s: 0 for s in STATUS_SEQUENCE}
    active_weight = 0.0
    active_boxes = 0.0
    transit_weight = 0.0
    transit_boxes = 0.0
    latest_active: Optional[dict] = None

    for o in orders:
        by_status[o.status] = by_status.get(o.status, 0) + 1
        order_weight = sum(l.computed_weight for l in o.lines)
        order_boxes  = sum(l.order_qty_box for l in o.lines)
        if o.status != "arrived":
            active_weight += order_weight
            active_boxes  += order_boxes
        if o.status == "transit":
            transit_weight += order_weight
            transit_boxes  += order_boxes
        # Most recent non-arrived order
        if latest_active is None and o.status != "arrived":
            latest_active = {
                "id": o.id,
                "order_date": o.order_date.isoformat(),
                "status": o.status,
                "status_label": STATUS_LABEL.get(o.status, o.status),
                "total_weight": round(order_weight, 2),
                "total_boxes":  round(order_boxes, 0),
            }

    active_count = sum(v for k, v in by_status.items() if k != "arrived")

    return {
        "total":          len(orders),
        "active":         active_count,
        "arrived":        by_status.get("arrived", 0),
        "by_status":      by_status,
        "active_weight":  round(active_weight, 2),
        "active_boxes":   round(active_boxes, 0),
        "transit_weight": round(transit_weight, 2),
        "transit_boxes":  round(transit_boxes, 0),
        "latest_active":  latest_active,
    }


@router.get("/master-check")
def master_check(_=Depends(get_current_user)):
    """Check if master_latest.xlsx exists."""
    exists = MASTER_FILE.exists()
    updated_at = None
    if exists:
        import os
        ts = os.path.getmtime(MASTER_FILE)
        from datetime import datetime, timezone
        updated_at = datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()
    return {"exists": exists, "updated_at": updated_at}


@router.get("")
def list_purchase_orders(
    status: Optional[str] = Query(None),
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    archived: Optional[str] = Query("false"),  # "false" | "true" | "only"
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    q = db.query(PurchaseOrder)

    # warehouse_clerk sees preparing and arrived orders
    if u.role == "warehouse_clerk":
        q = q.filter(PurchaseOrder.status.in_(["preparing", "arrived"]))
    # manager/supervisor/admin see all

    # Архив filter:
    # - "false" (default): архив бус
    # - "true": бүгд (архив + бус)
    # - "only": зөвхөн архивлагдсан (admin/manager л хандана)
    arch = (archived or "false").lower()
    if arch == "only":
        if u.role not in ("admin", "manager", "supervisor"):
            raise HTTPException(403, "Архив харах эрхгүй")
        q = q.filter(PurchaseOrder.is_archived == True)
    elif arch == "true":
        if u.role not in ("admin", "manager", "supervisor"):
            # Admin/manager биш бол архивгүй л буцаана
            q = q.filter(PurchaseOrder.is_archived == False)
    else:
        q = q.filter(PurchaseOrder.is_archived == False)

    if status:
        q = q.filter(PurchaseOrder.status == status)
    if date_from:
        q = q.filter(PurchaseOrder.order_date >= date_from)
    if date_to:
        q = q.filter(PurchaseOrder.order_date <= date_to)

    orders = q.order_by(PurchaseOrder.order_date.desc()).all()
    if not orders:
        return []

    # SQL aggregate for line stats (avoids loading 100K+ lines into memory)
    order_ids = [o.id for o in orders]
    line_stats = {
        row[0]: {"line_count": row[1], "total_boxes": float(row[2] or 0), "total_weight": float(row[3] or 0)}
        for row in db.query(
            PurchaseOrderLine.purchase_order_id,
            func.count(PurchaseOrderLine.id),
            func.sum(PurchaseOrderLine.order_qty_box),
            func.sum(PurchaseOrderLine.computed_weight),
        ).filter(
            PurchaseOrderLine.purchase_order_id.in_(order_ids)
        ).group_by(PurchaseOrderLine.purchase_order_id).all()
    }

    # Bulk load users + vehicles
    user_ids = {o.created_by_user_id for o in orders if o.created_by_user_id}
    vehicle_ids = {o.vehicle_id for o in orders if o.vehicle_id}
    user_map = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}
    vehicle_map = {v.id: v for v in db.query(Vehicle).filter(Vehicle.id.in_(vehicle_ids)).all()} if vehicle_ids else {}

    result = []
    for o in orders:
        creator = user_map.get(o.created_by_user_id)
        vehicle = vehicle_map.get(o.vehicle_id)
        stats = line_stats.get(o.id, {"line_count": 0, "total_boxes": 0, "total_weight": 0})
        result.append({
            "id": o.id,
            "order_date": o.order_date.isoformat(),
            "status": o.status,
            "status_label": STATUS_LABEL.get(o.status, o.status),
            "created_by_username": creator.username if creator else "",
            "line_count": stats["line_count"],
            "total_boxes": round(stats["total_boxes"], 2),
            "total_weight": round(stats["total_weight"], 2),
            "created_at": o.created_at.isoformat() if o.created_at else None,
            "vehicle_id": o.vehicle_id,
            "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
            "notes": o.notes or "",
            "is_archived": bool(o.is_archived),
        })
    return result


@router.post("")
def create_purchase_order(
    body: POCreateIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("manager", "admin", "supervisor")),
):
    # Check master file exists
    if not MASTER_FILE.exists():
        raise HTTPException(
            400,
            "Master Excel файл байхгүй байна. Эхлээд Мастер нэгтгэл хийнэ үү."
        )

    # Note: Нэг өдөр олон захиалга зөвшөөрөгдсөн (duplicate date check устгасан)

    # Create order
    po = PurchaseOrder(
        order_date=body.order_date,
        status="preparing",
        created_by_user_id=u.id,
        notes=body.notes,
    )
    db.add(po)
    db.flush()

    # Load products. Brand filter байвал зөвхөн тэдгээр брендийн бараа л.
    tag_ids = parse_tag_ids(u.tag_ids)
    q = db.query(Product)
    if tag_ids:
        q = q.filter(Product.warehouse_tag_id.in_(tag_ids + [0]))
    selected_brands = [b.strip() for b in (body.brands or []) if b and b.strip()]
    if selected_brands:
        q = q.filter(Product.brand.in_(selected_brands))
    products = q.order_by(Product.brand, Product.item_code).all()

    for p in products:
        db.add(PurchaseOrderLine(
            purchase_order_id=po.id,
            product_id=p.id,
            order_qty_box=0.0,
            order_qty_pcs=0.0,
            computed_weight=0.0,
        ))

    db.commit()
    db.refresh(po)
    return {"id": po.id, "ok": True, "line_count": len(products)}


@router.get("/pdf-templates")
def get_pdf_templates_early(_=Depends(get_current_user)):
    """Alias placed before /{order_id} so the path is not consumed as an int param."""
    return PDF_TEMPLATES


@router.get("/shipments/by-status")
def list_shipments_by_status(
    status: str = Query(...),
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    """Бүх захиалгын shipment-уудыг shipment.status-аар шүүж, vehicle_id-аар групплэн буцаана.

    PO status биш SHIPMENT status-аар шүүдэг тул нэг захиалгын зарим машин
    замд явж байхад бусад нь ачигдсаар байж болно.
    Ижил machine дээр олон захиалгын shipment байвал нэг vehicle group дотор нэгтгэнэ.
    """
    # warehouse_clerk зөвхөн "arrived" харна
    if u.role == "warehouse_clerk" and status != "arrived":
        return []

    shipments = (
        db.query(POShipment)
        .filter(POShipment.status == status)
        .order_by(POShipment.created_at.desc())
        .all()
    )

    # Bulk load POs and Vehicles to avoid N+1
    po_ids = {sh.purchase_order_id for sh in shipments}
    vehicle_ids = {sh.vehicle_id for sh in shipments if sh.vehicle_id}
    po_map = {po.id: po for po in db.query(PurchaseOrder).filter(PurchaseOrder.id.in_(po_ids)).all()} if po_ids else {}
    vehicle_bulk = {v.id: v for v in db.query(Vehicle).filter(Vehicle.id.in_(vehicle_ids)).all()} if vehicle_ids else {}

    # Vehicle-аар групплэх
    vehicle_groups: dict[int | None, list] = {}
    for sh in shipments:
        po = po_map.get(sh.purchase_order_id)
        if not po:
            continue
        data = _serialize_shipment(sh, db)
        data["order_date"] = po.order_date.isoformat() if po.order_date else None
        data["order_status"] = po.status

        vid = sh.vehicle_id
        if vid not in vehicle_groups:
            vehicle_groups[vid] = []
        vehicle_groups[vid].append(data)

    result = []
    for vid, ship_list in vehicle_groups.items():
        vehicle = vehicle_bulk.get(vid) if vid else None
        all_brands: set[str] = set()
        order_ids: set[int] = set()
        total_loaded = 0.0
        total_received = 0.0
        total_weight = 0.0
        total_lines = 0
        for s in ship_list:
            total_loaded += s.get("total_loaded_box", 0)
            total_received += s.get("total_received_box", 0)
            total_weight += s.get("total_weight", 0)
            total_lines += s.get("line_count", 0)
            all_brands.update(s.get("brands", []))
            order_ids.add(s["purchase_order_id"])

        result.append({
            "vehicle_id": vid,
            "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
            "driver_name": vehicle.driver_name if vehicle else None,
            "shipments": ship_list,
            "shipment_count": len(ship_list),
            "order_count": len(order_ids),
            "order_ids": sorted(order_ids),
            "total_loaded_box": round(total_loaded, 1),
            "total_received_box": round(total_received, 1),
            "total_weight": round(total_weight, 1),
            "total_lines": total_lines,
            "brands": sorted(all_brands),
        })

    # Нийт жингээр буурахаар эрэмбэлэх
    result.sort(key=lambda x: x["total_weight"], reverse=True)
    return result


@router.get("/{order_id}")
def get_purchase_order(
    order_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    # Архивлагдсан захиалгыг зөвхөн admin/manager/supervisor л харна
    if po.is_archived and u.role not in ("admin", "manager", "supervisor"):
        raise HTTPException(403, "Архивлагдсан захиалгыг харах эрхгүй")

    # warehouse_clerk can only see preparing and arrived orders
    if u.role == "warehouse_clerk" and po.status not in ("preparing", "arrived"):
        raise HTTPException(403, "Энэ захиалгыг харах эрх байхгүй")

    # warehouse_clerk sees only products from their assigned warehouses
    filter_tag_ids = parse_tag_ids(u.tag_ids) if u.role == "warehouse_clerk" else None
    return _serialize_order_detail(po, db, filter_tag_ids=filter_tag_ids)


@router.patch("/{order_id}/status")
def advance_status(
    order_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    next_st = _next_status(po.status)
    if not next_st:
        raise HTTPException(400, "Захиалга эцсийн статуст хүрсэн байна")

    # accountant: accounting → confirmed, confirmed → received
    allowed_roles = ["manager", "supervisor", "admin"]
    if po.status in ("accounting", "confirmed"):
        allowed_roles.append("accountant")
    if u.role not in allowed_roles:
        raise HTTPException(403, "Энэ үйлдлийг хийх эрх байхгүй")

    # When advancing arrived → accounting, pre-fill unit_price from last_purchase_price
    if po.status == "arrived" and next_st == "accounting":
        product_ids = [l.product_id for l in po.lines]
        products = db.query(Product).filter(Product.id.in_(product_ids)).all()
        pmap = {p.id: p for p in products}
        for line in po.lines:
            p = pmap.get(line.product_id)
            if p and (line.unit_price or 0) == 0.0 and (p.last_purchase_price or 0) > 0:
                line.unit_price = float(p.last_purchase_price)

    po.status = next_st

    # Batch advance all brand statuses that are at the old PO status
    old_st = STATUS_SEQUENCE[STATUS_SEQUENCE.index(next_st) - 1]
    brand_rows = db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id,
        PurchaseOrderBrandStatus.status == old_st,
    ).all()
    for bs in brand_rows:
        bs.status = next_st

    db.commit()
    return {
        "ok": True,
        "new_status": next_st,
        "new_status_label": STATUS_LABEL.get(next_st, next_st),
    }


class ForceStatusIn(BaseModel):
    status: str


class ArchiveIn(BaseModel):
    archived: bool = True


@router.patch("/{order_id}/archive")
def archive_order(
    order_id: int,
    body: ArchiveIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "supervisor")),
):
    """Захиалгыг архивлах / архиваас буцаах. Admin, manager, supervisor."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    po.is_archived = bool(body.archived)
    db.commit()
    return {"ok": True, "is_archived": po.is_archived}

@router.patch("/{order_id}/force-status")
def force_status(
    order_id: int,
    body: ForceStatusIn,
    brand: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin")),
):
    """
    Статус албадан өөрчлөх.
    - brand параметргүй: бүх бренд + PO status шинэчилнэ.
    - brand='X': зөвхөн тухайн брендийн статус өөрчилнэ, PO status нь brand-уудын минимумаас тооцоологдоно.
    """
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    if body.status not in STATUS_SEQUENCE:
        raise HTTPException(400, "Буруу статус")

    if brand:
        bs = db.query(PurchaseOrderBrandStatus).filter(
            PurchaseOrderBrandStatus.purchase_order_id == order_id,
            PurchaseOrderBrandStatus.brand == brand,
        ).first()
        if not bs:
            raise HTTPException(404, f"'{brand}' брендийн статус олдсонгүй")
        bs.status = body.status
        db.commit()
        _sync_po_status_from_brands(order_id, db)
        return {
            "ok": True,
            "brand": brand,
            "new_status": body.status,
            "new_status_label": STATUS_LABEL.get(body.status, body.status),
            "po_status": po.status,
        }

    # Бүх бренд + PO status нэг утгад шилжүүлэх
    po.status = body.status
    for bs in db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id
    ).all():
        bs.status = body.status
    db.commit()
    return {"ok": True, "new_status": po.status, "new_status_label": STATUS_LABEL.get(po.status, po.status)}


@router.post("/{order_id}/set-lines")
def set_lines(
    order_id: int,
    lines: List[POLineIn],
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    # warehouse_clerk — preparing + arrived; accountant — accounting; admin — all editable;
    # manager/supervisor — preparing, reviewing, loading + accounting (for unit_price editing)
    if u.role == "warehouse_clerk":
        allowed_statuses = ["preparing", "arrived"]
    elif u.role == "accountant":
        allowed_statuses = ["accounting"]
    elif u.role == "admin":
        allowed_statuses = ["preparing", "reviewing", "loading", "arrived", "accounting"]
    else:
        allowed_statuses = ["preparing", "reviewing", "loading", "accounting"]
    # Per-brand status lookup
    brand_status_map: dict[str, str] = {}
    for bs in db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id
    ).all():
        brand_status_map[bs.brand] = bs.status

    # PO-level check (fallback for backward compat)
    if po.status not in allowed_statuses:
        # Per-brand: any brand in an allowed status?
        has_allowed_brand = any(s in allowed_statuses for s in brand_status_map.values())
        if not has_allowed_brand:
            raise HTTPException(400, "Энэ статуст тоо өөрчлөх боломжгүй")

    # Permission check
    if u.role not in ("manager", "warehouse_clerk", "admin", "supervisor", "accountant"):
        raise HTTPException(403, "Энэ үйлдлийг хийх эрх байхгүй")

    # warehouse_clerk restricted to their assigned warehouses
    clerk_tag_ids = parse_tag_ids(u.tag_ids) if u.role == "warehouse_clerk" else None

    # Build lookup map for existing lines
    line_map = {l.product_id: l for l in po.lines}

    for li in lines:
        if li.product_id not in line_map:
            continue
        p = db.query(Product).filter(Product.id == li.product_id).first()
        if not p:
            continue
        # Clerk cannot update products outside their warehouses.
        if clerk_tag_ids and p.warehouse_tag_id != 0 and p.warehouse_tag_id not in clerk_tag_ids:
            continue
        line = line_map[li.product_id]

        # Use per-brand status if available, fallback to PO status
        effective_st = brand_status_map.get(p.brand, po.status)

        if effective_st == "arrived":
            if li.received_qty_box is not None:
                line.received_qty_box = float(li.received_qty_box)
            if li.received_qty_extra_pcs is not None:
                line.received_qty_extra_pcs = float(li.received_qty_extra_pcs)
            if li.remark is not None:
                line.line_remark = li.remark
        elif effective_st == "accounting":
            if li.unit_price is not None:
                line.unit_price = float(li.unit_price)
            if li.received_qty_box is not None:
                line.received_qty_box = float(li.received_qty_box)
            if li.received_qty_extra_pcs is not None:
                line.received_qty_extra_pcs = float(li.received_qty_extra_pcs)
            if li.remark is not None:
                line.line_remark = li.remark
        else:
            # preparing / reviewing / loading — update order qty + derived fields
            qty_box = float(li.order_qty_box or 0)
            qty_pcs = qty_box * float(p.pack_ratio or 1)
            weight = qty_pcs * float(p.unit_weight or 0)
            line.order_qty_box = qty_box
            line.order_qty_pcs = qty_pcs
            line.computed_weight = weight
            if effective_st == "loading":
                if li.supplier_qty_box is not None:
                    line.supplier_qty_box = float(li.supplier_qty_box)
                if li.loaded_qty_box is not None:
                    new_loaded = float(li.loaded_qty_box)
                    line.loaded_qty_box = new_loaded
                    # Хэрэв энэ PO line нь ЯГ НЭГ shipment line дээр байвал тэр дээр ч хадгална
                    # (харин олон хуваагдсан бол shipment detail UI ашиглана)
                    sh_lines = db.query(POShipmentLine).filter(
                        POShipmentLine.po_line_id == line.id
                    ).all()
                    if len(sh_lines) == 1:
                        sh_lines[0].loaded_qty_box = new_loaded

    db.commit()

    # Ensure brand status records exist for any new brands
    _ensure_brand_statuses(order_id, db)
    db.commit()

    return {"ok": True}


@router.delete("/{order_id}")
def delete_order(
    order_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    # Delete brand-vehicle assignments first (no cascade on this model)
    db.query(PurchaseOrderBrandVehicle).filter(
        PurchaseOrderBrandVehicle.purchase_order_id == order_id
    ).delete(synchronize_session=False)
    db.delete(po)
    db.commit()
    return {"ok": True}


@router.delete("/{order_id}/lines/{line_id}")
def delete_line(
    order_id: int,
    line_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("manager", "admin", "supervisor", "warehouse_clerk")),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    line = db.query(PurchaseOrderLine).filter(
        PurchaseOrderLine.id == line_id,
        PurchaseOrderLine.purchase_order_id == order_id,
    ).first()
    if not line:
        raise HTTPException(404, "Мөр олдсонгүй")

    # Эрх шалгах: тухайн line-ий brand status-г харгалзана
    product = db.query(Product).filter(Product.id == line.product_id).first()
    brand = product.brand if product else None
    bs = None
    if brand:
        bs = db.query(PurchaseOrderBrandStatus).filter(
            PurchaseOrderBrandStatus.purchase_order_id == order_id,
            PurchaseOrderBrandStatus.brand == brand,
        ).first()
    effective_st = bs.status if bs else po.status

    # warehouse_clerk-д зөвхөн arrived (Ачаа ирсэн) status-д устгах эрх
    # бусад role-д зөвхөн loading status-д устгах эрх
    if u.role == "warehouse_clerk":
        if effective_st != "arrived":
            raise HTTPException(400, "Нярав зөвхөн 'Ачаа ирсэн' статуст мөр устгана")
    else:
        if effective_st not in ("loading", "arrived"):
            raise HTTPException(400, "Зөвхөн 'Ачигдаж байна' болон 'Ачаа ирсэн' статуст мөр устгана")

    # Холбоотой shipment line-уудыг эхлээд устгах (FK clean)
    db.query(POShipmentLine).filter(POShipmentLine.po_line_id == line_id).delete()
    db.delete(line)
    db.commit()
    return {"ok": True}


@router.post("/{order_id}/add-line")
def add_line(
    order_id: int,
    body: AddLineIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("manager", "admin", "supervisor")),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    if po.status not in ("preparing", "loading"):
        raise HTTPException(400, "Энэ статуст бараа нэмэх боломжгүй")
    p = db.query(Product).filter(Product.id == body.product_id).first()
    if not p:
        raise HTTPException(404, "Бараа олдсонгүй")
    existing = next((l for l in po.lines if l.product_id == body.product_id), None)
    # override_brand шалгалт: зөвхөн admin
    override_brand = (body.override_brand or "").strip()
    if override_brand and u.role != "admin":
        raise HTTPException(403, "Бусад брендийн бараа нэмэх эрх зөвхөн админ эрхтэй")
    # override тохиолдолд барааны оригинал бренд override_brand-тэй адил байвал утгагүй
    if override_brand and (p.brand or "") == override_brand:
        override_brand = ""
    if existing:
        # Override-той бол давхар мөр зөвшөөрөх боломжтой (нэг бараа давхар брендэд)
        if not override_brand and not (existing.override_brand or ""):
            raise HTTPException(400, "Бараа аль хэдийн нэмэгдсэн байна")
        if override_brand and (existing.override_brand or "") == override_brand:
            raise HTTPException(400, f"Бараа {override_brand} брендэд аль хэдийн нэмэгдсэн байна")
    qty_box = float(body.order_qty_box or 0)
    qty_pcs = qty_box * float(p.pack_ratio or 1)
    weight = qty_pcs * float(p.unit_weight or 0)
    line = PurchaseOrderLine(
        purchase_order_id=order_id,
        product_id=body.product_id,
        order_qty_box=qty_box,
        order_qty_pcs=qty_pcs,
        computed_weight=weight,
        override_brand=override_brand,
    )
    db.add(line)
    db.commit()
    # Override бренд нэмсэн бол тухайн брендэд brand_status үүсгэнэ
    if override_brand:
        _ensure_brand_statuses(order_id, db)
        db.commit()
    return {"ok": True}


class BrandVehicleIn(BaseModel):
    brand: str
    vehicle_id: Optional[int] = None


@router.post("/{order_id}/brand-vehicles")
def set_brand_vehicles(
    order_id: int,
    items: List[BrandVehicleIn],
    db: Session = Depends(get_db),
    u: User = Depends(require_role("manager", "admin", "supervisor")),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    for item in items:
        existing = db.query(PurchaseOrderBrandVehicle).filter(
            PurchaseOrderBrandVehicle.purchase_order_id == order_id,
            PurchaseOrderBrandVehicle.brand == item.brand,
        ).first()
        if item.vehicle_id is None:
            if existing:
                db.delete(existing)
        elif existing:
            existing.vehicle_id = item.vehicle_id
        else:
            db.add(PurchaseOrderBrandVehicle(
                purchase_order_id=order_id,
                brand=item.brand,
                vehicle_id=item.vehicle_id,
            ))
    db.commit()
    return {"ok": True}


@router.post("/{order_id}/revert")
def revert_to_arrived(
    order_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("accountant", "supervisor", "admin")),
):
    """Нягтлан шалгаж байна → Ачаа ирсэн руу буцаана."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    if po.status != "accounting":
        raise HTTPException(400, "Зөвхөн 'Нягтлан шалгаж байна' статусаас буцаах боломжтой")
    po.status = "arrived"
    db.commit()
    return {"ok": True, "new_status": "arrived", "new_status_label": STATUS_LABEL["arrived"]}


@router.get("/{order_id}/export-excel")
def export_excel(
    order_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("accountant", "supervisor", "admin")),
):
    """Нягтлан Баталгаажсан — Excel файл татах."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    # Load only lines with order_qty > 0 (NOT lazy po.lines)
    export_lines = db.query(PurchaseOrderLine).filter(
        PurchaseOrderLine.purchase_order_id == po.id,
        PurchaseOrderLine.order_qty_box > 0,
    ).all()
    product_ids = [l.product_id for l in export_lines]
    products = db.query(Product).filter(Product.id.in_(product_ids)).all() if product_ids else []
    product_map = {p.id: p for p in products}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Захиалга {po.order_date}"

    # Header row
    headers = [
        "Брэнд", "Код", "Нэр", "Агуулах",
        "Захиалах тоо", "Нийлүүлэгч бэлдсэн", "Ачигдсан тоо", "Ирсэн тоо", "Зөрүү"
    ]
    header_fill = PatternFill("solid", fgColor="3258A0")
    header_font = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows sorted by brand, item_code
    lines_data = []
    for l in export_lines:
        p = product_map.get(l.product_id)
        if not p:
            continue
        diff = round((l.loaded_qty_box or 0) - (l.received_qty_box or 0), 2)
        lines_data.append([
            p.brand or "", p.item_code, p.name, p.warehouse_name or "",
            l.order_qty_box, l.supplier_qty_box, l.loaded_qty_box, l.received_qty_box, diff,
        ])
    lines_data.sort(key=lambda x: (x[0], x[1]))

    for ri, row in enumerate(lines_data, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    # Column widths
    for ci, w in enumerate([18, 14, 35, 14, 14, 18, 14, 12, 10], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"order_{po.order_date.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── ERP Import Excel Export ────────────────────────────────────────────────────

class ERPExcelConfigIn(BaseModel):
    company: str                     # "buten_orgil" | "orgil_khorum"
    date: str                        # "YYYY-MM-DD"
    document_note: str = ""          # Гүйлгээний утга
    related_account: str = ""        # Харьцсан данс (e.g. 310101)
    account: str = ""                # Данс (e.g. 150101)
    warehouse_map: dict = {}         # buten_orgil: {warehouse_name: erp_location_code}
    single_location: str = ""        # orgil_khorum: single location code
    brand_filter: str = ""           # Тодорхой бренд шүүх (хоосон = бүгд)


@router.post("/{order_id}/export-erp-excel")
def export_erp_excel(
    order_id: int,
    body: ERPExcelConfigIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "manager", "accountant", "supervisor")),
):
    """ERP-д импортлох Excel файл үүсгэх (confirmed статус)."""
    try:
        return _export_erp_excel_impl(order_id, body, db)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[export_erp_excel] ERROR: {e}\n{tb}", flush=True)
        raise HTTPException(500, f"Excel үүсгэхэд алдаа: {type(e).__name__}: {e}")


def _export_erp_excel_impl(order_id: int, body: "ERPExcelConfigIn", db: Session):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt_cls

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    # Shipment-аас aggregated loaded/received тоог бэлдэнэ (зарим PO дээр values нь зөвхөн
    # POShipmentLine дээр байдаг, PurchaseOrderLine-д 0 байдаг).
    from sqlalchemy import func as _func
    from sqlalchemy import or_ as _or_q
    sh_agg = dict(
        (row[0], {"loaded": float(row[1] or 0), "received": float(row[2] or 0)})
        for row in db.query(
            POShipmentLine.po_line_id,
            _func.sum(POShipmentLine.loaded_qty_box),
            _func.sum(POShipmentLine.received_qty_box),
        )
        .join(PurchaseOrderLine, PurchaseOrderLine.id == POShipmentLine.po_line_id)
        .filter(PurchaseOrderLine.purchase_order_id == po.id)
        .group_by(POShipmentLine.po_line_id)
        .all()
    )

    # Идэвхтэй line-ууд (order>0 эсвэл supplier>0). Received/loaded-оор дараа шүүнэ.
    recv_lines = db.query(PurchaseOrderLine).filter(
        PurchaseOrderLine.purchase_order_id == po.id,
        _or_q(PurchaseOrderLine.order_qty_box > 0, PurchaseOrderLine.supplier_qty_box > 0),
    ).all()
    # Filter: дор хаяж нэг хэмжээ (received эсвэл loaded) > 0 байх
    def _line_has_qty(l):
        agg = sh_agg.get(l.id, {})
        return (
            (l.received_qty_box or 0) > 0
            or (l.received_qty_extra_pcs or 0) > 0
            or (l.loaded_qty_box or 0) > 0
            or agg.get("received", 0) > 0
            or agg.get("loaded", 0) > 0
        )
    recv_lines = [l for l in recv_lines if _line_has_qty(l)]
    product_ids = [l.product_id for l in recv_lines]
    products = db.query(Product).filter(Product.id.in_(product_ids)).all() if product_ids else []
    product_map = {p.id: p for p in products}

    # Parse date
    try:
        date_val = dt_cls.strptime(body.date, "%Y-%m-%d").date()
    except Exception:
        date_val = po.order_date

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"

    # ── Column headers (row 1) ──
    col_headers = [
        "Огноо", "Баримтын дугаар", "Гүйлгээний утга", "Харилцагч",
        "Харьцсан данс", "Харьцсан ялгаатай харилцагч",
        "НӨАТ тай эсэх", "НӨАТ-н үзүүлэлт", "НӨАТ автоматаар бодох эсэх", "НӨАТ-н дүн",
        "НХАТ тай эсэх", "НХАТ автоматаар бодох эсэх", "НХАТ-н дүн",
        "Данс", "Бараа материал", "Барааны байршил",
        "Тоо хэмжээ", "Нэгж үнэ", "Хувийн жин", "Нийт дүн",
        "НӨАТ тооцох эсэх", "НХАТ тооцох эсэх", "НХАТ мөр",
    ]
    hdr_fill = PatternFill("solid", fgColor="3258A0")
    hdr_font = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    # ── Build data rows ──
    # Дээр recv_lines-ыг аль хэдийн filter хийсэн. Mөн бренд шүүлтүүр хэрэглэнэ.
    brand_filter = (body.brand_filter or "").strip()
    erp_lines = recv_lines
    valid = []
    for line in erp_lines:
        agg = sh_agg.get(line.id, {})
        extra_pcs = float(line.received_qty_extra_pcs or 0)
        # Effective received: PO line-ийн received > 0 бол түүнийг, эс бол shipment aggregated received.
        # Хэрэглэгч "0 хайрцаг + N ширхэг" гэж санаатайгаар оруулсан бол (extra_pcs > 0 бол)
        # loaded-оор fallback ХИЙХГҮЙ — хэрэглэгчийн оруулсан тоог нэн тэргүүнд тоолно.
        eff_received_box = float(line.received_qty_box or 0)
        if eff_received_box == 0:
            eff_received_box = float(agg.get("received", 0))
        if eff_received_box == 0 and extra_pcs == 0:
            # received + extra хоёр хоосон үед л loaded-аас fallback авна (автомат pre-fill)
            eff_received_box = float(line.loaded_qty_box or 0)
            if eff_received_box == 0:
                eff_received_box = float(agg.get("loaded", 0))
        if eff_received_box <= 0 and extra_pcs <= 0:
            continue
        if not ((line.order_qty_box or 0) > 0 or (line.supplier_qty_box or 0) > 0):
            continue
        p = product_map.get(line.product_id)
        if not p:
            continue
        # Brand filter — effective brand (override_brand-ийг түрүүнд)
        eff_brand = (line.override_brand or "").strip() or (p.brand or "")
        if brand_filter and eff_brand != brand_filter:
            continue
        if body.company == "orgil_khorum":
            location = body.single_location
        else:
            location = body.warehouse_map.get(p.warehouse_name, "")
        # ERP импортод тоо хэмжээ нь ширхгээр орох ёстой (бид хайрцгаар явдаг)
        # Ачаа ирсэн үед задгай ширхэг нэмж болно (жишээ: 4 хайрцаг + 2 ширхэг)
        # unit_price нь Орлого тайлангаас авсан ширхэгийн үнэ
        # Хэрэв received_qty_box=0 бол loaded_qty_box-г ашиглана (fallback).
        pack_ratio = float(p.pack_ratio or 1) or 1.0
        qty = eff_received_box * pack_ratio + extra_pcs  # нийт ширхэг
        price = float(line.unit_price or 0)              # нэгж үнэ/ширхэг
        total = round(qty * price, 2)                     # нийт дүн
        # Override брендтэй бол тухайн брендийн brand_code-ыг ашиглана (supplier солигдсон)
        if (line.override_brand or "").strip():
            ref = db.query(Product).filter(Product.brand == eff_brand, Product.brand_code != None).first()
            eff_brand_code = ref.brand_code if ref else (p.brand_code or "")
        else:
            eff_brand_code = p.brand_code or ""
        valid.append((eff_brand, p.item_code, p, line, location, qty, price, total, eff_brand_code))

    # Sort by brand_code then item_code so same-supplier items are grouped
    valid.sort(key=lambda x: (x[8] or "", x[0], x[1]))

    # ── Group by brand_code — one ERP document block per supplier ──
    from collections import defaultdict
    groups: dict = defaultdict(list)
    for item in valid:
        groups[item[8] or ""].append(item)

    current_row = 2
    for supplier_code, items in groups.items():
        for i, (brand, item_code, p, line, location, qty, price, total, _eff_bc) in enumerate(items):
            is_first = (i == 0)
            row = [
                date_val if is_first else None,            # Огноо
                None,                                       # Баримтын дугаар
                body.document_note if is_first else None,  # Гүйлгээний утга
                supplier_code if is_first else None,                    # Харилцагч (brand_code-оос)
                body.related_account if is_first else None, None,      # Харьцсан данс, Харьцсан ялгаатай
                0 if is_first else None,                   # НӨАТ тай эсэх
                None, None,                                 # НӨАТ-н үзүүлэлт, автоматаар
                0 if is_first else None,                   # НӨАТ-н дүн
                0 if is_first else None,                   # НХАТ тай эсэх
                None,                                       # НХАТ автоматаар
                0 if is_first else None,                   # НХАТ-н дүн
                body.account,                               # Данс*
                p.item_code,                                # Бараа материал*
                location,                                   # Барааны байршил*
                qty,                                        # Тоо хэмжээ*
                price,                                      # Нэгж үнэ*
                1.0,                                        # Хувийн жин* (тогтмол)
                total,                                      # Нийт дүн*
                0,                                          # НӨАТ тооцох эсэх*
                0,                                          # НХАТ тооцох эсэх*
                None,                                       # НХАТ мөр
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(row=current_row, column=ci, value=val)
            current_row += 1

    # ── Column widths ──
    widths = [14, 16, 28, 14, 14, 20, 14, 18, 22, 12, 14, 22, 12,
              16, 18, 20, 12, 12, 12, 14, 18, 18, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import re
    from urllib.parse import quote
    date_str = po.order_date.strftime("%Y%m%d")
    brand_part = re.sub(r'[\\/:*?"<>|]', '_', brand_filter) if brand_filter else "all"
    filename = f"{date_str}_PO{po.id}_{brand_part}.xlsx"
    # RFC 5987: ASCII fallback + UTF-8 encoded filename* (Cyrillic-д зориулж)
    # ASCII fallback нь зай/тусгай тэмдэггүй учир хашилт хэрэггүй
    ascii_fallback = re.sub(r"[^\w\-.]", "_", filename.encode("ascii", "ignore").decode("ascii")) or f"PO{po.id}.xlsx"
    utf8_quoted = quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename={ascii_fallback}; filename*=UTF-8''{utf8_quoted}"
            )
        },
    )


# ── PDF Export ─────────────────────────────────────────────────────────────────

FONT_REGULAR = "C:/Windows/Fonts/arial.ttf"
FONT_BOLD    = "C:/Windows/Fonts/arialbd.ttf"

# Default template data (user edits these in the modal)
PDF_TEMPLATES = {
    "buten_orgil": {
        "company_name": "Бүтэн-Оргил ХХК",
        "address": "Улаанбаатар хот, ...",
        "phone": "...",
        "truck_location": "",
        "driver": "",
    },
    "orgil_khorum": {
        "company_name": "Оргил-Хорум ХХК",
        "address": "Улаанбаатар хот, ...",
        "phone": "...",
        "truck_location": "",
        "driver": "",
    },
}


class PDFHeaderIn(BaseModel):
    company_name: str
    address: str = ""
    phone: str = ""
    truck_location: str = ""
    driver: str = ""
    extra_note: str = ""
    brand_filter: str = ""  # хоосон бол бүх брэнд


def _build_pdf(po: PurchaseOrder, body: PDFHeaderIn, db: Session) -> bytes:
    from fpdf import FPDF
    from datetime import datetime

    generated_at = datetime.now().strftime("%Y/%m/%d %H:%M")

    # Gather lines with qty > 0, grouped by brand (SQL filter instead of lazy load)
    active_lines = db.query(PurchaseOrderLine).filter(
        PurchaseOrderLine.purchase_order_id == po.id,
        PurchaseOrderLine.order_qty_box > 0,
    ).all()
    product_ids = [l.product_id for l in active_lines]
    products = db.query(Product).filter(Product.id.in_(product_ids)).all() if product_ids else []
    product_map = {p.id: p for p in products}

    grouped: dict[str, list[dict]] = {}
    total_boxes = 0.0
    total_weight = 0.0
    brand_filter = (body.brand_filter or "").strip()
    for l in active_lines:
        if l.order_qty_box <= 0:
            continue
        p = product_map.get(l.product_id)
        if not p:
            continue
        brand_override = (l.override_brand or "").strip()
        brand_raw = brand_override or (p.brand or "").strip()
        brand = brand_raw if brand_raw and brand_raw.lower() != "nan" else "Брэнд байхгүй"
        # брэнд шүүлт хэрэглэгдсэн бол зөвхөн тухайн брэнд
        if brand_filter and brand != brand_filter:
            continue
        grouped.setdefault(brand, []).append({
            "item_code": p.item_code,
            "name": p.name,
            "warehouse_name": p.warehouse_name or "",
            "qty_box": l.order_qty_box,
            "qty_pcs": l.order_qty_pcs,
            "weight": l.computed_weight,
        })
        total_boxes += l.order_qty_box
        total_weight += l.computed_weight

    # Extra lines — same brand group as regular lines
    for el in po.extra_lines:
        if el.qty_box <= 0:
            continue
        el_brand = (el.brand or "").strip() or "Нэмэлт бараа"
        if brand_filter and el_brand != brand_filter:
            continue
        grouped.setdefault(el_brand, []).append({
            "item_code": el.item_code or "—",
            "name": f"★ {el.name}",
            "warehouse_name": el.warehouse_name,
            "qty_box": el.qty_box,
            "qty_pcs": el.qty_box * el.pack_ratio,
            "weight": el.computed_weight,
        })
        total_boxes += el.qty_box
        total_weight += el.computed_weight

    # Sort brands and lines within each brand
    sorted_brands = sorted(grouped.keys())
    for brand in sorted_brands:
        grouped[brand].sort(key=lambda x: x["item_code"])

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_font("Arial", fname=FONT_REGULAR)
    pdf.add_font("Arial", style="B", fname=FONT_BOLD)
    pdf.set_margins(12, 12, 12)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Page usable width: 210 - 24 = 186mm
    W = pdf.w - pdf.l_margin - pdf.r_margin  # ~186

    # ── Company header ──────────────────────────────────────────
    pdf.set_font("Arial", style="B", size=15)
    pdf.cell(0, 9, body.company_name, new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.set_font("Arial", size=9)
    if body.address:
        pdf.cell(0, 5, f"Хаяг: {body.address}", new_x="LMARGIN", new_y="NEXT", align="C")
    if body.phone:
        pdf.cell(0, 5, f"Утас: {body.phone}", new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.ln(2)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(3)

    # ── Order meta ──────────────────────────────────────────────
    pdf.set_font("Arial", style="B", size=10)
    date_str = po.order_date.strftime("%Y/%m/%d")
    pdf.cell(W / 2, 6, f"Захиалга: {date_str}", new_x="RIGHT", new_y="TOP")
    pdf.cell(W / 2, 6, f"Статус: {STATUS_LABEL.get(po.status, po.status)}", new_x="LMARGIN", new_y="NEXT", align="R")

    pdf.set_font("Arial", size=9)
    if body.truck_location:
        pdf.cell(0, 5, f"Ачигдах байршил: {body.truck_location}", new_x="LMARGIN", new_y="NEXT")
    if body.driver:
        pdf.cell(0, 5, f"Жолооч / Машин: {body.driver}", new_x="LMARGIN", new_y="NEXT")
    if body.extra_note:
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 5, f"Тэмдэглэл: {body.extra_note}", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)

    # ── Table ───────────────────────────────────────────────────
    # Column widths (total = W ~186)
    COL_CODE = 22
    COL_NAME = 68
    COL_WH   = 30
    COL_BOX  = 20
    COL_PCS  = 18
    COL_KG   = W - COL_CODE - COL_NAME - COL_WH - COL_BOX - COL_PCS  # remainder

    def tbl_header():
        pdf.set_font("Arial", style="B", size=8)
        pdf.set_fill_color(50, 90, 160)
        pdf.set_text_color(255, 255, 255)
        ROW_H = 7
        pdf.cell(COL_CODE, ROW_H, "Код",      border=0, align="C", fill=True)
        pdf.cell(COL_NAME, ROW_H, "Нэр",      border=0, align="L", fill=True)
        pdf.cell(COL_WH,   ROW_H, "Агуулах",  border=0, align="C", fill=True)
        pdf.cell(COL_BOX,  ROW_H, "Хайрцаг",  border=0, align="C", fill=True)
        pdf.cell(COL_PCS,  ROW_H, "Ш",        border=0, align="C", fill=True)
        pdf.cell(COL_KG,   ROW_H, "Жин (кг)", border=0, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)

    tbl_header()
    ROW_H = 6

    for bi, brand in enumerate(sorted_brands):
        lines = grouped[brand]
        # Brand row
        pdf.set_font("Arial", style="B", size=8)
        pdf.set_fill_color(220, 230, 245)
        pdf.cell(W, ROW_H, f"  {brand}", border=0, align="L", fill=True, new_x="LMARGIN", new_y="NEXT")

        # Product rows
        pdf.set_font("Arial", size=8)
        for ri, row in enumerate(lines):
            fill = (ri % 2 == 1)
            if fill:
                pdf.set_fill_color(245, 247, 252)
            else:
                pdf.set_fill_color(255, 255, 255)

            pdf.cell(COL_CODE, ROW_H, row["item_code"],             border=0, align="C", fill=True)
            pdf.cell(COL_NAME, ROW_H, row["name"],                  border=0, align="L", fill=True)
            pdf.cell(COL_WH,   ROW_H, row["warehouse_name"],        border=0, align="C", fill=True)
            pdf.cell(COL_BOX,  ROW_H, f"{row['qty_box']:.0f}",      border=0, align="C", fill=True)
            pdf.cell(COL_PCS,  ROW_H, f"{row['qty_pcs']:.0f}",      border=0, align="C", fill=True)
            pdf.cell(COL_KG,   ROW_H, f"{row['weight']:.2f}",       border=0, align="R", fill=True, new_x="LMARGIN", new_y="NEXT")

    # ── Totals row ──────────────────────────────────────────────
    pdf.ln(1)
    pdf.set_font("Arial", style="B", size=9)
    pdf.set_fill_color(50, 90, 160)
    pdf.set_text_color(255, 255, 255)
    label_w = COL_CODE + COL_NAME + COL_WH
    pdf.cell(label_w, 7, "Нийт дүн", border=0, align="R", fill=True)
    pdf.cell(COL_BOX, 7, f"{total_boxes:.0f}", border=0, align="C", fill=True)
    pdf.cell(COL_PCS, 7, "", border=0, fill=True)
    pdf.cell(COL_KG,  7, f"{total_weight:.2f}", border=0, align="R", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)

    # ── Timestamp footer (last line of content) ─────────────────
    pdf.ln(3)
    pdf.set_font("Arial", size=7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f"Бэлдсэн: {generated_at}", align="R")
    pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output())


# ── Extra lines (supplier-added items not in product catalog) ──────────────────

class ExtraLineIn(BaseModel):
    brand: str = ""
    name: str
    item_code: str = ""
    warehouse_name: str = ""
    unit_weight: float = 0.0
    pack_ratio: float = 1.0
    qty_box: float = 0.0


def _extra_line_access(order_id: int, db: Session, u: User) -> PurchaseOrder:
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    if po.status != "loading":
        raise HTTPException(400, "Зөвхөн 'Ачигдаж байна' статуст нэмэлт мөр засах боломжтой")
    if u.role not in ("manager", "admin", "supervisor"):
        raise HTTPException(403, "Эрх хүрэлцэхгүй")
    return po


@router.post("/{order_id}/extra-lines")
def add_extra_line(
    order_id: int,
    body: ExtraLineIn,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = _extra_line_access(order_id, db, u)
    el = OrderExtraLine(
        purchase_order_id=po.id,
        brand=body.brand.strip(),
        name=body.name.strip(),
        item_code=body.item_code.strip(),
        warehouse_name=body.warehouse_name.strip(),
        unit_weight=body.unit_weight,
        pack_ratio=body.pack_ratio,
        qty_box=body.qty_box,
        computed_weight=round(body.qty_box * body.pack_ratio * body.unit_weight, 4),
    )
    db.add(el)
    db.commit()
    db.refresh(el)
    return {"id": el.id, "computed_weight": el.computed_weight}


@router.put("/{order_id}/extra-lines/{extra_id}")
def update_extra_line(
    order_id: int,
    extra_id: int,
    body: ExtraLineIn,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = _extra_line_access(order_id, db, u)
    el = db.query(OrderExtraLine).filter(
        OrderExtraLine.id == extra_id,
        OrderExtraLine.purchase_order_id == po.id,
    ).first()
    if not el:
        raise HTTPException(404, "Мөр олдсонгүй")
    el.brand = body.brand.strip()
    el.name = body.name.strip()
    el.item_code = body.item_code.strip()
    el.warehouse_name = body.warehouse_name.strip()
    el.unit_weight = body.unit_weight
    el.pack_ratio = body.pack_ratio
    el.qty_box = body.qty_box
    el.computed_weight = round(body.qty_box * body.pack_ratio * body.unit_weight, 4)
    db.commit()
    return {"id": el.id, "computed_weight": el.computed_weight}


@router.delete("/{order_id}/extra-lines/{extra_id}")
def delete_extra_line(
    order_id: int,
    extra_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = _extra_line_access(order_id, db, u)
    el = db.query(OrderExtraLine).filter(
        OrderExtraLine.id == extra_id,
        OrderExtraLine.purchase_order_id == po.id,
    ).first()
    if not el:
        raise HTTPException(404, "Мөр олдсонгүй")
    db.delete(el)
    db.commit()
    return {"ok": True}


@router.post("/{order_id}/export-pdf")
def export_pdf(
    order_id: int,
    body: PDFHeaderIn,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")

    pdf_bytes = _build_pdf(po, body, db)
    filename = f"order_{po.order_date.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# Shipment (машинаар ачилт) endpoints
# ══════════════════════════════════════════════════════════════════════════════

SHIPMENT_STATUS_SEQ = ["loading", "transit", "arrived", "accounting", "confirmed", "received"]
SHIPMENT_STATUS_LABEL = {
    "loading":    "Ачигдаж байна",
    "transit":    "Замд явж байна",
    "arrived":    "Ачаа ирсэн",
    "accounting": "Нягтлан шалгаж байна",
    "confirmed":  "Баталгаажсан",
    "received":   "Орлого авагдсан",
}


def _serialize_shipment(sh: POShipment, db: Session) -> dict:
    vehicle = db.query(Vehicle).filter(Vehicle.id == sh.vehicle_id).first() if sh.vehicle_id else None
    lines = db.query(POShipmentLine).filter(POShipmentLine.shipment_id == sh.id).all()

    total_loaded = sum(sl.loaded_qty_box for sl in lines)
    total_received = sum(sl.received_qty_box for sl in lines)

    # Compute total weight from PO lines
    po_line_ids = [sl.po_line_id for sl in lines]
    total_weight = 0.0
    brands = set()
    if po_line_ids:
        po_lines = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.id.in_(po_line_ids)).all()
        po_line_map = {pl.id: pl for pl in po_lines}
        prod_ids = [pl.product_id for pl in po_lines]
        prods = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()}
        for sl in lines:
            pl = po_line_map.get(sl.po_line_id)
            if not pl:
                continue
            p = prods.get(pl.product_id)
            if p:
                total_weight += sl.loaded_qty_box * float(p.pack_ratio or 1) * float(p.unit_weight or 0)
                brands.add(p.brand)

    return {
        "id": sh.id,
        "purchase_order_id": sh.purchase_order_id,
        "vehicle_id": sh.vehicle_id,
        "vehicle_name": f"{vehicle.name} ({vehicle.plate})" if vehicle else None,
        "status": sh.status,
        "status_label": SHIPMENT_STATUS_LABEL.get(sh.status, sh.status),
        "notes": sh.notes,
        "created_at": sh.created_at.isoformat() if sh.created_at else None,
        "line_count": len(lines),
        "total_loaded_box": total_loaded,
        "total_received_box": total_received,
        "total_weight": round(total_weight, 1),
        "brand_count": len(brands),
        "brands": sorted(brands),
    }


def _serialize_shipment_detail(sh: POShipment, db: Session) -> dict:
    """Shipment with full line details (for shipment detail page)."""
    base = _serialize_shipment(sh, db)
    lines = db.query(POShipmentLine).filter(POShipmentLine.shipment_id == sh.id).all()

    line_details = []
    for sl in lines:
        pl = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.id == sl.po_line_id).first()
        if not pl:
            continue
        p = db.query(Product).filter(Product.id == pl.product_id).first()
        line_details.append({
            "id": sl.id,
            "po_line_id": sl.po_line_id,
            "product_id": pl.product_id,
            "item_code": p.item_code if p else "",
            "name": p.name if p else "",
            "brand": p.brand if p else "",
            "warehouse_name": p.warehouse_name if p else "",
            "unit_weight": float(p.unit_weight or 0) if p else 0,
            "pack_ratio": float(p.pack_ratio or 1) if p else 1,
            "order_qty_box": pl.order_qty_box,
            "loaded_qty_box": sl.loaded_qty_box,
            "received_qty_box": sl.received_qty_box,
            "computed_weight": round(sl.loaded_qty_box * float(p.pack_ratio or 1) * float(p.unit_weight or 0), 2) if p else 0,
        })
    base["lines"] = line_details
    return base


# ── List shipments for a PO ──────────────────────────────────────────────────

@router.get("/{order_id}/shipments")
def list_shipments(
    order_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    shipments = (
        db.query(POShipment)
        .filter(POShipment.purchase_order_id == order_id)
        .order_by(POShipment.id)
        .all()
    )

    # Also compute unassigned quantities per PO line
    # unassigned = order_qty_box - SUM(loaded_qty_box across all shipments)
    all_ship_lines = (
        db.query(POShipmentLine)
        .join(POShipment, POShipment.id == POShipmentLine.shipment_id)
        .filter(POShipment.purchase_order_id == order_id)
        .all()
    )
    assigned_map: dict[int, float] = {}  # po_line_id → total loaded
    for sl in all_ship_lines:
        assigned_map[sl.po_line_id] = assigned_map.get(sl.po_line_id, 0) + sl.loaded_qty_box

    unassigned_lines = []
    for pl in po.lines:
        if pl.order_qty_box <= 0:
            continue
        assigned = assigned_map.get(pl.id, 0)
        remaining = pl.order_qty_box - assigned
        if remaining > 0:
            p = db.query(Product).filter(Product.id == pl.product_id).first()
            unassigned_lines.append({
                "po_line_id": pl.id,
                "product_id": pl.product_id,
                "item_code": p.item_code if p else "",
                "name": p.name if p else "",
                "brand": p.brand if p else "",
                "warehouse_name": p.warehouse_name if p else "",
                "unit_weight": float(p.unit_weight or 0) if p else 0,
                "pack_ratio": float(p.pack_ratio or 1) if p else 1,
                "order_qty_box": pl.order_qty_box,
                "assigned_qty_box": assigned,
                "remaining_qty_box": remaining,
            })

    return {
        "shipments": [_serialize_shipment(s, db) for s in shipments],
        "unassigned_lines": unassigned_lines,
    }


# ── Create shipment ──────────────────────────────────────────────────────────

class CreateShipmentIn(BaseModel):
    vehicle_id: Optional[int] = None
    notes: str = ""

@router.post("/{order_id}/shipments")
def create_shipment(
    order_id: int,
    body: CreateShipmentIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "supervisor", "manager")),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Захиалга олдсонгүй")
    if po.status != "loading":
        # Per-brand: ядаж 1 бренд "loading" статуст байвал зөвшөөрнө
        has_loading_brand = db.query(PurchaseOrderBrandStatus).filter(
            PurchaseOrderBrandStatus.purchase_order_id == order_id,
            PurchaseOrderBrandStatus.status == "loading",
        ).first()
        if not has_loading_brand:
            raise HTTPException(400, "Ачилт үүсгэхэд loading статустай бренд байхгүй")

    sh = POShipment(
        purchase_order_id=order_id,
        vehicle_id=body.vehicle_id,
        status="loading",
        notes=body.notes,
    )
    db.add(sh)
    db.commit()
    db.refresh(sh)
    return _serialize_shipment(sh, db)


# ── Get shipment detail ──────────────────────────────────────────────────────

@router.get("/{order_id}/shipments/{shipment_id}")
def get_shipment(
    order_id: int,
    shipment_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")
    return _serialize_shipment_detail(sh, db)


# ── Add/update lines to shipment (assign brands/products to truck) ────────────

class ShipmentLineIn(BaseModel):
    po_line_id: int
    loaded_qty_box: float = 0

@router.post("/{order_id}/shipments/{shipment_id}/lines")
def set_shipment_lines(
    order_id: int,
    shipment_id: int,
    lines: List[ShipmentLineIn],
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "supervisor", "manager")),
):
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")
    if sh.status != "loading":
        raise HTTPException(400, "Зөвхөн 'Ачигдаж байна' статуст бараа нэмж болно")

    existing_map = {sl.po_line_id: sl for sl in sh.lines}

    for li in lines:
        if li.loaded_qty_box <= 0:
            # Remove if exists
            if li.po_line_id in existing_map:
                db.delete(existing_map[li.po_line_id])
            continue

        if li.po_line_id in existing_map:
            existing_map[li.po_line_id].loaded_qty_box = li.loaded_qty_box
        else:
            db.add(POShipmentLine(
                shipment_id=shipment_id,
                po_line_id=li.po_line_id,
                loaded_qty_box=li.loaded_qty_box,
            ))
    db.commit()
    return _serialize_shipment_detail(sh, db)


# ── Assign entire brand to shipment (convenience) ────────────────────────────

class AssignBrandIn(BaseModel):
    brand: str

@router.post("/{order_id}/shipments/{shipment_id}/assign-brand")
def assign_brand_to_shipment(
    order_id: int,
    shipment_id: int,
    body: AssignBrandIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "supervisor", "manager")),
):
    """Брэндийн хуваарилагдаагүй бүх бараанууд → тухайн ачилтад нэмэгдэнэ."""
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")
    if sh.status != "loading":
        raise HTTPException(400, "Зөвхөн 'Ачигдаж байна' статуст бараа нэмж болно")

    # Brand status шалгалт
    brand_bs = db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id,
        PurchaseOrderBrandStatus.brand == body.brand,
    ).first()
    if brand_bs and brand_bs.status != "loading":
        raise HTTPException(400, f"'{body.brand}' бренд 'loading' статуст байх ёстой")

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()

    # All shipment lines for this PO (to compute assigned qty)
    all_ship_lines = (
        db.query(POShipmentLine)
        .join(POShipment, POShipment.id == POShipmentLine.shipment_id)
        .filter(POShipment.purchase_order_id == order_id)
        .all()
    )
    assigned_map: dict[int, float] = {}
    for sl in all_ship_lines:
        assigned_map[sl.po_line_id] = assigned_map.get(sl.po_line_id, 0) + sl.loaded_qty_box

    existing_in_shipment = {sl.po_line_id: sl for sl in sh.lines}
    added = 0

    for pl in po.lines:
        if pl.order_qty_box <= 0:
            continue
        p = db.query(Product).filter(Product.id == pl.product_id).first()
        if not p or p.brand != body.brand:
            continue
        remaining = pl.order_qty_box - assigned_map.get(pl.id, 0)
        if remaining <= 0:
            continue
        if pl.id in existing_in_shipment:
            existing_in_shipment[pl.id].loaded_qty_box += remaining
        else:
            db.add(POShipmentLine(
                shipment_id=shipment_id,
                po_line_id=pl.id,
                loaded_qty_box=remaining,
            ))
        assigned_map[pl.id] = assigned_map.get(pl.id, 0) + remaining
        added += 1

    db.commit()
    return {"ok": True, "added": added, "shipment": _serialize_shipment(sh, db)}


# ── Advance shipment status ──────────────────────────────────────────────────

@router.patch("/{order_id}/shipments/{shipment_id}/advance")
def advance_shipment(
    order_id: int,
    shipment_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")

    try:
        idx = SHIPMENT_STATUS_SEQ.index(sh.status)
        if idx + 1 >= len(SHIPMENT_STATUS_SEQ):
            raise HTTPException(400, "Ачилт эцсийн статуст хүрсэн")
        next_st = SHIPMENT_STATUS_SEQ[idx + 1]
    except ValueError:
        raise HTTPException(400, "Буруу статус")

    # Ачилтад бараа нэмэгдсэн эсэхийг шалгах (loading → transit)
    if sh.status == "loading" and next_st == "transit":
        if not sh.lines:
            raise HTTPException(400, "Ачилтад бараа нэмэгдээгүй байна")

    sh.status = next_st
    db.commit()

    # Sync brand statuses based on shipment brands
    # Find which brands are in this shipment
    sh_lines = db.query(POShipmentLine).filter(POShipmentLine.shipment_id == sh.id).all()
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if po and sh_lines:
        po_line_ids = [sl.po_line_id for sl in sh_lines]
        po_lines = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.id.in_(po_line_ids)).all()
        prod_ids = [pl.product_id for pl in po_lines]
        products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()}
        brands_in_shipment = {products[pl.product_id].brand for pl in po_lines if pl.product_id in products and products[pl.product_id].brand}

        # Map shipment status to brand status (shipment stages start from "loading")
        shipment_to_brand = {
            "loading": "loading", "transit": "transit", "arrived": "arrived",
            "accounting": "accounting", "confirmed": "confirmed", "received": "received",
        }
        target_brand_status = shipment_to_brand.get(next_st)
        if target_brand_status:
            for brand in brands_in_shipment:
                bs = db.query(PurchaseOrderBrandStatus).filter(
                    PurchaseOrderBrandStatus.purchase_order_id == order_id,
                    PurchaseOrderBrandStatus.brand == brand,
                ).first()
                if bs:
                    bs_idx = STATUS_SEQUENCE.index(bs.status) if bs.status in STATUS_SEQUENCE else 0
                    target_idx = STATUS_SEQUENCE.index(target_brand_status)
                    if target_idx > bs_idx:
                        bs.status = target_brand_status
            db.commit()

    # Auto-advance PO status from brands
    _sync_po_status_from_brands(order_id, db)
    # Fallback to legacy shipment sync if no brand statuses
    _sync_po_status(order_id, db)

    return _serialize_shipment(sh, db)


# ── Update received qty on shipment lines (arrived stage) ─────────────────────

class ShipmentReceivedIn(BaseModel):
    lines: List[dict]  # [{"shipment_line_id": int, "received_qty_box": float}]

@router.post("/{order_id}/shipments/{shipment_id}/received")
def set_shipment_received(
    order_id: int,
    shipment_id: int,
    body: ShipmentReceivedIn,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
):
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")
    if sh.status not in ("arrived", "accounting"):
        raise HTTPException(400, "Зөвхөн 'Ачаа ирсэн' статуст тоо оруулах боломжтой")

    line_map = {sl.id: sl for sl in sh.lines}
    for item in body.lines:
        sl = line_map.get(item.get("shipment_line_id"))
        if sl:
            sl.received_qty_box = float(item.get("received_qty_box", 0))
    db.commit()
    return _serialize_shipment_detail(sh, db)


# ── Delete shipment (only loading status) ─────────────────────────────────────

@router.delete("/{order_id}/shipments/{shipment_id}")
def delete_shipment(
    order_id: int,
    shipment_id: int,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "supervisor", "manager")),
):
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")
    if sh.status != "loading":
        raise HTTPException(400, "Зөвхөн 'Ачигдаж байна' статустай ачилтыг устгах боломжтой")
    db.delete(sh)
    db.commit()
    return {"ok": True}


# ── Update shipment vehicle ────────────────────────────────────────────────────

class UpdateShipmentIn(BaseModel):
    vehicle_id: Optional[int] = None
    notes: Optional[str] = None

@router.patch("/{order_id}/shipments/{shipment_id}")
def update_shipment(
    order_id: int,
    shipment_id: int,
    body: UpdateShipmentIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "supervisor", "manager")),
):
    sh = db.query(POShipment).filter(
        POShipment.id == shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not sh:
        raise HTTPException(404, "Ачилт олдсонгүй")
    if body.vehicle_id is not None:
        sh.vehicle_id = body.vehicle_id if body.vehicle_id else None
    if body.notes is not None:
        sh.notes = body.notes
    db.commit()
    return _serialize_shipment(sh, db)


# ── Move line(s) between shipments / back to unassigned ────────────────────────

class MoveLineIn(BaseModel):
    shipment_line_id: int
    target_shipment_id: Optional[int] = None   # None = буцааж хуваарилагдаагүй болгох

@router.post("/{order_id}/shipments/move-line")
def move_shipment_line(
    order_id: int,
    body: MoveLineIn,
    db: Session = Depends(get_db),
    u: User = Depends(require_role("admin", "supervisor", "manager")),
):
    """Нэг бараа (shipment_line)-г өөр ачилт руу шилжүүлэх эсвэл буцааж хуваарилагдаагүй болгох."""
    sl = db.query(POShipmentLine).filter(POShipmentLine.id == body.shipment_line_id).first()
    if not sl:
        raise HTTPException(404, "Ачилтын мөр олдсонгүй")

    src_sh = db.query(POShipment).filter(POShipment.id == sl.shipment_id).first()
    if not src_sh or src_sh.purchase_order_id != order_id:
        raise HTTPException(404, "Ачилт олдсонгүй")
    if src_sh.status != "loading":
        raise HTTPException(400, "Зөвхөн 'Ачигдаж байна' статуст шилжүүлэх боломжтой")

    if body.target_shipment_id is None:
        # Буцааж unassigned → shipment line устгана
        db.delete(sl)
        db.commit()
        return {"ok": True, "action": "unassigned"}

    # Өөр ачилт руу шилжүүлэх
    target_sh = db.query(POShipment).filter(
        POShipment.id == body.target_shipment_id,
        POShipment.purchase_order_id == order_id,
    ).first()
    if not target_sh:
        raise HTTPException(404, "Зорилтот ачилт олдсонгүй")
    if target_sh.status != "loading":
        raise HTTPException(400, "Зорилтот ачилт 'Ачигдаж байна' статуст байх ёстой")

    # Check if target already has this po_line
    existing = db.query(POShipmentLine).filter(
        POShipmentLine.shipment_id == body.target_shipment_id,
        POShipmentLine.po_line_id == sl.po_line_id,
    ).first()
    if existing:
        existing.loaded_qty_box += sl.loaded_qty_box
    else:
        db.add(POShipmentLine(
            shipment_id=body.target_shipment_id,
            po_line_id=sl.po_line_id,
            loaded_qty_box=sl.loaded_qty_box,
        ))
    db.delete(sl)
    db.commit()
    return {"ok": True, "action": "moved", "target_shipment_id": body.target_shipment_id}


# ── Brand status helpers ──────────────────────────────────────────────────────

def _ensure_brand_statuses(order_id: int, db: Session):
    """PO lines-аас бренд олж, brand_status record байхгүй бол үүсгэнэ."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        return

    lines = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.purchase_order_id == order_id).all()
    prod_ids = [l.product_id for l in lines]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()}

    # Brands with order_qty > 0 (override_brand-ийг түрүүнд харгалзана)
    active_brands: set[str] = set()
    for l in lines:
        if l.order_qty_box > 0:
            p = products.get(l.product_id)
            eff_brand = (l.override_brand or "").strip()
            if not eff_brand and p and p.brand and p.brand.lower() != "nan":
                eff_brand = p.brand
            if eff_brand and eff_brand.lower() != "nan":
                active_brands.add(eff_brand)

    existing = {
        bs.brand
        for bs in db.query(PurchaseOrderBrandStatus).filter(
            PurchaseOrderBrandStatus.purchase_order_id == order_id
        ).all()
    }

    for brand in active_brands - existing:
        db.add(PurchaseOrderBrandStatus(
            purchase_order_id=order_id,
            brand=brand,
            status=po.status,  # inherit current PO status
        ))

    if active_brands - existing:
        db.flush()


def _sync_po_status_from_brands(order_id: int, db: Session):
    """Бүх brand status-аас PO.status = min(brand statuses) тооцоолно."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        return

    brand_statuses = db.query(PurchaseOrderBrandStatus).filter(
        PurchaseOrderBrandStatus.purchase_order_id == order_id
    ).all()
    if not brand_statuses:
        return  # no brand records yet — don't touch PO status

    min_idx = min(
        STATUS_SEQUENCE.index(bs.status)
        for bs in brand_statuses
        if bs.status in STATUS_SEQUENCE
    )
    po.status = STATUS_SEQUENCE[min_idx]
    db.commit()


# ── Sync PO status from shipment statuses ─────────────────────────────────────

def _sync_po_status(order_id: int, db: Session):
    """Бүх shipment-ийн статусаас PO-ийн статусыг автоматаар тодорхойлно."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po or po.status not in ("loading", "transit", "arrived", "accounting", "confirmed", "received"):
        return

    shipments = db.query(POShipment).filter(POShipment.purchase_order_id == order_id).all()
    if not shipments:
        return

    statuses = [s.status for s in shipments]

    # Бүх shipment "received" → PO = "received"
    if all(s == "received" for s in statuses):
        po.status = "received"
    # Бүх shipment "confirmed" or higher → PO = "confirmed"
    elif all(SHIPMENT_STATUS_SEQ.index(s) >= SHIPMENT_STATUS_SEQ.index("confirmed") for s in statuses):
        po.status = "confirmed"
    # Бүх shipment "accounting" or higher → PO = "accounting"
    elif all(SHIPMENT_STATUS_SEQ.index(s) >= SHIPMENT_STATUS_SEQ.index("accounting") for s in statuses):
        po.status = "accounting"
    # Бүх shipment "arrived" or higher → PO = "arrived"
    elif all(SHIPMENT_STATUS_SEQ.index(s) >= SHIPMENT_STATUS_SEQ.index("arrived") for s in statuses):
        po.status = "arrived"
    # Ядаж 1 shipment "transit" or higher → PO = "transit" (хуваарилагдаагүй бараа байсан ч)
    elif any(SHIPMENT_STATUS_SEQ.index(s) >= SHIPMENT_STATUS_SEQ.index("transit") for s in statuses):
        po.status = "transit"

    db.commit()
