from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
from pathlib import Path
import io
import re

from app.api.deps import get_db, require_role

from app.models.inventory_count import InventoryCount, InventoryCountFile
from app.models.kpi import KpiAdminDailyTask, KpiChecklistEntry, KpiDailyChecklist

router = APIRouter(prefix="/inventory-count", tags=["inventory-count"])

UPLOAD_DIR = Path(__file__).resolve().parent.parent / "data" / "uploads" / "inventory_count"

# ── Constants ────────────────────────────────────────────────────────────────

WAREHOUSES = [
    {"key": "drink_alcohol",    "label": "Ус ундаа архи агуулах", "color": "text-blue-700",    "bg": "bg-blue-50 border-blue-200"},
    {"key": "retail",           "label": "Жижиглэн агуулах",      "color": "text-emerald-700", "bg": "bg-emerald-50 border-emerald-200"},
    {"key": "wholesale",        "label": "Бөөний агуулах",        "color": "text-violet-700",  "bg": "bg-violet-50 border-violet-200"},
    {"key": "contract",         "label": "Гэрээт агуулах",        "color": "text-amber-700",   "bg": "bg-amber-50 border-amber-200"},
    {"key": "alcohol_hall",     "label": "Архины заал",           "color": "text-rose-700",    "bg": "bg-rose-50 border-rose-200"},
    {"key": "hall",             "label": "Заал",                   "color": "text-sky-700",     "bg": "bg-sky-50 border-sky-200"},
    {"key": "kharkhorin_wh",    "label": "Хархорин агуулах",      "color": "text-orange-700",  "bg": "bg-orange-50 border-orange-200"},
    {"key": "kharkhorin_hall",  "label": "Хархорин заал",         "color": "text-teal-700",    "bg": "bg-teal-50 border-teal-200"},
]

WAREHOUSE_MAP = {w["key"]: w for w in WAREHOUSES}


# ── Schemas ──────────────────────────────────────────────────────────────────

class CountCreate(BaseModel):
    warehouse_key: str
    count_date: date
    description: str = ""
    target_employee_ids: List[int] = []
    points: float = 0.0


class CountUpdate(BaseModel):
    count_date: Optional[date] = None
    description: Optional[str] = None
    target_employee_ids: Optional[List[int]] = None
    points: Optional[float] = None


class ChecklistUpdate(BaseModel):
    check_all_synced: Optional[bool] = None
    check_no_partial: Optional[bool] = None
    check_no_wh14_sales: Optional[bool] = None
    check_balance_unchanged: Optional[bool] = None
    check_red_blocked_fixed: Optional[bool] = None


def _build_kpi_task_name(warehouse_label: str, description: str) -> str:
    desc = (description or "").strip()
    if desc:
        return f"{warehouse_label} — {desc}"
    return warehouse_label


def _get_or_create_checklist(db: Session, employee_id: int, on_date: date) -> KpiDailyChecklist:
    """Тухайн ажилтны тухайн өдрийн checklist row-ыг буцаана. Байхгүй бол
    нэг 'auto-created' төлөвт үүсгэнэ — ингэснээр ажилтан тэр өдөр аппыг
    нээгээгүй ч KPI entry нь бэлэн орших юм. Ажилтан дараа нь нэвтрэхэд
    мөнхүү checklist-д шууд нэмэгдсэн байх болно."""
    cl = db.query(KpiDailyChecklist).filter(
        KpiDailyChecklist.employee_id == employee_id,
        KpiDailyChecklist.date == on_date,
    ).first()
    if cl:
        return cl
    cl = KpiDailyChecklist(
        employee_id=employee_id,
        date=on_date,
        status="open",  # ажилтан өөрөө submit хийх боломжтой хэвээр
    )
    db.add(cl)
    db.flush()
    return cl


def _sync_inventory_kpi_entries(db: Session, kpi_task: KpiAdminDailyTask, default_points: float) -> None:
    """Inventory KPI admin task-тай холбоотой ажилтан бүрд entry-г eager-аар
    үүсгэдэг + хасагдсан ажилтнаас entry-г устгана.

    Дараах хатуу дүрэмтэй:
      - approved entry-г хэзээ ч устгадаггүй (audit trail хадгална)
      - тэдгээрийг "is_active" өөрчилдөггүй
      - шинээр insert хийх entry-ийн monetary_value = default_points (тооллогын
        default оноо — KpiSettings.inventory_default_points-аас ирнэ)
    """
    target_ids = [
        int(x) for x in (kpi_task.target_employee_ids or "").split(",") if x.strip()
    ]
    target_set = set(target_ids)

    # 1) Одоогийн entries-ийг авч ангилна (approved биш үлдэх вэ хадгална)
    existing = db.query(KpiChecklistEntry).filter(
        KpiChecklistEntry.admin_task_id == kpi_task.id
    ).all()
    by_emp: dict[int, KpiChecklistEntry] = {}
    for e in existing:
        # admin_task_id-ээр холбоотой entry бүрийн checklist-ийг tied хийнэ
        cl = db.query(KpiDailyChecklist).filter(
            KpiDailyChecklist.id == e.checklist_id
        ).first()
        if cl:
            by_emp[cl.employee_id] = e

    # 2) Target-аас хасагдсан ажилтнуудын entry-ийг (approved биш бол) устгана
    for emp_id, entry in list(by_emp.items()):
        if emp_id in target_set:
            continue
        if (entry.approval_status or "pending") == "approved":
            # approved entry-г устгахгүй, "соронзон" хадгална
            continue
        db.delete(entry)
    db.flush()

    # 3) Target-д шинэ ажилтан нэмэгдсэн бол entry үүсгэнэ (давхар үүсэх асуудал
    # үгүй — by_emp дотор аль хэдий нь байгаа эсэхийг шалгана)
    for emp_id in target_ids:
        if emp_id in by_emp:
            # Шинээр үүсэхгүй — байгаа entry хэвээр (approved ч, pending ч аль аль нь)
            continue
        cl = _get_or_create_checklist(db, emp_id, kpi_task.date)
        entry = KpiChecklistEntry(
            checklist_id=cl.id,
            template_id=None,
            config_id=None,
            task_name=kpi_task.task_name,
            monetary_value=default_points,
            task_category="inventory",
            is_checked=False,
            approval_status="pending",
            admin_task_id=kpi_task.id,
            approver_id=kpi_task.approver_id,
            is_adhoc=False,
        )
        db.add(entry)
    db.flush()


def _normalize_code(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".")[0]
    return s


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    s = str(value).strip().replace(",", "")
    if not s or s.lower() == "nan":
        return None
    try:
        return float(s)
    except Exception:
        return None


def _decode_text_file(path: Path) -> str:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1251", "cp1252"):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("utf-8", errors="ignore")


def _build_txt_code_index(txt_files: list[InventoryCountFile]) -> dict[str, list[tuple[str, int]]]:
    code_index: dict[str, list[tuple[str, int]]] = {}
    for f in txt_files:
        p = Path(f.saved_path or "")
        if not p.exists():
            continue
        text = _decode_text_file(p)
        person_name = (Path(f.original_filename or p.name).stem or p.stem).strip()
        for line_no, line in enumerate(text.splitlines(), start=1):
            raw = (line or "").strip()
            if not raw:
                continue
            m = re.search(r"\d{5,}", raw)
            if not m:
                continue
            code = _normalize_code(m.group(0))
            if not code:
                continue
            code_index.setdefault(code, []).append((person_name, line_no))
    return code_index


def _read_discrepancy_rows(excel_path: Path) -> list[dict]:
    import pandas as pd

    df = pd.read_excel(excel_path, header=None, dtype=object)
    if df.empty:
        return []

    header_row = -1
    scan_rows = min(len(df), 30)
    for i in range(scan_rows):
        c0 = str(df.iat[i, 0]).strip().lower() if df.shape[1] > 0 else ""
        c3 = str(df.iat[i, 3]).strip().lower() if df.shape[1] > 3 else ""
        c4 = str(df.iat[i, 4]).strip().lower() if df.shape[1] > 4 else ""
        if "код" in c0 and ("зөрүү" in c4 or "тоо" in c3):
            header_row = i
            break

    start_idx = header_row + 1 if header_row >= 0 else 0
    out: list[dict] = []
    for i in range(start_idx, len(df)):
        code = _normalize_code(df.iat[i, 0] if df.shape[1] > 0 else None)
        if not code:
            continue

        name_raw = df.iat[i, 1] if df.shape[1] > 1 else ""
        name = "" if name_raw is None else str(name_raw).strip()
        if name.lower() == "nan":
            name = ""

        balance = _to_float(df.iat[i, 2] if df.shape[1] > 2 else None)
        counted = _to_float(df.iat[i, 3] if df.shape[1] > 3 else None)
        diff = _to_float(df.iat[i, 4] if df.shape[1] > 4 else None)
        if diff is None:
            diff = (counted or 0.0) - (balance or 0.0)
        unit_price = _to_float(df.iat[i, 5] if df.shape[1] > 5 else None)

        out.append({
            "code": code,
            "name": name,
            "balance": balance,
            "counted": counted,
            "diff": diff,
            "unit_price": unit_price,
        })
    return out


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/warehouses")
def list_warehouses(_=Depends(require_role("admin", "supervisor", "manager"))):
    return WAREHOUSES


@router.get("/counts")
def list_counts(
    warehouse: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    from sqlalchemy import extract

    q = db.query(InventoryCount).order_by(InventoryCount.count_date.desc())
    if warehouse:
        q = q.filter(InventoryCount.warehouse_key == warehouse)
    if year:
        q = q.filter(extract("year", InventoryCount.count_date) == year)
    if month:
        q = q.filter(extract("month", InventoryCount.count_date) == month)

    rows = q.all()
    result = []
    for r in rows:
        wh = WAREHOUSE_MAP.get(r.warehouse_key, {})
        kpi_points = None
        kpi_target_ids: list[int] = []
        kpi_active = None
        if r.kpi_admin_task_id:
            kt = db.query(KpiAdminDailyTask).filter(
                KpiAdminDailyTask.id == r.kpi_admin_task_id
            ).first()
            if kt:
                kpi_points = kt.monetary_value
                kpi_target_ids = [
                    int(x) for x in (kt.target_employee_ids or "").split(",") if x.strip()
                ]
                kpi_active = kt.is_active
        result.append({
            "id": r.id,
            "warehouse_key": r.warehouse_key,
            "warehouse_label": wh.get("label", r.warehouse_key),
            "count_date": r.count_date.isoformat(),
            "description": r.description,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "created_by": r.created_by,
            "file_count": len(r.files),
            "files": [
                {
                    "id": f.id,
                    "file_type": f.file_type,
                    "original_filename": f.original_filename,
                    "uploaded_at": f.uploaded_at.isoformat() if f.uploaded_at else None,
                }
                for f in r.files
            ],
            "kpi_admin_task_id": r.kpi_admin_task_id,
            "kpi_points": kpi_points,
            "kpi_target_employee_ids": kpi_target_ids,
            "kpi_task_active": kpi_active,
            # Checklist
            "check_all_synced": bool(getattr(r, "check_all_synced", False)),
            "check_no_partial": bool(getattr(r, "check_no_partial", False)),
            "check_no_wh14_sales": bool(getattr(r, "check_no_wh14_sales", False)),
            "check_balance_unchanged": bool(getattr(r, "check_balance_unchanged", False)),
            "check_red_blocked_fixed": bool(getattr(r, "check_red_blocked_fixed", False)),
        })
    return result


@router.post("/counts")
def create_count(
    body: CountCreate,
    db: Session = Depends(get_db),
    u=Depends(require_role("admin", "supervisor", "manager")),
):
    if body.warehouse_key not in WAREHOUSE_MAP:
        raise HTTPException(400, "Агуулах олдсонгүй")
    wh_label = WAREHOUSE_MAP[body.warehouse_key]["label"]

    c = InventoryCount(
        warehouse_key=body.warehouse_key,
        count_date=body.count_date,
        description=body.description,
        created_by=u.username if hasattr(u, "username") else "",
    )
    db.add(c)
    db.flush()

    # KPI admin task үүсгэх (ажилтан сонгосон бол)
    kpi_task_id = None
    if body.target_employee_ids:
        from app.api.kpi import get_inventory_default_points
        default_points = get_inventory_default_points(db)
        task_name = _build_kpi_task_name(wh_label, body.description)
        target_ids_str = ",".join(str(x) for x in body.target_employee_ids)
        # monetary_value-ыг хэвээр хадгална (хуучин данс/audit-д ишлэл болгоно)
        # гэхдээ entry бүрд ашиглагдах оноо нь default_points
        kpi_task = KpiAdminDailyTask(
            task_name=task_name,
            monetary_value=body.points if body.points and body.points > 0 else default_points,
            task_category="inventory",
            date=body.count_date,
            approver_id=u.id,
            created_by=u.id,
            is_active=True,
            target_employee_ids=target_ids_str,
        )
        db.add(kpi_task)
        db.flush()
        c.kpi_admin_task_id = kpi_task.id
        kpi_task_id = kpi_task.id
        # Eager: оролцогч ажилтан бүрд KpiChecklistEntry-г шууд үүсгэнэ. Ажилтан тэр
        # өдөр аппыг нээгээгүй ч bulk approve-аар батлагдах боломжтой үлдэнэ.
        _sync_inventory_kpi_entries(db, kpi_task, default_points)

    db.commit()
    db.refresh(c)
    return {"id": c.id, "ok": True, "kpi_admin_task_id": kpi_task_id}


@router.put("/counts/{count_id}")
def update_count(
    count_id: int,
    body: CountUpdate,
    db: Session = Depends(get_db),
    u=Depends(require_role("admin")),
):
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")

    if body.count_date is not None:
        c.count_date = body.count_date
    if body.description is not None:
        c.description = body.description

    # Sync KPI admin task
    wh_label = WAREHOUSE_MAP.get(c.warehouse_key, {}).get("label", c.warehouse_key)
    kpi_task = None
    if c.kpi_admin_task_id:
        kpi_task = db.query(KpiAdminDailyTask).filter(
            KpiAdminDailyTask.id == c.kpi_admin_task_id
        ).first()

    from app.api.kpi import get_inventory_default_points
    default_points = get_inventory_default_points(db)

    if kpi_task:
        kpi_task.task_name = _build_kpi_task_name(wh_label, c.description)
        kpi_task.date = c.count_date
        if body.points is not None and body.points > 0:
            kpi_task.monetary_value = body.points
        if body.target_employee_ids is not None:
            kpi_task.target_employee_ids = ",".join(str(x) for x in body.target_employee_ids)
        # Sync entries: оролцогч нэмэгдвэл шинэ entry, хасагдвал устгана (approved-аас бусдыг)
        _sync_inventory_kpi_entries(db, kpi_task, default_points)
    else:
        # KPI task байхгүй — шинээр target/points орж ирсэн бол үүсгэнэ
        if body.target_employee_ids:
            new_task = KpiAdminDailyTask(
                task_name=_build_kpi_task_name(wh_label, c.description),
                monetary_value=(body.points if body.points and body.points > 0 else default_points),
                task_category="inventory",
                date=c.count_date,
                approver_id=u.id,
                created_by=u.id,
                is_active=True,
                target_employee_ids=",".join(str(x) for x in body.target_employee_ids),
            )
            db.add(new_task)
            db.flush()
            c.kpi_admin_task_id = new_task.id
            _sync_inventory_kpi_entries(db, new_task, default_points)

    db.commit()
    return {"ok": True}


@router.delete("/counts/{count_id}")
def delete_count(
    count_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")

    # Guard: батлагдсан KPI entry-тэй тооллого устгахыг хориглоно. Эс тэгвэл
    # ажилтны цалин тооцоонд дотоод зөрчил үүсэх магадлалтай (data integrity).
    kpi_task_id = c.kpi_admin_task_id
    if kpi_task_id:
        approved_count = db.query(KpiChecklistEntry).filter(
            KpiChecklistEntry.admin_task_id == kpi_task_id,
            KpiChecklistEntry.approval_status == "approved",
        ).count()
        if approved_count > 0:
            raise HTTPException(
                400,
                f"Энэ тооллогод {approved_count} ажилтны KPI аль хэдий нь батлагдсан "
                "тул устгах боломжгүй. Эхлээд тэр KPI-уудыг буцаах хэрэгтэй."
            )

    # Deactivate or delete linked KPI admin task
    if kpi_task_id:
        kpi_task = db.query(KpiAdminDailyTask).filter(
            KpiAdminDailyTask.id == kpi_task_id
        ).first()
        if kpi_task:
            has_refs = db.query(KpiChecklistEntry).filter(
                KpiChecklistEntry.admin_task_id == kpi_task_id
            ).first()
            if has_refs:
                kpi_task.is_active = False
                # Үлдсэн (pending/rejected) entries-ийг устгана — approved нь дээр шалгасан
                db.query(KpiChecklistEntry).filter(
                    KpiChecklistEntry.admin_task_id == kpi_task_id,
                    KpiChecklistEntry.approval_status != "approved",
                ).delete(synchronize_session=False)
            else:
                # Линкийг эхлээд салгаж дараа нь устгах (FK constraint-аас сэргийлж)
                c.kpi_admin_task_id = None
                db.flush()
                db.delete(kpi_task)

    # Delete files from disk
    for f in c.files:
        p = Path(f.saved_path)
        if p.exists():
            p.unlink()
    db.delete(c)
    db.commit()
    return {"ok": True}


@router.patch("/counts/{count_id}/checklist")
def update_checklist(
    count_id: int,
    body: ChecklistUpdate,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    """Тооллогын урьдчилсан шалгалтын checkbox-уудыг toggle хийх."""
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")
    if body.check_all_synced is not None:
        c.check_all_synced = bool(body.check_all_synced)
    if body.check_no_partial is not None:
        c.check_no_partial = bool(body.check_no_partial)
    if body.check_no_wh14_sales is not None:
        c.check_no_wh14_sales = bool(body.check_no_wh14_sales)
    if body.check_balance_unchanged is not None:
        c.check_balance_unchanged = bool(body.check_balance_unchanged)
    if body.check_red_blocked_fixed is not None:
        c.check_red_blocked_fixed = bool(body.check_red_blocked_fixed)
    db.commit()
    return {
        "ok": True,
        "check_all_synced": bool(c.check_all_synced),
        "check_no_partial": bool(c.check_no_partial),
        "check_no_wh14_sales": bool(c.check_no_wh14_sales),
        "check_balance_unchanged": bool(c.check_balance_unchanged),
        "check_red_blocked_fixed": bool(c.check_red_blocked_fixed),
    }


@router.delete("/files/{file_id}")
def delete_file(
    file_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Зөвхөн админ — тооллогын нэмэлт файл (TXT/Excel)-ыг устгах."""
    f = db.query(InventoryCountFile).filter(InventoryCountFile.id == file_id).first()
    if not f:
        raise HTTPException(404, "Файл олдсонгүй")
    p = Path(f.saved_path)
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass  # Disk-н алдаа хэвийн зүйл биш ч DB record-ыг үргэлжлүүлэн устгана
    db.delete(f)
    db.commit()
    return {"ok": True}


@router.post("/counts/{count_id}/upload-txt")
async def upload_txt_files(
    count_id: int,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")

    target_dir = UPLOAD_DIR / c.warehouse_key / c.count_date.isoformat()
    target_dir.mkdir(parents=True, exist_ok=True)

    saved = []
    for f in files:
        suffix = Path(f.filename or "file.txt").suffix.lower()
        if suffix not in (".txt", ".text"):
            continue
        ts = datetime.now().strftime("%H%M%S")
        safe_name = f"{ts}_{f.filename}"
        dest = target_dir / safe_name
        dest.write_bytes(await f.read())

        rec = InventoryCountFile(
            inventory_count_id=c.id,
            file_type="txt",
            original_filename=f.filename or "",
            saved_path=str(dest),
        )
        db.add(rec)
        saved.append(f.filename)

    db.commit()
    return {"ok": True, "saved": saved, "count": len(saved)}


@router.post("/counts/{count_id}/upload-excel")
async def upload_excel_file(
    count_id: int,
    f: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")

    suffix = Path(f.filename or "file.xlsx").suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(400, "Excel файл оруулна уу (.xlsx, .xls)")

    target_dir = UPLOAD_DIR / c.warehouse_key / c.count_date.isoformat()
    target_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%H%M%S")
    safe_name = f"{ts}_{f.filename}"
    dest = target_dir / safe_name
    dest.write_bytes(await f.read())

    rec = InventoryCountFile(
        inventory_count_id=c.id,
        file_type="excel",
        original_filename=f.filename or "",
        saved_path=str(dest),
    )
    db.add(rec)
    db.commit()
    return {"ok": True, "filename": f.filename}


@router.get("/counts/{count_id}/files")
def list_files(
    count_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")
    return [
        {
            "id": f.id,
            "file_type": f.file_type,
            "original_filename": f.original_filename,
            "uploaded_at": f.uploaded_at.isoformat() if f.uploaded_at else None,
        }
        for f in c.files
    ]


@router.get("/files/{file_id}/download")
def download_file(
    file_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    f = db.query(InventoryCountFile).filter(InventoryCountFile.id == file_id).first()
    if not f:
        raise HTTPException(404, "Файл олдсонгүй")
    p = Path(f.saved_path)
    if not p.exists():
        raise HTTPException(404, "Файл дискнээс олдсонгүй")
    return FileResponse(
        path=str(p),
        filename=f.original_filename,
        media_type="application/octet-stream",
    )


@router.get("/counts/{count_id}/export-discrepancy")
def export_discrepancy_excel(
    count_id: int,
    excel_file_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    c = db.query(InventoryCount).filter(InventoryCount.id == count_id).first()
    if not c:
        raise HTTPException(404, "Тооллого олдсонгүй")

    excel_q = db.query(InventoryCountFile).filter(
        InventoryCountFile.inventory_count_id == count_id,
        InventoryCountFile.file_type == "excel",
    )
    if excel_file_id is not None:
        excel_file = excel_q.filter(InventoryCountFile.id == excel_file_id).first()
    else:
        excel_file = excel_q.order_by(InventoryCountFile.uploaded_at.desc(), InventoryCountFile.id.desc()).first()
    if not excel_file:
        raise HTTPException(400, "Тооллогоны Эксэл файл (.xlsx) оруулаагүй байна")

    excel_path = Path(excel_file.saved_path or "")
    if not excel_path.exists():
        raise HTTPException(404, "Тооллогоны Эксэл файл дискнээс олдсонгүй")

    txt_files = db.query(InventoryCountFile).filter(
        InventoryCountFile.inventory_count_id == count_id,
        InventoryCountFile.file_type == "txt",
    ).all()
    txt_index = _build_txt_code_index(txt_files)

    source_rows = _read_discrepancy_rows(excel_path)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    wb = Workbook()
    ws = wb.active
    ws.title = "Тооллогоны зөрүү"
    headers = [
        "Код",
        "Нэр",
        "Үлдэгдэл",
        "Тоолсон тоо",
        "Зөрүү",
        "Нэгж үнэ",
        "TXT эх сурвалж (Нэр + мөр)",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    out_count = 0
    for r in source_rows:
        diff = float(r.get("diff") or 0.0)
        if abs(diff) < 1e-9:
            continue

        code = r.get("code") or ""
        refs = txt_index.get(code, [])
        seen = set()
        ref_parts: list[str] = []
        for person, line_no in refs:
            key = (person, line_no)
            if key in seen:
                continue
            seen.add(key)
            ref_parts.append(f"{person} (мөр {line_no})")
        source_info = "; ".join(ref_parts) if ref_parts else "Олдсонгүй"

        ws.append([
            code,
            r.get("name") or "",
            r.get("balance"),
            r.get("counted"),
            diff,
            r.get("unit_price"),
            source_info,
        ])
        out_count += 1

    widths = [16, 40, 14, 14, 12, 14, 48]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=6, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.##"
    for cell in ws["G"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = c.count_date.isoformat()
    filename = f"{date_str}_count_{count_id}_zoruutai_baraa_{out_count}.xlsx"
    ascii_fallback = re.sub(r"[^\w\-.]", "_", filename.encode("ascii", "ignore").decode("ascii")) or f"count_{count_id}_diff.xlsx"
    utf8_quoted = quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={ascii_fallback}; filename*=UTF-8''{utf8_quoted}"},
    )


@router.get("/calendar")
def calendar_data(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin", "supervisor", "manager")),
):
    from sqlalchemy import extract

    rows = (
        db.query(InventoryCount)
        .filter(
            extract("year", InventoryCount.count_date) == year,
            extract("month", InventoryCount.count_date) == month,
        )
        .order_by(InventoryCount.count_date)
        .all()
    )

    # Group by day
    day_map: dict[int, list] = {}
    for r in rows:
        day = r.count_date.day
        wh = WAREHOUSE_MAP.get(r.warehouse_key, {})
        entry = {
            "id": r.id,
            "warehouse_key": r.warehouse_key,
            "warehouse_label": wh.get("label", r.warehouse_key),
            "color": wh.get("color", "text-gray-700"),
            "bg": wh.get("bg", "bg-gray-50"),
            "description": r.description,
            "file_count": len(r.files),
        }
        day_map.setdefault(day, []).append(entry)

    return {"year": year, "month": month, "days": day_map}
